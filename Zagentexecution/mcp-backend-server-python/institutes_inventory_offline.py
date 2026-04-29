"""
Build per-institute inventory tabs OFFLINE from Session #031 captures:
- MGIE: phase2_e071_MGIE.json + phase2_e071k_MGIE.json
- ICBA: icba_e071_full.json + icba_e071k_full.json
- ICTP: ictp_e071k_full.json (E071K only — derive object list from there)
"""
import sys
import json
from collections import Counter, defaultdict
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

base = 'Zagentexecution/mcp-backend-server-python/blueprint_output/'

# Reuse explanation + golden mapping from STEM script
GOLDEN_MAP = {
    'T001':[(1,'Create / Copy Company Code')], 'T001Q':[(1,'Create / Copy Company Code')],
    'T001A':[(2,'Add Group Currency')], 'T001B':[(69,'Open posting periods')],
    'T010O':[(69,'Open posting periods')], 'T010P':[(69,'Open posting periods')],
    'SKB1':[(4,'Extend GL per co code')], 'SKA1':[(4,'GL master')],
    'T012':[(16,'Create house banks')], 'T012D':[(16,'Create house banks')],
    'T012K':[(18,'Bank accounts')], 'T012T':[(16,'Create house banks')],
    'T018V':[(17,'Bank directory entries')], 'T028':[(17,'Bank directory')],'T028B':[(17,'Bank directory')],
    'T030H':[(14,'OBA1 / Exchange rate diff')],
    'T035D':[(28,'Cash management planning groups')],'T035U':[(28,'Cash management planning groups')],
    'T042':[(20,'FBZP — All co codes')],'T042B':[(23,'Payment methods per co code')],'T042E':[(22,'Payment methods per country')],
    'T042I':[(24,'Bank determination/ranking')],'T042T':[(21,'FBZP Paying co codes')],'T042V':[(25,'FBZP extras')],
    'T042A':[(25,'FBZP extras')],'T042D':[(25,'FBZP extras')],'T042P':[(25,'FBZP extras')],
    'T043G':[(35,'Tolerance groups')],'T043GT':[(35,'Tolerance groups')],
    'T043S':[(35,'Tolerance groups')],'T043ST':[(35,'Tolerance groups')],'T043T':[(35,'Tolerance groups')],
    'T169G':[(37,'MM/IV tolerances')],'T169P':[(37,'MM/IV tolerances')],'T169V':[(37,'MM/IV tolerances')],
    'T093B':[(44,'Asset class → GL determination')],'T093C':[(43,'Chart of depreciation assignment')],
    'T093D':[(46,'Asset transfer date')],'T093U':[(45,'Depreciation keys')],'T093V':[(43,'Chart of depreciation')],
    'AAACC_OBJ':[(43,'Account assignment objects')], 'ATPRA':[(46,'Asset transfer')], 'MARV':[(46,'MM/AA period')],
    'T047':[(42,'Dunning procedure')],
    'T882':[(55,'Consolidation unit')],
    'GB90':[(29,'GGB1 substitution')],'GB901':[(29,'GGB1 substitution')],'GB903':[(29,'GGB1 substitution')],
    'GB90T':[(29,'GGB1 substitution')],'GB92':[(29,'GGB1 substitution')],'GB921':[(29,'GGB1 substitution')],
    'GB922':[(29,'GGB1 substitution')],'GB92T':[(29,'GGB1 substitution')],'GB921T':[(29,'GGB1 substitution')],
    'GB93':[(31,'GGB0 validation')],
    'GGB1':[(29,'GGB1 substitution bundle')], 'GGB0':[(31,'GGB0 validation bundle')],
    'TKA00':[(9,'CO area')],'TKA01':[(9,'CO area')],'TKA02':[(9,'CO area assignment')],
    'TKA07':[(9,'CO area')],'TKA09':[(9,'CO area')],'TKT09':[(9,'CO area text')],
    'V_TKA3':[(10,'CO versions')],
    'CSKA':[(11,'Cost elements chart')], 'CSKB':[(11,'Cost elements per CO area')],
    'CSKU':[(11,'Cost elements text')], 'CSKS':[(47,'Cost centers')], 'CSKT':[(47,'Cost centers text')],
    'FM01':[(5,'FM area')], 'FMCI':[(8,'Commitment items')], 'FMCIT':[(8,'Commitment items text')],
    'FMFCTR':[(50,'Fund centers')], 'FMFCTRT':[(50,'Fund centers text')],
    'FMFINT':[(51,'Funds')],
    'FMFUNDTYPE':[(6,'Fund types')], 'V_FMFUNDTYPE':[(6,'Fund types maintenance')],
    'TPIR2':[(59,'Travel integration')], 'TPIR2T':[(59,'Travel integration text')],
    'T542A':[(58,'Personnel area subarea')], 'T500P':[(58,'Personnel area')],
    'T024':[(61,'Purchasing org')], 'T024E':[(61,'Purchasing org')], 'T024W':[(60,'Plant assignment')],
    'T001W':[(60,'Plant')], 'T001K':[(60,'Valuation area')],
    'T077K':[(0,'Vendor account groups')],
    'TCJ_C_JOURNALS':[(56,'Cash journal setup')],'TCJ_CJ_NAMES':[(56,'Cash journal setup')],
    'TCJ_PRINT':[(56,'Cash journal setup')],'TCJ_TRANSACTIONS':[(56,'Cash journal setup')],
    'TCJ_TRANS_NAMES':[(56,'Cash journal setup')],
    'FMCFC0':[(54,'FM Carry-forward')],'FMCFC1':[(54,'FM Carry-forward')],'FMCFC2':[(54,'FM Carry-forward')],
    'T007V':[(38,'Tax codes')],'TAXIN':[(38,'Tax codes (India)')],
}


def explain(obj, name):
    name_u = (name or '').upper().strip()
    if obj == 'TABU':
        if name_u in ('T001','T001Q','T001A'): return 'Company code core record'
        if name_u in ('T012','T012D','T012K','T012T'): return 'House bank / bank account'
        if name_u in ('T018V','T018C','T018D','T028','T028B'): return 'Bank directory + bank type'
        if name_u.startswith('T042'): return 'FBZP payment program config'
        if name_u.startswith('T043') or name_u.startswith('T169'): return 'Tolerance group config'
        if name_u.startswith('T093') or name_u in ('AAACC_OBJ','ATPRA'): return 'Asset accounting config'
        if name_u == 'SKB1': return 'GL accounts at company-code level'
        if name_u in ('SKA1','SKAT'): return 'GL master / GL text'
        if name_u in ('T035D','T035U'): return 'Cash management planning groups'
        if name_u == 'T047': return 'Dunning procedure'
        if name_u == 'T882': return 'Consolidation unit'
        if name_u.startswith('GB'): return 'Validation/Substitution rule'
        if name_u.startswith('TKA'): return 'Controlling area config'
        if name_u in ('CSKA','CSKB','CSKS','CSKT','CSKU'): return 'CO master data'
        if name_u.startswith('FM') or 'FUND' in name_u: return 'Funds Management config'
        if name_u.startswith('TPIR') or name_u in ('T500P','T542A'): return 'HR / Travel personnel config'
        if name_u.startswith('TCJ'): return 'Cash journal config'
        if name_u in ('T010O','T010P','T001B'): return 'Posting periods'
        if name_u.startswith('T024'): return 'Purchasing org config'
        if name_u.startswith('T001W') or name_u == 'T001K': return 'Plant / valuation area'
        if name_u in ('UNESGBB','UNESWRX'): return 'UNESCO custom MM account determination'
        if name_u.startswith('T007') or 'TAXIN' in name_u: return 'Tax code config'
        if name_u.startswith('T030'): return 'GL automatic account determination'
        return f'Customizing table {name}'
    if obj == 'TDAT':
        if name_u == 'GGB1': return 'GGB1 substitution bundle'
        if name_u == 'GGB0': return 'GGB0 validation bundle'
        if 'ADDRESS' in name_u: return 'Address tables bundle'
        if name_u == 'GRW_SET': return 'GS02 sets bundle'
        if name_u.startswith('KE34'): return 'CO-PA operating concern'
        return f'Transport data {name}'
    if obj == 'VDAT': return f'View data {name}'
    if obj == 'CDAT': return f'Customer data {name}'
    if obj == 'F01U': return f'Cross-company clearing pair {name}'
    if obj == 'CORR': return 'Sub-task release'
    if obj == 'RELE': return 'Released sub-task'
    if obj == 'ACGR': return f'Authorization role {name}'
    if obj.startswith('CL'): return f'ABAP class {name}'
    if obj == 'PROG': return f'ABAP program {name}'
    return f'{obj} {name}'


def load_json(path):
    try:
        with open(base + path) as f: return json.load(f)
    except FileNotFoundError:
        return None


def build_inventory(institute, e071_data, e071k_data):
    """Build inventory rows from E071/E071K data."""
    rows = []
    # Group by transport
    by_tr_objs = defaultdict(list)
    by_tr_keys = defaultdict(lambda: defaultdict(int))

    if e071_data:
        for r in e071_data:
            tr = r.get('TRKORR','').strip()
            by_tr_objs[tr].append({
                'PGMID': r.get('PGMID','').strip(),
                'OBJECT': r.get('OBJECT','').strip(),
                'OBJ_NAME': r.get('OBJ_NAME','').strip(),
            })
    if e071k_data:
        for r in e071k_data:
            tr = r.get('TRKORR','').strip()
            obj = r.get('OBJECT','').strip()
            name = r.get('OBJNAME','').strip()
            by_tr_keys[tr][(obj, name)] += 1

    # If no E071 data (ICTP case), derive object list from E071K
    for tr in by_tr_keys:
        if tr not in by_tr_objs:
            for (obj, name) in by_tr_keys[tr]:
                by_tr_objs[tr].append({
                    'PGMID': 'R3TR', 'OBJECT': obj, 'OBJ_NAME': name,
                })

    # Build rows
    for tr in sorted(by_tr_objs):
        for o in by_tr_objs[tr]:
            cnt = by_tr_keys[tr].get((o['OBJECT'], o['OBJ_NAME']), 0)
            plain = explain(o['OBJECT'], o['OBJ_NAME'])
            gold = GOLDEN_MAP.get(o['OBJ_NAME'].upper(), [])
            gold_str = '; '.join(f'#{g[0]} {g[1]}' for g in gold) if gold else ''
            rows.append({
                'transport': tr, 'pgmid': o['PGMID'], 'object': o['OBJECT'],
                'obj_name': o['OBJ_NAME'], 'keys': cnt,
                'explanation': plain, 'maps_to_golden': gold_str,
            })
    return rows


# Load and build for each institute
inventories = {}

mgie_e = load_json('phase2_e071_MGIE.json')
mgie_k = load_json('phase2_e071k_MGIE.json')
inventories['MGIE'] = build_inventory('MGIE', mgie_e, mgie_k)
print(f'MGIE: {len(inventories["MGIE"])} inventory rows from {len(set(r["transport"] for r in inventories["MGIE"]))} transports')

icba_e = load_json('icba_e071_full.json')
icba_k = load_json('icba_e071k_full.json')
inventories['ICBA'] = build_inventory('ICBA', icba_e, icba_k)
print(f'ICBA: {len(inventories["ICBA"])} inventory rows from {len(set(r["transport"] for r in inventories["ICBA"]))} transports')

ictp_k = load_json('ictp_e071k_full.json')
inventories['ICTP'] = build_inventory('ICTP', None, ictp_k)
print(f'ICTP: {len(inventories["ICTP"])} inventory rows from {len(set(r["transport"] for r in inventories["ICTP"]))} transports (E071K-only)')


# Add tabs to Excel
SRC = 'Zagentexecution/Golden_live_matrix_2026-04-16_realigned_v3.xlsx'
DST = 'Zagentexecution/Golden_live_matrix_2026-04-16_realigned_v4.xlsx'
wb = load_workbook(SRC)
hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
hdr_font = Font(color='FFFFFF', bold=True)
tr_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')

for inst, inv in inventories.items():
    sheet_name = f'{inst} Inventory'
    if sheet_name in wb.sheetnames: del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(['Transport','PGMID','OBJECT','OBJ_NAME','Keys','Explanation','Maps to Golden Step'])
    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    cur_tr = None
    for r in inv:
        if r['transport'] != cur_tr:
            ws.append([f"== {r['transport']} =="] + ['']*6)
            for c in ws[ws.max_row]:
                c.fill = tr_fill; c.font = Font(bold=True)
            cur_tr = r['transport']
        ws.append([r['transport'], r['pgmid'], r['object'], r['obj_name'],
                   r['keys'], r['explanation'], r['maps_to_golden']])
    widths = [16, 6, 7, 22, 7, 50, 30]
    for col_letter, w in zip('ABCDEFG', widths):
        ws.column_dimensions[col_letter].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:G{ws.max_row}'
    print(f'  Tab "{sheet_name}": {ws.max_row - 1} rows')

# Migrated institutes tab
sheet = 'Migrated Institutes'
if sheet in wb.sheetnames: del wb[sheet]
ws = wb.create_sheet(sheet)
ws.append(['Institute','Status','Note'])
for c in ws[1]:
    c.fill = hdr_fill; c.font = hdr_font
for inst in ['IBE','UBO','UIL','UIS']:
    ws.append([inst,'Pre-D01 / Migrated','No T001 creation transport found in D01 E071K. Config exists in live tables but the creation trail is in a predecessor SAP system that was migrated.'])
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 100

wb.save(DST)
print(f'\nSaved: {DST}')
print(f'Tabs: {wb.sheetnames}')
