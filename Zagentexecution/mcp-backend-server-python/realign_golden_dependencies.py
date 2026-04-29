"""
Read user-modified Golden_live_matrix_2026-04-16.xlsx, renumber items
sequentially in the new order, and update all dep references.

Strategy:
1. Read all data rows (skip group headers)
2. Build map old_item → new_item (1..N in current row order)
3. Translate Dep column ("19,42" → new_item lookups)
4. Normalize phase header typos (P19 Activation, P20 Testing → consistent)
5. Save with original styling preserved
"""
import sys
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from copy import copy

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

SRC = 'Zagentexecution/Golden_live_matrix_2026-04-16.xlsx'
DST = 'Zagentexecution/Golden_live_matrix_2026-04-16_realigned.xlsx'

wb = load_workbook(SRC)
ws = wb['Golden × Live']

# Read all data rows
rows = []
for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    rows.append({'row_idx': r_idx, 'values': list(row)})

# First pass: identify group headers (item is None) vs data rows
data_rows = [r for r in rows if r['values'][1] is not None]
header_rows = [r for r in rows if r['values'][1] is None]
print(f'Data rows: {len(data_rows)}  Group headers: {len(header_rows)}')

# Build old → new item mapping (preserve order)
old_to_new = {}
for new_item, r in enumerate(data_rows, start=1):
    old_item = r['values'][1]
    old_to_new[old_item] = new_item
    r['new_item'] = new_item

# Special tokens (don't translate)
SPECIAL = {'all', '', None}

def translate_dep(dep):
    """Translate '19,20' → new numbers."""
    if dep is None or dep == '' or str(dep).strip().lower() == 'all':
        return dep
    # Handle ranges like "52-56" — expand to explicit list since reordering breaks contiguity
    if '-' in str(dep) and str(dep).count('-') == 1:
        try:
            a, b = str(dep).split('-')
            a = a.strip(); b = b.strip()
            if a.isdigit() and b.isdigit():
                old_range = list(range(int(a), int(b) + 1))
                new_items = [old_to_new[o] for o in old_range if o in old_to_new]
                new_items.sort()
                return ','.join(str(n) for n in new_items)
        except: pass
        return dep
    parts = re.split(r'[,;\s]+', str(dep).strip())
    out = []
    for p in parts:
        if not p or p in SPECIAL: continue
        if p.isdigit():
            old = int(p)
            new = old_to_new.get(old)
            if new:
                out.append(str(new))
            else:
                out.append(f'?{p}')  # marker for missing
        else:
            out.append(p)  # keep textual labels
    return ','.join(out)

# Apply changes
phase_normalize = {
    'P19 Activation': 'P19 Activation',
    'P20  Activation': 'P19 Activation',
    'P20 Activation': 'P19 Activation',
    'P19 Testing': 'P20 Testing',
    'P20 Testing': 'P20 Testing',
}

print('\n=== Renumbering + dep translation ===')
print(f'{"Old":>4s} → {"New":>4s}  {"Phase":20s} {"Action":40s} {"Old Dep":>10s} → {"New Dep":>10s}')
for r in data_rows:
    v = r['values']
    phase, item, system, action, tcode, typ, dep, tables = v[:8]
    new_item = r['new_item']
    new_dep = translate_dep(dep)
    new_phase = phase_normalize.get(phase, phase)
    print(f'{item:>4} → {new_item:>4}  {new_phase[:20]:20s} {(action or "")[:40]:40s} {str(dep or ""):>10s} → {str(new_dep or ""):>10s}')
    # write back
    ws.cell(row=r['row_idx'], column=1).value = new_phase
    ws.cell(row=r['row_idx'], column=2).value = new_item
    ws.cell(row=r['row_idx'], column=7).value = new_dep

# Also fix group header phase names
for r in header_rows:
    phase = r['values'][0]
    new_phase = phase_normalize.get(phase, phase)
    if new_phase != phase:
        ws.cell(row=r['row_idx'], column=1).value = new_phase

wb.save(DST)
print(f'\nSaved: {DST}')
print(f'Items renumbered: {len(data_rows)}')
print(f'Phase headers normalized: {sum(1 for r in header_rows if phase_normalize.get(r["values"][0]) != r["values"][0])}')
