"""
Pull E071/E071K for every CURRENT STEM transport and write a STEM Transport
Inventory tab to the Excel with explanations + Golden step mapping.
"""
import sys
import json
from collections import Counter, defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

sys.path.insert(0, 'Zagentexecution/mcp-backend-server-python')
from rfc_helpers import get_connection
g = get_connection('D01')

# Current STEM creation cluster (2024-11-14 → 2026-04-15)
STEM_TRANSPORTS = [
    ('D01K9B0CDR','2024-11-14','A_SEFIANI','W','HR - PA International Institute for STEM Shanghai'),
    ('D01K9B0CDS','2024-11-14','A_SEFIANI','Q','HR - PA task (D01K9B0CDR)'),
    ('D01K9B0CBF','2026-03-26','JP_LOPEZ','W','FI New company code IISTEM'),
    ('D01K9B0F40','2026-03-27','R_RIOS','W','Logistic STEM Company Code'),
    ('D01K9B0F3V','-','I_KONAKOV','W','FM Area + config (parent of F3W)'),
    ('D01K9B0F3W','2026-04-14','I_KONAKOV','Q','FM task (D01K9B0F3V)'),
    ('D01K9B0F41','2026-04-14','R_RIOS','Q','Logistic task (D01K9B0F40)'),
    ('D01K9B0F4I','2026-04-14','R_RIOS','Q','FI periods (D01K9B0CBF)'),
    ('D01K9B0CBG','2026-04-15','JP_LOPEZ','Q','FI task (D01K9B0CBF) — our SKB1+CSKB work'),
    ('D01K9B0F7I','2026-04-15','JP_LOPEZ','K','Workbench: NEW Company Code STEM'),
    ('D01K9B0F7J','2026-04-15','JP_LOPEZ','S','Workbench task (D01K9B0F7I)'),
]

# Golden step mapping: object/table → list of (step_id, step_name)
GOLDEN_MAP = {
    'T001':[(1,'Create / Copy Company Code')], 'T001Q':[(1,'Create / Copy Company Code')],
    'T001A':[(2,'Add Group Currency')], 'T001B':[(69,'Open posting periods')],
    'T010O':[(69,'Open posting periods')], 'T010P':[(69,'Open posting periods')],
    'SKB1':[(4,'Extend GL per co code')], 'SKA1':[(4,'GL master')],
    'T012':[(16,'Create house banks')], 'T012D':[(16,'Create house banks')],
    'T012K':[(18,'Bank accounts')], 'T012T':[(16,'Create house banks')],
    'T018V':[(17,'Bank directory entries')], 'T028':[(17,'Bank directory')],'T028B':[(17,'Bank directory')],
    'T030H':[(14,'OBA1 / Exchange rate diff')],
    'T035D':[(28,'Cash management planning groups')],'T035U':[(28,'Cash management planning groups')],
    'T042':[(20,'FBZP — All co codes')],'T042B':[(23,'Payment methods per co code')],'T042E':[(22,'Payment methods per country')],
    'T042I':[(24,'Bank determination/ranking')],'T042T':[(21,'FBZP Paying co codes')],'T042V':[(25,'FBZP extras')],
    'T042A':[(25,'FBZP extras')],'T042D':[(25,'FBZP extras')],'T042P':[(25,'FBZP extras')],
    'T043G':[(35,'Tolerance groups')],'T043GT':[(35,'Tolerance groups')],
    'T043S':[(35,'Tolerance groups')],'T043ST':[(35,'Tolerance groups')],'T043T':[(35,'Tolerance groups')],
    'T169G':[(37,'MM/IV tolerances')],'T169P':[(37,'MM/IV tolerances')],'T169V':[(37,'MM/IV tolerances')],
    'T093B':[(44,'Asset class → GL determination')],'T093C':[(43,'Chart of depreciation assignment')],
    'T093D':[(46,'Asset transfer date')],'T093U':[(45,'Depreciation keys')],'T093V':[(43,'Chart of depreciation')],
    'AAACC_OBJ':[(43,'Account assignment objects')], 'ATPRA':[(46,'Asset transfer')], 'MARV':[(46,'MM/AA period')],
    'T047':[(42,'Dunning procedure')],
    'T882':[(55,'Consolidation unit')],
    'GB90':[(29,'GGB1 substitution')],'GB901':[(29,'GGB1 substitution')],'GB903':[(29,'GGB1 substitution')],
    'GB90T':[(29,'GGB1 substitution')],'GB92':[(29,'GGB1 substitution')],'GB921':[(29,'GGB1 substitution')],
    'GB922':[(29,'GGB1 substitution')],'GB92T':[(29,'GGB1 substitution')],'GB921T':[(29,'GGB1 substitution')],
    'GB93':[(31,'GGB0 validation')],
    'GGB1':[(29,'GGB1 substitution bundle')], 'GGB0':[(31,'GGB0 validation bundle')],
    'TKA00':[(9,'CO area')],'TKA01':[(9,'CO area')],'TKA02':[(9,'CO area assignment')],
    'TKA07':[(9,'CO area')],'TKA09':[(9,'CO area')],'TKT09':[(9,'CO area text')],
    'V_TKA3':[(10,'CO versions')],
    'CSKA':[(11,'Cost elements chart')], 'CSKB':[(11,'Cost elements per CO area')], 'CSKU':[(11,'Cost elements text')],
    'CSKS':[(47,'Cost centers')], 'CSKT':[(47,'Cost centers text')],
    'FM01':[(5,'FM area')], 'FMCI':[(8,'Commitment items')], 'FMCIT':[(8,'Commitment items text')],
    'FMFCTR':[(50,'Fund centers')], 'FMFCTRT':[(50,'Fund centers text')],
    'FMFINT':[(51,'Funds')],
    'FMFUNDTYPE':[(6,'Fund types')], 'V_FMFUNDTYPE':[(6,'Fund types maintenance')],
    'TPIR2':[(59,'Travel integration')], 'TPIR2T':[(59,'Travel integration text')],
    'T542A':[(58,'Personnel area subarea')], 'T500P':[(58,'Personnel area')],
    'T527O':[(80,'HCM customizing')],
    'T024':[(61,'Purchasing org')], 'T024E':[(61,'Purchasing org')], 'T024W':[(60,'Plant assignment')],
    'T001W':[(60,'Plant')], 'T001K':[(60,'Valuation area')],
    'T077K':[(0,'Vendor account groups')],
    'T7UNPAD_DS0P':[(80,'HR custom')],
    # Add patterns for special objects
}

def explain(obj, name):
    """Plain-English explanation of an object."""
    name_u = name.upper().strip()
    if obj == 'TABU':
        if name_u in ('T001','T001Q','T001A'): return 'Company code core record (T001 family)'
        if name_u in ('T012','T012D','T012K','T012T'): return 'House bank / bank account / bank text'
        if name_u in ('T018V','T018C','T018D','T028','T028B'): return 'Bank directory + bank type'
        if name_u.startswith('T042'): return 'FBZP payment program config'
        if name_u.startswith('T043') or name_u.startswith('T169'): return 'Tolerance group config'
        if name_u.startswith('T093') or name_u in ('AAACC_OBJ','ATPRA'): return 'Asset accounting config'
        if name_u == 'SKB1': return 'GL accounts at company-code level'
        if name_u in ('SKA1','SKAT'): return 'GL master / GL text'
        if name_u in ('T035D','T035U'): return 'Cash management planning groups'
        if name_u == 'T047': return 'Dunning procedure assignment'
        if name_u == 'T882': return 'Consolidation unit'
        if name_u.startswith('GB'): return 'Validation/Substitution rule (GGB0/GGB1 family)'
        if name_u.startswith('TKA'): return 'Controlling area config'
        if name_u in ('CSKA','CSKB','CSKS','CSKT','CSKU'): return 'CO master data (cost element/center)'
        if name_u.startswith('FM') or 'FUND' in name_u: return 'Funds Management config'
        if name_u.startswith('TPIR') or name_u in ('T500P','T542A'): return 'HR / Travel personnel config'
        if name_u.startswith('TCJ'): return 'Cash journal config'
        if name_u in ('T010O','T010P','T001B'): return 'Posting periods config'
        if name_u.startswith('T024'): return 'Purchasing org config'
        if name_u.startswith('T001W') or name_u == 'T001K': return 'Plant / valuation area config'
        if name_u.startswith('T527') or name_u.startswith('T7UN'): return 'HR personnel customizing'
        if name_u in ('UNESGBB','UNESWRX'): return 'UNESCO custom MM account determination keys'
        return f'Customizing table {name}'
    if obj == 'TDAT':
        if name_u == 'GGB1': return 'GGB1 substitution bundle (carries GB90/921/922/T)'
        if name_u == 'GGB0': return 'GGB0 validation bundle'
        if name_u == 'ADDRESS': return 'Address tables bundle (ADR2..ADR13, ADRC...)'
        if name_u == 'GRW_SET': return 'GS02 sets (BSET) bundle'
        if name_u.startswith('KE34'): return 'CO-PA operating concern data'
        return f'Transport data bundle {name}'
    if obj == 'VDAT':
        return f'View data {name}'
    if obj == 'CDAT':
        return f'Customer table data {name}'
    if obj == 'F01U':
        return f'Cross-company clearing pair {name}'
    if obj == 'CORR':
        return f'Sub-task release record (parent transport)'
    if obj == 'RELE':
        return f'Released sub-task record'
    return f'{obj} {name}'

# Pull E071/E071K for each current STEM transport
print('=== Pulling E071/E071K for current STEM transports ===\n')
inventory = []
for tr, date, user, fn, desc in STEM_TRANSPORTS:
    res = g.call('RFC_READ_TABLE', QUERY_TABLE='E071',
                 FIELDS=[{'FIELDNAME':'PGMID'},{'FIELDNAME':'OBJECT'},{'FIELDNAME':'OBJ_NAME'}],
                 OPTIONS=[{'TEXT':f"TRKORR = '{tr}'"}], DELIMITER='|', ROWCOUNT=500)
    objs = []
    for d in res.get('DATA', []):
        wa = [x.strip() for x in d['WA'].split('|')]
        objs.append({'PGMID':wa[0],'OBJECT':wa[1],'OBJ_NAME':wa[2]})
    # Key counts per (object, name)
    kres = g.call('RFC_READ_TABLE', QUERY_TABLE='E071K',
                  FIELDS=[{'FIELDNAME':'OBJECT'},{'FIELDNAME':'OBJNAME'},{'FIELDNAME':'TABKEY'}],
                  OPTIONS=[{'TEXT':f"TRKORR = '{tr}'"}], DELIMITER='|', ROWCOUNT=20000)
    key_counter = Counter()
    sample_keys = defaultdict(list)
    for d in kres.get('DATA', []):
        wa = d['WA'].split('|', 2)
        ob = wa[0].strip(); nm = wa[1].strip(); tk = wa[2].strip() if len(wa)>2 else ''
        key_counter[(ob, nm)] += 1
        if len(sample_keys[(ob, nm)]) < 3:
            sample_keys[(ob, nm)].append(tk)
    print(f'{tr} ({date} {user} fn={fn}): {len(objs)} objects, {sum(key_counter.values())} keys')
    for o in objs:
        cnt = key_counter.get((o['OBJECT'], o['OBJ_NAME']), 0)
        plain = explain(o['OBJECT'], o['OBJ_NAME'])
        gold = GOLDEN_MAP.get(o['OBJ_NAME'].upper(), [])
        gold_str = '; '.join(f'#{g[0]} {g[1]}' for g in gold) if gold else ''
        samples = '; '.join(sample_keys.get((o['OBJECT'], o['OBJ_NAME']), [])[:2])
        inventory.append({
            'transport': tr, 'date': date, 'user': user, 'fn': fn, 'tr_desc': desc,
            'pgmid': o['PGMID'], 'object': o['OBJECT'], 'obj_name': o['OBJ_NAME'],
            'keys': cnt, 'sample_keys': samples,
            'explanation': plain, 'maps_to_golden': gold_str,
        })

print(f'\nTotal inventory rows: {len(inventory)}')

# Save to JSON
with open('Zagentexecution/mcp-backend-server-python/stem_inventory_2026-04-16.json','w') as f:
    json.dump(inventory, f, indent=2)

# Add tab to Excel
SRC = 'Zagentexecution/Golden_live_matrix_2026-04-16_realigned.xlsx'
DST = 'Zagentexecution/Golden_live_matrix_2026-04-16_realigned_v3.xlsx'
wb = load_workbook(SRC)
if 'STEM Inventory' in wb.sheetnames:
    del wb['STEM Inventory']
ws = wb.create_sheet('STEM Inventory')

hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
hdr_font = Font(color='FFFFFF', bold=True)
tr_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')

headers = ['Transport','Date','Owner','FN','Desc','PGMID','OBJECT','OBJ_NAME','Keys','Sample TABKEY','Explanation','Maps to Golden Step']
ws.append(headers)
for c in ws[1]:
    c.fill = hdr_fill; c.font = hdr_font
    c.alignment = Alignment(horizontal='center', wrap_text=True)

current_tr = None
for inv in inventory:
    if inv['transport'] != current_tr:
        # Group separator
        ws.append([f"== {inv['transport']} == ({inv['date']} {inv['user']} fn={inv['fn']}) {inv['tr_desc']}"] + ['']*11)
        for c in ws[ws.max_row]:
            c.fill = tr_fill
            c.font = Font(bold=True)
        current_tr = inv['transport']
    ws.append([
        inv['transport'], inv['date'], inv['user'], inv['fn'], inv['tr_desc'],
        inv['pgmid'], inv['object'], inv['obj_name'], inv['keys'],
        inv['sample_keys'], inv['explanation'], inv['maps_to_golden'],
    ])

widths = [16, 12, 12, 4, 35, 6, 7, 22, 7, 30, 50, 30]
for col_letter, w in zip('ABCDEFGHIJKL', widths):
    ws.column_dimensions[col_letter].width = w
ws.row_dimensions[1].height = 32
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:L{ws.max_row}'

wb.save(DST)
print(f'\nSaved: {DST}')
print(f'New tab "STEM Inventory" has {ws.max_row - 1} rows')
g.close()
