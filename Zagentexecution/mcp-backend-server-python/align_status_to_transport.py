"""
Align DONE column to STEM transport-order reality:
- STEM = ✓  → Done
- STEM = ◐  → In Progress (partial)
- STEM = ✗  → Todo (not yet in transport)
- STEM = ∅  → N/A (migrated, not applicable to STEM)
- STEM = —  → Manual (no transport signature)

For NEW items P21-P28: leave as-set (no transport reality, manual entry).
"""
import sys
from openpyxl import load_workbook
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

SRC = 'Zagentexecution/Golden_live_matrix_2026-04-16_realigned.xlsx'
wb = load_workbook(SRC)
ws = wb['Golden × Live']

# STEM column is at index 18 (R) — let me confirm
hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
stem_col = hdr.index('STEM') + 1  # openpyxl is 1-indexed
print(f'Headers: {hdr}')
print(f'STEM column index: {stem_col} (= column {chr(ord("A")+stem_col-1)})')

# Map STEM symbol → DONE value
STATUS_MAP = {
    '✓': 'Done',
    '✗': 'Todo',
    '∅': 'N/A',
    '—': 'Manual',
}
def stem_to_done(stem_val):
    if stem_val is None or stem_val == '': return None
    s = str(stem_val).strip()
    if s in STATUS_MAP: return STATUS_MAP[s]
    if s.startswith('◐'): return 'In Progress'
    return None

# Use STEM transport reality for ALL 75 main items — no curated overrides
updated = 0
for r_idx in range(2, ws.max_row + 1):
    item = ws.cell(row=r_idx, column=2).value
    if item is None: continue
    if not isinstance(item, int): continue
    if item > 75: continue  # skip new P21-P28 items (manual entries)
    stem_val = ws.cell(row=r_idx, column=stem_col).value
    derived = stem_to_done(stem_val)
    if derived is not None:
        ws.cell(row=r_idx, column=22, value=derived)  # column V = DONE
        updated += 1

# Print summary
print(f'\nUpdated DONE column on {updated} existing items based on STEM transport reality.')

# Show the status distribution
from collections import Counter
status_count = Counter()
for r_idx in range(2, ws.max_row + 1):
    item = ws.cell(row=r_idx, column=2).value
    if item is None: continue
    if not isinstance(item, int): continue
    if item > 75: continue
    s = ws.cell(row=r_idx, column=22).value or '(blank)'
    status_count[s] += 1

print('\nDONE distribution (75 main items):')
for s, c in sorted(status_count.items(), key=lambda x: -x[1]):
    print(f'  {s:15s} {c}')

wb.save(SRC)
print(f'\nSaved: {SRC}')
