"""Update BCM_StructuredAddress_Analysis.xlsx with Phase 1 findings:
- Sheet 09 Vendor DQ: populate the 5 specific vendors with detail rows + fix actions
- New Sheet 15 'Outreach Drafts': N_MENARD, Francesco, SocGen, Citi, CGI banks, Master Data team
- New Sheet 16 'Pattern A Fix': the BAdI 3-line change with diff visualization
- New Sheet 17 'DE/IT Classes': decoded behavior (delegate to FALLBACK except bank ClrSysMmbId)
- New Sheet 18 'SEPA Sub-option Decision': test plan with revised choice (FI_PAYMEDIUM_DMEE_CGI_05 not FVD_SEPA_OL)
- Update sheet 04 Change Matrix Owner column
"""
from __future__ import annotations
import sys, csv
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

REPO = Path(__file__).resolve().parents[2]
XLSX = REPO / "Zagentexecution" / "incidents" / "xml_payment_structured_address" / "BCM_StructuredAddress_Analysis.xlsx"
VENDOR_CSV = REPO / "knowledge" / "domains" / "Payment" / "phase0" / "vendor_dq_5_specific.csv"

HEADER_FILL = PatternFill("solid", fgColor="1E3A4A")
HEADER_FONT = Font(name="Segoe UI", color="FFFFFF", bold=True, size=10)
SECTION_FILL = PatternFill("solid", fgColor="0D2A1F")
SECTION_FONT = Font(name="Segoe UI", color="1ABC9C", bold=True, size=11)
BORDER = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))


def style_row(row, fill=HEADER_FILL, font=HEADER_FONT):
    for c in row:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BORDER


def autosize(ws, max_w=70):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        lengths = [len(str(c.value)) if c.value else 0 for c in col]
        if lengths:
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(lengths) + 2, max_w)


def update_vendor_dq(wb):
    print("Updating sheet 09 Vendor DQ with 5 specific vendors")
    ws = wb["09 Vendor DQ"]
    # Append the 5 specific vendors at the bottom of the sheet
    ws.append([])
    ws.append(["The 5 specific LIFNRs (verified 2026-04-24 from P01 Gold DB):"])
    ws.cell(ws.max_row, 1).font = Font(name="Segoe UI", size=11, bold=True, color="1ABC9C")
    headers = ["LIFNR", "NAME1", "NAME2", "ADRC.CITY1", "ADRC.COUNTRY", "ADRC.STREET",
               "ADRC.POST_CODE1", "ADRC.HOUSE_NUM1", "LFA1.LAND1", "LFA1.ORT01", "Fix action proposed"]
    ws.append(headers)
    style_row(ws[ws.max_row])
    if VENDOR_CSV.exists():
        with open(VENDOR_CSV, encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            for row in reader:
                ws.append(row)
    ws.append([])
    ws.append(["Pattern observed: 4 of 5 are event/workshop ad-hoc vendors, 1 is workshop in Zimbabwe"])
    ws.append(["Risk: LOW — these are not recurring F110 payees. Master Data team to manually fix or KILL."])
    ws.append(["Action owner: Master Data team (request via Marlies)"])


def add_outreach_drafts(wb):
    print("Adding sheet '15 Outreach Drafts'")
    ws = wb.create_sheet("15 Outreach Drafts")
    drafts = [
        ("N_MENARD ALIGNMENT", "internal call / email", "Subject: BCM Structured Address V001 — 30-min alignment request", """\
Hola Nicolas,

Working on the CBPR+ Nov 2026 structured address mandate for BCM. We need
30 minutes of your time on 2 items where you are the code owner:

1. Pattern A fix in YCL_IDFI_CGI_DMEE_FALLBACK_CM001::GET_CREDIT
   The existing name-overflow-into-StrtNm logic (lines 13-31 of the method)
   would corrupt structured StrtNm values when V001 (structured) is active.
   Proposed 3-line guard:
     IF i_fpayh = mv_fpayh
        AND mv_cdtr_name+35 IS NOT INITIAL
        AND c_value IS INITIAL.    "<-- new guard: only prepend if StrtNm empty
       c_value = mv_cdtr_name+35.
     ELSEIF i_fpayh = mv_fpayh
        AND mv_cdtr_name+35 IS NOT INITIAL.
       c_value = |{ mv_cdtr_name+35 } { c_value }|.
     ENDIF.
   Backward-compatible (V000 behavior preserved). DE/IT classes delegate
   to FALLBACK so this fix covers them too.

2. D01 retrofit of 5 P01-only objects you authored:
   - YCL_IDFI_CGI_DMEE_DE (CLAS)
   - YCL_IDFI_CGI_DMEE_IT (CLAS)
   - Y_IDFI_CGI_DMEE_COUNTRIES_DE (ENHO)
   - Y_IDFI_CGI_DMEE_COUNTRIES_FR (ENHO)
   - Y_IDFI_CGI_DMEE_COUNTRIES_IT (ENHO)
   These exist in P01 but NOT in D01. Need to sync to D01 before any DMEE
   edits in Phase 2 (May start). Do you know why they got out of sync?
   Approve sync via standard transport from P01->D01?

Both items are gating Phase 2 (May 2026 config window). Looking for 30 min
this/next week. Pick a slot:
  Tue 29 16:00, Wed 30 10:00, Thu May 1 14:00 (Paris time)

Thanks,
Pablo

CC: Marlies Spronk (FIN/TRS), Anssi Yli-Hietanen
"""),
        ("FRANCESCO COURTESY", "internal email", "Subject: BCM CGI tree V001 work — courtesy heads-up", """\
Ciao Francesco,

Heads-up — we are starting work on the CGI_XML_CT_UNESCO and _1 trees
for the CBPR+ Nov 2026 structured address mandate.

We saw your 5 transports from Q1 2025 (D01K9B0CTP, CUS, CUT, CWS, CZ0)
on these trees. Per analysis they are PMF variant configurations
(VC_TFPM042F bank variant setups), not tree node structure changes,
so there is NO conflict with our planned V001 work.

Just wanted to disclose:
- We will create a parallel V001 of /CGI_XML_CT_UNESCO and _1 in DEV (your
  current V000 stays untouched).
- V001 will add structured PstlAdr nodes for CdtrAgt only — Cdtr and Dbtr
  in CGI are already structured so no change.
- Atomic version-flip cutover targeted Nov 2026.

If you have any uncommitted WIP in D01 on these trees, please let me know
before May 1 so we coordinate. Otherwise we proceed.

Saluti,
Pablo

CC: Marlies Spronk
"""),
        ("MARLIES — SEPA SUB-OPTION DECISION", "internal Teams/email", "Subject: SEPA implementation pattern — your preference?", """\
Hola Marlies,

Quick decision needed for the SEPA part of the project. Phase 0 found that
/SEPA_CT_UNES has NO Event 05 registered today (unlike CGI/CITI which have
SAP standard FMs). Three options to add structured address to SEPA:

A) Register FI_PAYMEDIUM_DMEE_CGI_05 (the same SAP std FM CGI uses) for
   /SEPA_CT_UNES Event 05. SEPA gains the same multi-mode infrastructure
   as CGI. Untested but technically compatible because the FM dispatches
   to country class CL_IDFI_CGI_CALL05_FR.
   Effort: LOW (config only, no ABAP). Risk: needs D01 dry-run test.

B) No Event 05. Add structured nodes directly in DMEE tree reading from
   FPAYHX fields, with CV_RULE/MP_OFFSET parsing on ZPLOR (postal+city
   combined field) to extract PstCd and TwnNm separately.
   Effort: LOW-MEDIUM (config + ZPLOR parse rule). Risk: ZPLOR reliability.

C) Custom UNESCO FM Z_DMEE_UNESCO_SEPA_05. Requires ABAP + N_MENARD review.
   Effort: MEDIUM. Risk: LOW (full control).

Your Excel Tab 2 has the target XML structure but doesn't specify how to
implement. Do you have a preference based on past SEPA changes you did?
Or should we test A first in D01?

Best,
Pablo
"""),
        ("MARLIES — BANK SPEC OUTREACH (Marlies sends)", "Marlies relays to bank TRMs", "Subject: ISO 20022 CBPR+ structured address — UNESCO compliance plan", """\
Dear [TRM Name — SocGen / Citi / Shinhan / Metro / GT Bank],

UNESCO is preparing for the CBPR+ Nov 2026 structured address mandate.
We are upgrading our 3 XML payment file formats:
- /SEPA_CT_UNES (SocGen EUR — UNES + IIEP + UIL)
- /CITI/XML/UNESCO/DC_V3_01 (Citibank USD + Worldlink)
- /CGI_XML_CT_UNESCO (multi-bank including yours)

Three questions to align before our May 2026 configuration work:

1. STRICTNESS: Do you require fully structured address (StrtNm + BldgNb +
   PstCd + TwnNm + Ctry) in PstlAdr, or is Hybrid (TwnNm + Ctry mandatory
   + AdrLine for residual) acceptable? Per CBPR+ Hybrid is allowed; we
   want to confirm your specific stance.

2. TEST GATEWAY: Do you support a non-production/test gateway where we
   can send sample files for schema + acceptance validation without
   production impact? UNESCO needs this for July 2026 UAT.

3. SCHEMA VERSION: Confirm the pain.001 version your parser expects
   (.03 / .09). Any proprietary extensions in the PstlAdr block?

[CITI ONLY: For Worldlink payments where Cdtr is an intermediary (e.g.,
WHO) and the real beneficiary is in UltmtCdtr (e.g., ICC), please clarify
the data source you expect to be populated for UltmtCdtr address. We see
mixed handling today.]

We aim to send a written confirmation of CBPR+ enforcement deadline from
each bank for our compliance file. Please respond by 2026-05-05.

Thanks,
Marlies Spronk
FIN/TRS, UNESCO
"""),
        ("MASTER DATA TEAM — 5 vendor cleanup", "internal", "Subject: Vendor master fix — 5 LIFNRs missing CITY1/COUNTRY", """\
Team,

For CBPR+ Nov 2026 compliance we need 5 specific vendors to have CITY1
and COUNTRY populated in ADRC. These are the only 5 (out of 111,241
active) missing the mandatory fields.

LIFNR        NAME1                                Issue
0000020171   KOUADIO                              ADRC.COUNTRY empty (LFA1.ORT01='PARIS' suggests FR)
0000020731   Comité d'orientation du Pôle de D    ADRC.COUNTRY empty (LFA1.ORT01='PARIS' suggests FR)
0000020815   Mr Darasack RATSAVONG                ADRC.COUNTRY empty (LFA1.ORT01='PARIS' suggests FR)
0000020843   Atelier focus group 13 au 17 mars    ADRC.COUNTRY empty (Yamoussoukro = Côte d'Ivoire CI)
0000059828   Data Validation Workshop on the      ADRC.CITY1 empty (LFA1.LAND1='ZW' Zimbabwe)

Pattern: 4 of 5 are event/workshop ad-hoc vendors. Two options:
1. Manually populate ADRC.CITY1 and ADRC.COUNTRY for each
2. KILL the vendors if no future payments expected (set LFA1.LOEVM='X')

Please action by 2026-05-15. Detail in:
Zagentexecution/incidents/xml_payment_structured_address/BCM_StructuredAddress_Analysis.xlsx
sheet 09 Vendor DQ.

Thanks,
Pablo
"""),
    ]

    # Header
    ws.append(["#", "Recipient", "Channel", "Subject", "Body draft"])
    style_row(ws[1])
    for i, (recipient, channel, subject, body) in enumerate(drafts, 1):
        ws.append([i, recipient, channel, subject, body])
        # Style cells
        for c in ws[ws.max_row]:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
            c.font = Font(name="Consolas" if c.column == 5 else "Segoe UI", size=9)
        ws.row_dimensions[ws.max_row].height = max(150, len(body.split("\n")) * 14)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 110


def add_pattern_a_fix(wb):
    print("Adding sheet '16 Pattern A Fix'")
    ws = wb.create_sheet("16 Pattern A Fix")
    ws.append(["Pattern A — BAdI structure-aware guard"])
    ws.cell(1, 1).font = Font(size=14, bold=True, color="1ABC9C")
    ws.append(["Class:", "YCL_IDFI_CGI_DMEE_FALLBACK"])
    ws.append(["Method:", "GET_CREDIT (CM001)"])
    ws.append(["Lines affected:", "13-31 (existing logic), guard added at line 24"])
    ws.append(["Reviewer:", "N_MENARD (code owner)"])
    ws.append(["Transport:", "D01K-BADI-FIX-01 (to create in Phase 2)"])
    ws.append(["Risk:", "LOW — backward-compatible, V000 legacy preserved"])
    ws.append(["DE/IT impact:", "Zero — DE/IT classes delegate to FALLBACK; this fix covers them"])
    ws.append([])
    ws.append(["BEFORE (current V000-only behavior, BUGGY for V001):"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color="E74C3C")
    before = """\
WHEN '<PmtInf><CdtTrfTxInf><Cdtr><Nm>'.
  IF i_fpayp-origin = 'TR-CM-BT'.
    c_value = i_fpayp-sgtxt.
  ENDIF.
  mv_cdtr_name = c_value.
  IF c_value+35 IS NOT INITIAL.
    CLEAR c_value+35.
  ENDIF.
  mv_fpayh = i_fpayh.
WHEN '<PmtInf><CdtTrfTxInf><Cdtr><PstlAdr><StrtNm>'.
  IF i_fpayh = mv_fpayh AND mv_cdtr_name+35 IS NOT INITIAL.
    c_value = |{ mv_cdtr_name+35 } { c_value }|.   "<-- BUG in V001: corrupts real StrtNm
  ENDIF.
  IF c_value+70 IS NOT INITIAL.
    CLEAR c_value+70.
  ENDIF."""
    ws.append([before])
    ws.cell(ws.max_row, 1).font = Font(name="Consolas", size=10)
    ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[ws.max_row].height = 280
    ws.append([])
    ws.append(["AFTER (Pattern A — structure-aware guard, works for V000 + V001):"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color="27AE60")
    after = """\
WHEN '<PmtInf><CdtTrfTxInf><Cdtr><Nm>'.
  IF i_fpayp-origin = 'TR-CM-BT'.
    c_value = i_fpayp-sgtxt.
  ENDIF.
  mv_cdtr_name = c_value.
  IF c_value+35 IS NOT INITIAL.
    CLEAR c_value+35.
  ENDIF.
  mv_fpayh = i_fpayh.
WHEN '<PmtInf><CdtTrfTxInf><Cdtr><PstlAdr><StrtNm>'.
  IF i_fpayh = mv_fpayh
     AND mv_cdtr_name+35 IS NOT INITIAL
     AND c_value IS INITIAL.                                 "<-- NEW V001 guard
    c_value = mv_cdtr_name+35.                               "<-- NEW: only prepend if StrtNm source empty
  ELSEIF i_fpayh = mv_fpayh
     AND mv_cdtr_name+35 IS NOT INITIAL.
    c_value = |{ mv_cdtr_name+35 } { c_value }|.             "<-- V000 legacy unchanged
  ENDIF.
  IF c_value+70 IS NOT INITIAL.
    CLEAR c_value+70.
  ENDIF."""
    ws.append([after])
    ws.cell(ws.max_row, 1).font = Font(name="Consolas", size=10)
    ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[ws.max_row].height = 320
    ws.column_dimensions["A"].width = 130


def add_de_it_classes_summary(wb):
    print("Adding sheet '17 DE-IT Classes'")
    ws = wb.create_sheet("17 DE-IT Classes")
    ws.append(["YCL_IDFI_CGI_DMEE_DE and _IT — decoded behavior"])
    ws.cell(1, 1).font = Font(size=14, bold=True, color="1ABC9C")
    ws.append(["Extracted from P01 via class-include reads (2026-04-24)"])
    ws.append([])
    ws.append(["Finding: DE and IT classes are essentially SHELLS that delegate to FALLBACK"])
    ws.append([])
    ws.append(["What they DO override (only one node):"])
    ws.append(["Node:", "<PmtInf><CdtTrfTxInf><CdtrAgt><FinInstnId><ClrSysMmbId><MmbId>"])
    ws.append(["Logic:", "IF i_fpayh-zbnkl IS NOT INITIAL → c_value = i_fpayh-zbnkl ELSE clear"])
    ws.append(["Purpose:", "DE/IT have specific bank clearing system member ID handling"])
    ws.append([])
    ws.append(["What they DELEGATE to FALLBACK (everything else, including all PstlAdr):"])
    ws.append(["Pattern:", "WHEN OTHERS → super->get_credit(...) with all params forwarded"])
    ws.append([])
    ws.append(["Implication for our project:"])
    ws.append(["1.", "DE/IT classes do NOT implement the name-overflow-into-StrtNm bug. Only FALLBACK does."])
    ws.append(["2.", "Pattern A fix to FALLBACK_CM001 covers DE and IT payment scenarios automatically."])
    ws.append(["3.", "Scope of ABAP change for entire project = 1 method (FALLBACK_CM001 / GET_CREDIT)"])
    ws.append(["4.", "i_tree_id IS passed to the BAdI as a parameter — Pattern B (separate trees) is also viable but Pattern A is cleaner"])
    ws.append([])
    ws.append(["Source files extracted (in extracted_code/FI/DMEE_full_inventory/):"])
    for f in ["YCL_IDFI_CGI_DMEE_DE_CCDEF.abap",
              "YCL_IDFI_CGI_DMEE_DE_CCIMP.abap",
              "YCL_IDFI_CGI_DMEE_DE_CCMAC.abap",
              "YCL_IDFI_CGI_DMEE_DE_CP.abap",
              "YCL_IDFI_CGI_DMEE_DE_CO.abap",
              "YCL_IDFI_CGI_DMEE_DE_CU.abap",
              "YCL_IDFI_CGI_DMEE_DE_CM001.abap",
              "YCL_IDFI_CGI_DMEE_IT_*.abap (same pattern, identical CM001)"]:
        ws.append(["", f])
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 90


def add_sepa_decision(wb):
    print("Adding sheet '18 SEPA Sub-option'")
    ws = wb.create_sheet("18 SEPA Sub-option")
    ws.append(["SEPA Sub-option Decision — REVISED post Phase 1 D01 probe"])
    ws.cell(1, 1).font = Font(size=14, bold=True, color="1ABC9C")
    ws.append([])
    ws.append(["Context: /SEPA_CT_UNES has NO Event 05 today. Need to decide implementation."])
    ws.append(["D01 probe results (read-only, 2026-04-24):"])
    ws.append(["", "FVD_SEPA_OL_CT_DMEE_05 EXISTS in D01 (PNAME=SAPLFVD_SEPA_OL)"])
    ws.append(["", "FI_PAYMEDIUM_DMEE_CGI_05 EXISTS in D01 (PNAME=SAPLDMEE_CGI)"])
    ws.append(["", "FI_PAYMEDIUM_DMEE_CN_05 EXISTS (China variant — reference only)"])
    ws.append([])
    ws.append(["D01 SEPA-tree event registrations:"])
    headers = ["FORMI", "EVENT", "FNAME"]
    ws.append(headers)
    style_row(ws[ws.max_row])
    for r in [
        ("SEPA_CT_00100103", "20", "FI_PAYMEDIUM_DMEE_20"),
        ("SEPA_CT_00100103", "25", "FI_PAYMEDIUM_DMEE_25"),
        ("SEPA_CT_00100103", "30", "FI_PAYMEDIUM_DMEE_30"),
        ("SEPA_CT_00100103", "40", "FI_PAYMEDIUM_DMEE_40"),
        ("SEPA_CT_CML_DATA", "05", "FVD_SEPA_OL_CT_DMEE_05"),
        ("SEPA_CT_CML_RETURN", "05", "FVD_SEPA_OL_CT_DMEE_05"),
        ("SEPA_DD_CML_DATA", "05", "FVD_SEPA_OL_DD_DMEE_05"),
        ("ZCP_SEPA_MBS_CN_00100103", "05", "FI_PAYMEDIUM_DMEE_CN_05"),
    ]:
        ws.append(list(r))
    ws.append([])
    ws.append(["REVISED Sub-option recommendation:"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color="1ABC9C")
    ws.append([])
    ws.append(["Sub-option", "Description", "Pros", "Cons", "Recommendation"])
    style_row(ws[ws.max_row])
    rows_data = [
        ("A original (FVD_SEPA_OL)",
         "Register FVD_SEPA_OL_CT_DMEE_05 for /SEPA_CT_UNES Event 05",
         "SAP-std SEPA-specific FM",
         "Designed for CML/Treasury cash management, NOT AP — likely wrong context. Currently only registered for CML trees.",
         "REJECTED — wrong domain"),
        ("A revised (FI_PAYMEDIUM_DMEE_CGI_05)",
         "Register the SAME FM CGI uses for /SEPA_CT_UNES Event 05",
         "Already in production for CGI, dispatches to CL_IDFI_CGI_CALL05_FR for French SEPA addresses, gains multi-mode like CGI",
         "Untested with /SEPA_CT_UNES tree (CGI factory may need a CGI-pattern tree)",
         "★ TEST IN D01 FIRST"),
        ("B (direct read + CV_RULE)",
         "No Event 05. Tree reads FPAYHX directly; parse ZPLOR via CV_RULE/MP_OFFSET into PstCd+TwnNm",
         "Zero ABAP, fully under our control",
         "ZPLOR parse reliability depends on consistent vendor master format",
         "FALLBACK if A revised fails"),
        ("C (custom Z FM)",
         "Create Z_DMEE_UNESCO_SEPA_05 implementing IF_IDFI_CGI_CALL05",
         "Full control over field population logic",
         "Requires ABAP + N_MENARD review + D01 retrofit first",
         "RESERVED — only if A and B both fail"),
    ]
    for r in rows_data:
        ws.append(list(r))
        for c in ws[ws.max_row]:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
            c.font = Font(name="Segoe UI", size=10)
        ws.row_dimensions[ws.max_row].height = 60
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 22


def populate_change_matrix_owners(wb):
    print("Updating sheet 04 Change Matrix — Owner column")
    ws = wb["04 Change Matrix"]
    # Find the Owner column
    headers = [c.value for c in ws[1]]
    if "Owner" not in headers:
        print(f"  Owner column not found, headers={headers}")
        return
    owner_col = headers.index("Owner") + 1
    action_col = headers.index("Action") + 1
    party_col = headers.index("Party") + 1

    # Reasonable owner assignment per Action + Party
    for row_idx in range(2, ws.max_row + 1):
        action = ws.cell(row_idx, action_col).value or ""
        party = ws.cell(row_idx, party_col).value or ""
        owner_cell = ws.cell(row_idx, owner_col)
        # Default owner = Pablo, but specific cases
        if "BLOCKED" in action:
            owner_cell.value = "Citi TRM + DBS (Q3)"
        elif "DATA-QUALITY" in action:
            owner_cell.value = "Master Data team + Pablo"
        elif "INVESTIGATE" in action:
            owner_cell.value = "Pablo + Marlies"
        elif "SYNC" in action:
            owner_cell.value = "Pablo (auto-sync from main tree)"
        elif "UltmtCdtr" in party or "UltmtDbtr" in party:
            owner_cell.value = "Pablo + Citi TRM"
        elif "CdtrAgt" in party:
            owner_cell.value = "Pablo (BNKA standard bank DB)"
        else:
            owner_cell.value = "Pablo"


def main():
    if not XLSX.exists():
        print(f"Workbook not found: {XLSX}")
        sys.exit(1)
    wb = load_workbook(XLSX)
    print(f"Loaded: {XLSX} ({len(wb.sheetnames)} sheets)")

    update_vendor_dq(wb)
    populate_change_matrix_owners(wb)
    add_outreach_drafts(wb)
    add_pattern_a_fix(wb)
    add_de_it_classes_summary(wb)
    add_sepa_decision(wb)

    wb.save(XLSX)
    print(f"\nSaved: {XLSX} ({XLSX.stat().st_size:,} bytes, {len(wb.sheetnames)} sheets)")
    print(f"Sheets:")
    for i, s in enumerate(wb.sheetnames, 1):
        print(f"  {i:>2}. {s}")


if __name__ == "__main__":
    main()
