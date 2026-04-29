"""
Extract Funds (funds) + Fund Centers (fund_centers) + Projects (proj/prps)
for the 3 reference institutes MGIE / IBE / ICBA.

All data sourced from Gold DB. CSKS (Cost Centers) and T036FT texts
require RFC and are appended later when VPN is up.

Output: Zagentexecution/STEM_3institutes_reference_<date>.xlsx
"""
import sqlite3
import sys
from pathlib import Path
from datetime import datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

HERE = Path(__file__).resolve().parent
REPO = HERE.parents[1]
GOLD = REPO / "Zagentexecution" / "sap_data_extraction" / "sqlite" / "p01_gold_master_data.db"
INSTITUTES = ["MGIE", "IBE", "ICBA"]
OUT_XLSX = REPO / "Zagentexecution" / f"STEM_3institutes_reference_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

FUND_TYPE_CODES_HINT = {
    "100": "Regular Programme (RP) — UNESCO standard hint",
    "200": "Extra-Budgetary Funds-In-Trust (FIT) — UNESCO standard hint",
    "300": "Self-Benefitting / Cost Recovery — UNESCO standard hint",
    "301": "Institute-specific operations — UNESCO standard hint",
    "302": "Institute-specific operations — UNESCO standard hint",
    "303": "Institute-specific operations — UNESCO standard hint",
    "304": "Institute-specific operations — UNESCO standard hint",
    "305": "Institute-specific operations — UNESCO standard hint",
    "306": "Institute-specific operations — UNESCO standard hint",
}


def query_all(con, sql, params=()):
    cur = con.cursor()
    cur.execute(sql, params)
    cols = [d[0] for d in cur.description]
    return cols, cur.fetchall()


def main():
    print(f"Gold DB: {GOLD}")
    con = sqlite3.connect(GOLD)
    placeholders = ",".join(["?"] * len(INSTITUTES))

    funds_cols, funds = query_all(
        con,
        f"SELECT FIKRS, FINCODE, TYPE, ERFDAT, ERFNAME FROM funds WHERE FIKRS IN ({placeholders}) ORDER BY FIKRS, FINCODE",
        INSTITUTES,
    )
    print(f"  funds: {len(funds)}")

    fc_cols, fc = query_all(
        con,
        f"SELECT FIKRS, FICTR, ERFDAT, ERFNAME FROM fund_centers WHERE FIKRS IN ({placeholders}) ORDER BY FIKRS, FICTR",
        INSTITUTES,
    )
    print(f"  fund_centers: {len(fc)}")

    proj_cols, proj = query_all(
        con,
        f"SELECT VBUKR, PSPID, POST1, VERNR, ERDAT, PSPNR FROM proj WHERE VBUKR IN ({placeholders}) ORDER BY VBUKR, PSPID",
        INSTITUTES,
    )
    print(f"  proj: {len(proj)}")

    prps_cols, prps = query_all(
        con,
        f"SELECT PBUKR, POSID, POST1, VERNR, ERDAT, PSPHI, PSPNR, OBJNR FROM prps WHERE PBUKR IN ({placeholders}) ORDER BY PBUKR, POSID",
        INSTITUTES,
    )
    print(f"  prps: {len(prps)}")

    # Distinct fund TYPE codes per institute (lieu of T036FT texts)
    ft_cols, ft = query_all(
        con,
        f"""
        SELECT FIKRS, TYPE AS FONDSART, COUNT(*) AS num_funds, MIN(ERFDAT) AS first_seen, MAX(ERFDAT) AS last_seen
          FROM funds WHERE FIKRS IN ({placeholders})
          GROUP BY FIKRS, TYPE ORDER BY FIKRS, TYPE
        """,
        INSTITUTES,
    )
    print(f"  fund_types_distinct: {len(ft)}")
    con.close()

    # ----- Excel -----
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    bold = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    note_fill = PatternFill("solid", fgColor="FFF2CC")

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["STEM Company-Code Copy — Reference Data (MGIE / IBE / ICBA)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Generated: {datetime.now().isoformat()}"])
    ws.append(["Source: Gold DB (Zagentexecution/sap_data_extraction/sqlite/p01_gold_master_data.db)"])
    ws.append([])
    ws.append(["Object", "MGIE", "IBE", "ICBA", "Total"])
    for c in range(1, 6):
        ws.cell(row=ws.max_row, column=c).font = bold
        ws.cell(row=ws.max_row, column=c).fill = header_fill

    def cnt(rows, key_idx, code):
        return sum(1 for r in rows if r[key_idx] == code)

    summary_data = [
        ("Fund Types (distinct codes, FMFINCODE.TYPE)", ft, 0),
        ("Funds (FMFINCODE -> funds)", funds, 0),
        ("Fund Centers (FMFCTR -> fund_centers)", fc, 0),
        ("Projects (PROJ)", proj, 0),
        ("WBS Elements (PRPS)", prps, 0),
        ("Cost Centers (CSKS) — RFC required", [], 0),
    ]
    for label, rows, idx in summary_data:
        if rows:
            counts = [cnt(rows, idx, i) for i in INSTITUTES]
            if label.startswith("Fund Types"):
                # ft already grouped by FIKRS, count distinct TYPEs per institute
                counts = [sum(1 for r in rows if r[idx] == i) for i in INSTITUTES]
        else:
            counts = ["pending"] * 3
        ws.append([label, *counts, sum(counts) if all(isinstance(x, int) for x in counts) else "—"])

    ws.append([])
    ws.append(["Notes:"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["• ICBA shows 0 funds and 0 fund_centers in Gold DB — verify whether ICBA reuses UNES funds or extraction missed them"])
    ws.append(["• T036FT (Fund Type DESCRIPTIONS) and CSKS (Cost Centers) are not in Gold DB — pending RFC pull on next VPN restore"])
    ws.append(["• Fund Type codes shown are inferred from FMFINCODE.TYPE distinct values (no text descriptions yet)"])
    for r in range(ws.max_row - 2, ws.max_row + 1):
        ws.cell(row=r, column=1).fill = note_fill

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 42 if col == 1 else 14

    # Generic adder
    def add_sheet(title, cols, rows, hint_col=None, hint_map=None):
        s = wb.create_sheet(title)
        header = list(cols)
        if hint_col:
            header.append("Hint (UNESCO convention)")
        s.append(header)
        for c in range(1, len(header) + 1):
            s.cell(row=1, column=c).font = bold
            s.cell(row=1, column=c).fill = header_fill
        for r in rows:
            row = list(r)
            if hint_col:
                key = r[cols.index(hint_col)]
                row.append(hint_map.get(str(key), ""))
            s.append(row)
        s.freeze_panes = "A2"
        for i, c in enumerate(header, 1):
            s.column_dimensions[get_column_letter(i)].width = max(12, min(45, len(c) + 4))

    add_sheet("FundTypes_distinct_codes",
              ["FIKRS", "FONDSART", "num_funds", "first_seen", "last_seen"], ft,
              hint_col="FONDSART", hint_map=FUND_TYPE_CODES_HINT)
    add_sheet("Funds_FMFINCODE", funds_cols, funds)
    add_sheet("FundCenters_FMFCTR", fc_cols, fc)
    add_sheet("Projects_PROJ", proj_cols, proj)
    add_sheet("WBS_PRPS", prps_cols, prps)

    # Empty placeholder for CSKS to make structure obvious
    s = wb.create_sheet("CostCenters_CSKS")
    s.append(["BUKRS", "KOKRS", "KOSTL", "KTEXT", "LTEXT", "KOSAR", "VERAK", "DATAB", "DATBI"])
    for c in range(1, 10):
        s.cell(row=1, column=c).font = bold
        s.cell(row=1, column=c).fill = header_fill
    s.append(["(awaiting RFC pull — Gold DB does not contain CSKS yet)"])
    s.cell(row=2, column=1).fill = note_fill

    wb.save(OUT_XLSX)
    print(f"\nExcel: {OUT_XLSX}")


if __name__ == "__main__":
    main()
