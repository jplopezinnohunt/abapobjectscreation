"""Build BCM_StructuredAddress_Analysis.xlsx — the canonical analysis workbook.

Multi-sheet workbook combining:
- Marlies Spronk's observations (verbatim from her "XML Address un structured.xlsx")
- Our Phase 0 technical evidence (DMEE nodes, conditions, headers, D01 vs P01)
- Change matrix per-tree per-node with action + owner + status + Marlies remark link
- Bank spec matrix (to fill in Phase 1)
- Q1-Q8 status tracker
- SAP Notes reference catalog

Output: Zagentexecution/incidents/xml_payment_structured_address/BCM_StructuredAddress_Analysis.xlsx
"""
from __future__ import annotations
import csv, sys
from pathlib import Path
from datetime import datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as e:
    print(f"Missing openpyxl: {e}")
    sys.exit(1)

REPO = Path(__file__).resolve().parents[2]
MARLIES_XLSX = REPO / "Zagentexecution" / "incidents" / "xml_payment_structured_address" / "original_marlies" / "XML Address un structured.xlsx"
PHASE0 = REPO / "knowledge" / "domains" / "Payment" / "phase0"
OUT = REPO / "Zagentexecution" / "incidents" / "xml_payment_structured_address" / "BCM_StructuredAddress_Analysis.xlsx"

# Styling
HEADER_FILL = PatternFill("solid", fgColor="1E3A4A")
HEADER_FONT = Font(name="Segoe UI", color="FFFFFF", bold=True, size=10)
CELL_FONT = Font(name="Segoe UI", size=10)
SECTION_FILL = PatternFill("solid", fgColor="0D1E2A")
SECTION_FONT = Font(name="Segoe UI", color="1ABC9C", bold=True, size=11)
ALT_FILL = PatternFill("solid", fgColor="F5F8FA")
BORDER_THIN = Border(left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"))
STATUS_COLORS = {
    "OPEN": "F39C12", "PENDING": "F39C12",
    "CLOSED": "27AE60", "DONE": "27AE60", "COMPLETED": "27AE60",
    "BLOCKED": "E74C3C", "REJECTED": "E74C3C",
    "PARKED": "95A5A6", "N/A": "95A5A6",
    "PARTIAL": "3498DB",
}


def style_header(ws, row=1, cols=None):
    if cols is None:
        cols = range(1, ws.max_column + 1)
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER_THIN


def autosize(ws, max_width=60):
    for col in ws.columns:
        lengths = [len(str(c.value)) if c.value is not None else 0 for c in col]
        if not lengths:
            continue
        w = min(max(lengths) + 2, max_width)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(w, 10)


def add_overview(wb):
    ws = wb.create_sheet("01 Overview", 0)
    rows = [
        ["UNESCO BCM Structured Address Change — Analysis Workbook"],
        [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
        [""],
        ["Purpose", "Combined analysis of Marlies's 10 production cases + Phase 0 technical evidence + change matrix for V001 (structured) parallel version"],
        ["Drivers", "CBPR+ Nov 2026 ISO 20022 mandate · bank requests (SocGen, Citi, CGI) · Marlies email 2026-04-14 · Agent handoff doc"],
        ["Strategy", "2-file + DMEE native versioning: V000 stays ACTIVE (legacy), V001 INACTIVE (new structured), atomic cutover Nov 2026"],
        [""],
        ["Target trees", ""],
        ["  /SEPA_CT_UNES", "Société Générale EUR — UNES + IIEP + UIL — 95 nodes P01 — VERSION 000 — last M_SPRONK 2021-11-23"],
        ["  /CITI/XML/UNESCO/DC_V3_01", "Citibank USD + Worldlink — UNES + UBO + UIS — 610 nodes — XSLT CGI_XML_CT_XSLT post-proc — last M_SPRONK 2023-01-31"],
        ["  /CGI_XML_CT_UNESCO", "SocGen non-SEPA + other banks — 631 nodes — last FP_SPEZZANO 2025-03-20 (D01K9B0CZ0)"],
        ["  /CGI_XML_CT_UNESCO_1", "Variant of CGI_XML_CT_UNESCO — 639 nodes — last FP_SPEZZANO 2025-02-14"],
        [""],
        ["Sheets in this workbook"],
        ["01 Overview", "(this sheet)"],
        ["02 Marlies Analysis", "Verbatim from her Excel, 10 cases across SEPA/CITI/CGI"],
        ["03 Marlies SEPA Proposal", "Her proposed SEPA target XML (Tab 2 of her Excel)"],
        ["04 Change Matrix", "Per-tree per-node action list for V001 (main execution tracker)"],
        ["05 DMEE Headers", "All 4 target trees' header attributes (VERSION, PARAM_STRUC, CHARSET, last-changed)"],
        ["06 Address Nodes", "339 address+party nodes with source fields, MP_OFFSET, CV_RULE, MP_EXIT_FUNC"],
        ["07 Conditions", "614 DMEE_TREE_COND rows gating emission of address nodes"],
        ["08 D01 vs P01", "5 P01-ONLY objects (retrofit blocker) + shared objects"],
        ["09 Vendor DQ", "5 vendors missing mandatory CITY1/COUNTRY (placeholder until query run)"],
        ["10 Francesco Audit", "5 FP_SPEZZANO transports classified"],
        ["11 Q1-Q8 Status", "From handoff doc §11, tracked to closure"],
        ["12 Bank Spec Matrix", "Per-bank requirement — to fill in Phase 1"],
        ["13 SAP Notes", "Canonical refs: 1665873 + 2795667 + 2668719 + 2819590 + 2845063"],
        ["14 Phase Plan", "5-phase timeline with status"],
    ]
    for r in rows:
        ws.append(r)
    ws.cell(row=1, column=1).font = Font(name="Segoe UI", size=16, bold=True, color="1ABC9C")
    ws.cell(row=2, column=1).font = Font(name="Segoe UI", size=9, italic=True, color="5D7A92")
    for r in range(4, 8):
        ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=8, column=1).font = Font(bold=True, color="1ABC9C")
    ws.cell(row=14, column=1).font = Font(bold=True, color="1ABC9C")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 80
    return ws


def copy_marlies_analysis(wb):
    """Copy Marlies's 2 sheets verbatim."""
    if not MARLIES_XLSX.exists():
        return None
    mwb = load_workbook(MARLIES_XLSX, data_only=True)
    for idx, sheet_name in enumerate(mwb.sheetnames):
        src = mwb[sheet_name]
        dest_name = f"0{2+idx} Marlies {sheet_name}"[:31]
        dest = wb.create_sheet(dest_name)
        for row in src.iter_rows(values_only=True):
            dest.append(row)
        # Style header
        if dest.max_row >= 1:
            style_header(dest)
            for row_idx in range(2, dest.max_row + 1):
                for col_idx in range(1, dest.max_column + 1):
                    cell = dest.cell(row=row_idx, column=col_idx)
                    cell.font = CELL_FONT
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = BORDER_THIN
        autosize(dest, max_width=50)
        for r in dest.iter_rows():
            dest.row_dimensions[r[0].row].height = 60
    return True


def add_change_matrix(wb):
    """Change matrix — per-tree per-node action for V001 creation."""
    ws = wb.create_sheet("04 Change Matrix")
    headers = ["Tree", "Party", "Node path (target)", "V000 current source", "V001 target source",
               "Action", "Version", "Marlies remark", "Evidence", "Owner", "Status", "Transport", "Test status"]
    ws.append(headers)

    # Seed with the core changes per Marlies Excel + our probe findings
    rows = [
        # === SEPA_CT_UNES ===
        ("SEPA_CT_UNES", "Dbtr", "Dbtr/Nm",                   "FPAYHX-ZNM1S (UNESCO)",          "FPAYHX-ZNM1S (keep)",
         "KEEP", "V001", "Nm OK (name only, no truncation risk)", "Marlies row 1 Dbtr OK Nm", "Pablo", "OPEN", "", ""),
        ("SEPA_CT_UNES", "Dbtr", "Dbtr/PstlAdr/StrtNm",       "❌ NOT EMITTED today",            "FPAYHX-ZPFST",
         "ADD", "V001", "From 'UNESCO' → 'Place de Fontenoy'", "Marlies Tab 2 SEPA proposal", "Pablo", "OPEN", "", ""),
        ("SEPA_CT_UNES", "Dbtr", "Dbtr/PstlAdr/BldgNb",       "❌ NOT EMITTED today",            "FPAYHX (no source — skip or hardcode '7')",
         "ADD (optional)", "V001", "BldgNb '7' per Marlies proposal — skip if no reliable source", "Marlies Tab 2", "Pablo", "OPEN", "", ""),
        ("SEPA_CT_UNES", "Dbtr", "Dbtr/PstlAdr/PstCd",        "❌ NOT EMITTED today",            "FPAYHX-ZPLOR (parse) OR hardcode '75007'",
         "ADD", "V001", "PstCd '75007' per Marlies proposal", "Marlies Tab 2", "Pablo", "OPEN", "", ""),
        ("SEPA_CT_UNES", "Dbtr", "Dbtr/PstlAdr/TwnNm",        "❌ NOT EMITTED today (ORT1Z used for AdrLine)", "FPAYHX-ORT1Z OR hardcode 'PARIS'",
         "ADD (MANDATORY)", "V001", "TwnNm mandatory per CBPR+", "Marlies Tab 2 + handoff §2.1", "Pablo", "OPEN", "", ""),
        ("SEPA_CT_UNES", "Dbtr", "Dbtr/PstlAdr/Ctry",         "FPAYHX-LAND1 and FPAYHX-ZLISO",  "FPAYHX-ZLISO (keep)",
         "KEEP", "V001", "Ctry OK already", "Marlies row 1 Dbtr Ctry OK", "Pablo", "OPEN", "", ""),
        ("SEPA_CT_UNES", "Dbtr", "Dbtr/PstlAdr/AdrLine",      "FPAYHX-ORT1Z, ZPFST, ZPLOR (3 lines)", "SUPPRESS (or keep for Hybrid fallback)",
         "DEL", "V001", "Marlies: 'Not accepted — postal code missing'", "Marlies row 1 verdict", "Pablo", "OPEN", "", ""),

        ("SEPA_CT_UNES", "Cdtr", "Cdtr/Nm",                   "FPAYHX-ZNM1S (vendor)",          "FPAYHX-ZNM1S (keep)",
         "KEEP", "V001", "Cdtr Nm OK", "Marlies row 1 Cdtr Nm OK", "Pablo", "OPEN", "", ""),
        ("SEPA_CT_UNES", "Cdtr", "Cdtr/PstlAdr/StrtNm",       "❌ NOT EMITTED (tree has 0 StrtNm node in SEPA)", "FPAYHX-ZPFST OR event 05 fill",
         "ADD", "V001", "Structured street per Marlies proposal", "Marlies Tab 2 SEPA", "Pablo", "OPEN", "", ""),
        ("SEPA_CT_UNES", "Cdtr", "Cdtr/PstlAdr/PstCd",        "❌ NOT EMITTED",                  "FPAYHX-ZPLOR parsed (first 5 digits)",
         "ADD", "V001", "Postal split from ZPLOR", "Marlies Tab 2", "Pablo", "OPEN", "", ""),
        ("SEPA_CT_UNES", "Cdtr", "Cdtr/PstlAdr/TwnNm",        "❌ NOT EMITTED",                  "FPAYHX-ZPLOR parsed (after postal)",
         "ADD (MANDATORY)", "V001", "Town per CBPR+", "Marlies Tab 2 + CBPR+", "Pablo", "OPEN", "", ""),
        ("SEPA_CT_UNES", "Cdtr", "Cdtr/PstlAdr/Ctry",         "FPAYHX-LAND1 and FPAYHX-ZLISO",  "FPAYHX-ZLISO (keep)",
         "KEEP", "V001", "Ctry OK", "Marlies row 1 Cdtr Ctry OK", "Pablo", "OPEN", "", ""),
        ("SEPA_CT_UNES", "Cdtr", "Cdtr/PstlAdr/AdrLine",      "FPAYHX-ZPFST + ZPLOR (2 lines)",  "SUPPRESS in V001 (AdrLine only in V000 fallback)",
         "DEL", "V001", "Marlies: 'Not accepted — postal code should be in structured part'", "Marlies row 1 Cdtr verdict", "Pablo", "OPEN", "", ""),
        ("SEPA_CT_UNES", "CdtrAgt", "CdtrAgt/FinInstnId/PstlAdr",      "❌ NOT EMITTED (only BIC)",  "Decision — to confirm with SG TRM",
         "INVESTIGATE", "V001", "Marlies open question: 'Because of IBAN no need to enter bank addresses?'", "Marlies row 1 remark", "Pablo+Marlies", "OPEN", "", ""),

        # === CITI DC_V3_01 ===
        ("CITI/XML/UNESCO/DC_V3_01", "Dbtr", "Dbtr/PstlAdr/Ctry", "❌ MISSING today!",  "FPAYHX-ZLISO or hardcode 'FR'",
         "ADD (CRITICAL)", "V001", "Marlies: 'Unstructured, not accepted'", "Marlies row 4 Dbtr verdict", "Pablo", "OPEN", "", ""),
        ("CITI/XML/UNESCO/DC_V3_01", "Dbtr", "Dbtr/PstlAdr/StrtNm", "❌ NOT EMITTED",  "FPAYH-ZBSTR or FPAYHX via CITIPMW FREF",
         "ADD", "V001", "Structured street for UNESCO HQ", "Marlies row 4", "Pablo", "OPEN", "", ""),
        ("CITI/XML/UNESCO/DC_V3_01", "Dbtr", "Dbtr/PstlAdr/PstCd", "❌ NOT EMITTED",  "FPAYH-ZPSTL or hardcode '75007'",
         "ADD", "V001", "Postal for UNESCO HQ", "Marlies row 4", "Pablo", "OPEN", "", ""),
        ("CITI/XML/UNESCO/DC_V3_01", "Dbtr", "Dbtr/PstlAdr/TwnNm", "❌ NOT EMITTED (AdrLine has 'PARIS' + '75007 PARIS' + 'FRANCE')",  "FPAYH-ZORT1 or hardcode 'PARIS'",
         "ADD (MANDATORY)", "V001", "Town mandatory", "Marlies row 4 + CBPR+", "Pablo", "OPEN", "", ""),
        ("CITI/XML/UNESCO/DC_V3_01", "Dbtr", "Dbtr/PstlAdr/AdrLine (3 lines)", "FPAYHX-AUSTO/AUST2/AUST3 or FPAYH-ZNME3/4",  "SUPPRESS in V001",
         "DEL", "V001", "Marlies: 'Recommend structured, content doesn't change'", "Marlies row 4 verdict", "Pablo", "OPEN", "", ""),

        ("CITI/XML/UNESCO/DC_V3_01", "Cdtr", "Cdtr/PstlAdr/StrtNm", "FPAYH-ZBSTR",  "FPAYH-ZBSTR (keep)",
         "KEEP", "V001", "Cdtr Cdtr OK structured for most cases", "Marlies rows 4,5 Cdtr OK", "Pablo", "OPEN", "", ""),
        ("CITI/XML/UNESCO/DC_V3_01", "Cdtr", "Cdtr/PstlAdr/PstCd", "FPAYH-ZPSTL / FPAYHX-REF01", "Same, review MP_OFFSET",
         "REVIEW", "V001", "Check offset for non-Canada cases", "Probe found FPAYP-REF01 used for multiple nodes", "Pablo", "OPEN", "", ""),
        ("CITI/XML/UNESCO/DC_V3_01", "Cdtr", "Cdtr/PstlAdr/TwnNm", "FPAYH-ZORT1 / FPAYHX-ORT1Z", "FPAYH-ZORT1 (keep)",
         "KEEP", "V001", "OK", "Marlies row 5 Cdtr OK", "Pablo", "OPEN", "", ""),
        ("CITI/XML/UNESCO/DC_V3_01", "Cdtr", "Cdtr/PstlAdr/Ctry", "FPAYH-ZLAND / FPAYHX-ZLISO (multiple)", "Keep",
         "KEEP", "V001", "OK", "Marlies rows 4,5 Cdtr OK", "Pablo", "OPEN", "", ""),

        ("CITI/XML/UNESCO/DC_V3_01", "Cdtr", "Cdtr/PstlAdr/AdrLine (WHO case)", "FPAYH-ZNME2 (2nd line of name)", "REVIEW — may indicate vendor master DQ issue",
         "DATA-QUALITY", "V001", "Marlies row 4: 'AdrLine contains RAVEL — 2nd line of name vendor 904874'", "Marlies row 4 Cdtr remark", "Pablo+MasterData", "OPEN", "", ""),

        ("CITI/XML/UNESCO/DC_V3_01", "UltmtCdtr", "UltmtCdtr/PstlAdr (WHO→ICC)", "FPAYHX-UBSTR/UBORT/UBISO (our probe) / FPAYH-Z* for intermediary", "TBD — Q3 BLOCKED",
         "BLOCKED", "V001 or V002", "Marlies: 'Why is this not in structured?'", "Marlies row 4 UltmtCdtr + handoff Q3", "Citi TRM + DBS", "BLOCKED", "", ""),

        ("CITI/XML/UNESCO/DC_V3_01", "Dbtr", "Dbtr (Brazil UBO case)", "FPAYHX-ZPFST/ZPLOR (hybrid Brazil)", "Restructure: separate ZPLOR → PstCd + TwnNm",
         "MOD", "V001", "Marlies row 7: 'Postal code should be in structured part'", "Marlies row 7 (CITI BR)", "Pablo", "OPEN", "", ""),

        # === CGI_XML_CT_UNESCO ===
        ("CGI_XML_CT_UNESCO", "Dbtr", "Dbtr/PstlAdr/StrtNm", "FPAYP-REF01 (offset 0)", "KEEP — already structured",
         "KEEP", "V001", "Marlies row 8: 'Structured OK'", "Marlies row 8 Dbtr OK", "Pablo", "OPEN", "", ""),
        ("CGI_XML_CT_UNESCO", "Dbtr", "Dbtr/PstlAdr/BldgNb", "FPAYP-REF01 (offset N)", "KEEP",
         "KEEP", "V001", "Structured already", "Marlies row 8 Dbtr OK", "Pablo", "OPEN", "", ""),
        ("CGI_XML_CT_UNESCO", "Dbtr", "Dbtr/PstlAdr/PstCd", "FPAYH-ZPSTL / FPAYP-REF01", "KEEP",
         "KEEP", "V001", "Already OK", "Marlies row 8 Dbtr OK", "Pablo", "OPEN", "", ""),
        ("CGI_XML_CT_UNESCO", "Dbtr", "Dbtr/PstlAdr/TwnNm", "FPAYH-ZORT1 / FPAYP-BORT1/ORT01", "KEEP",
         "KEEP", "V001", "Already OK", "Marlies row 8 Dbtr OK", "Pablo", "OPEN", "", ""),
        ("CGI_XML_CT_UNESCO", "Dbtr", "Dbtr/PstlAdr/Ctry", "FPAYP-LANDL / FPAYP-LAND1 / FPAYHX-LAND1", "KEEP",
         "KEEP", "V001", "Already OK", "Marlies row 8 Dbtr OK", "Pablo", "OPEN", "", ""),

        ("CGI_XML_CT_UNESCO", "CdtrAgt", "CdtrAgt/FinInstnId/PstlAdr/AdrLine", "FPAYH-ZNME* / FPAYHX-UBSTR/UBORT", "SUPPRESS in V001",
         "DEL", "V001", "Marlies: 'Not accepted, recommend structured with standard bank DB'", "Marlies rows 8,9,10 CdtrAgt verdict", "Pablo", "OPEN", "", ""),
        ("CGI_XML_CT_UNESCO", "CdtrAgt", "CdtrAgt/FinInstnId/PstlAdr/Ctry", "FPAYHX-UBISO / FPAYHX-LDISO / FPAYH-BNKS1", "KEEP",
         "KEEP", "V001", "Ctry already OK in most cases", "Marlies probe", "Pablo", "OPEN", "", ""),
        ("CGI_XML_CT_UNESCO", "CdtrAgt", "CdtrAgt/FinInstnId/PstlAdr/StrtNm", "❌ NOT EMITTED today", "FPAYHX-ZBSTR or BNKA read",
         "ADD", "V001", "Replace AdrLine with structured", "Marlies rows 8,9,10 proposal", "Pablo", "OPEN", "", ""),
        ("CGI_XML_CT_UNESCO", "CdtrAgt", "CdtrAgt/FinInstnId/PstlAdr/TwnNm", "❌ NOT EMITTED today", "FPAYHX-ZBORT or BNKA read",
         "ADD (MANDATORY)", "V001", "Mandatory per CBPR+", "Marlies + CBPR+", "Pablo", "OPEN", "", ""),
        ("CGI_XML_CT_UNESCO", "CdtrAgt", "CdtrAgt/FinInstnId/PstlAdr/PstCd", "❌ NOT EMITTED today", "BNKA read or derived",
         "ADD", "V001", "Recommended", "Marlies + CBPR+", "Pablo", "OPEN", "", ""),

        ("CGI_XML_CT_UNESCO", "Cdtr", "Cdtr/PstlAdr/*", "FPAYP-BSTRAS / BORT1 / LAND1 / REGIO (structured)", "KEEP",
         "KEEP", "V001", "Marlies row 8: 'Structured, correct no need to change'", "Marlies rows 8,9,10 Cdtr OK", "Pablo", "OPEN", "", ""),

        ("CGI_XML_CT_UNESCO", "Dbtr (Treasury Transfer case)", "Dbtr/PstlAdr", "Only FPAYHX-ZLISO emitted (Ctry='FR')",
         "Add TwnNm + possibly full structured; investigate XSTRA='' flag behavior",
         "INVESTIGATE", "V001", "Marlies row 11: 'Postal code missing, why different from 3rd party?'", "Marlies row 11 + T042Z XSTRA='' for FR+A", "Pablo", "OPEN", "", ""),

        # === CGI_XML_CT_UNESCO_1 ===
        ("CGI_XML_CT_UNESCO_1", "ALL nodes", "Mirror of CGI_XML_CT_UNESCO", "Same as parent",  "Apply same changes as CGI_XML_CT_UNESCO V001",
         "SYNC", "V001", "Same FIRSTNODE_ID (N_5649723900) as parent — share changes", "Phase 0 Finding on FIRSTNODE_ID equality", "Pablo", "OPEN", "", ""),
    ]
    for r in rows:
        ws.append(r)
    style_header(ws)
    autosize(ws, max_width=50)
    # Row heights
    for r in ws.iter_rows(min_row=2):
        ws.row_dimensions[r[0].row].height = 45
    # Borders + alt colors
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = CELL_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER_THIN
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL
    # Freeze top + first col
    ws.freeze_panes = "B2"
    return ws


def add_dmee_headers(wb):
    ws = wb.create_sheet("05 DMEE Headers")
    src = PHASE0 / "dmee_full" / "dmee_tree_head_p01_full.csv"
    if not src.exists():
        return
    with open(src, encoding="utf-8") as f:
        for row in csv.reader(f):
            ws.append(row)
    style_header(ws)
    autosize(ws, max_width=60)
    ws.freeze_panes = "B2"


def add_address_nodes(wb):
    ws = wb.create_sheet("06 Address Nodes")
    src = PHASE0 / "dmee_full" / "dmee_tree_node_p01_all48.csv"
    if not src.exists():
        # fallback
        src = PHASE0 / "gap006_dmee_nodes_with_exit.csv"
    if not src.exists():
        return
    with open(src, encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)
        ws.append(header)
        # Filter to address+party nodes only
        address_kw = ['StrtNm','BldgNb','PstCd','TwnNm','Ctry','AdrLine','PstlAdr','BldgNmOrId','CtrySubDvsn','CtryOfRes','Dbtr','Cdtr','UltmtCdtr','CdtrAgt','UltmtDbtr']
        tech_idx = header.index("TECH_NAME") if "TECH_NAME" in header else None
        for row in reader:
            if tech_idx is not None and any(kw in (row[tech_idx] or "") for kw in address_kw):
                ws.append(row)
    style_header(ws)
    autosize(ws, max_width=30)
    ws.freeze_panes = "B2"


def add_conditions(wb):
    ws = wb.create_sheet("07 Conditions")
    src = PHASE0 / "dmee_full" / "dmee_tree_cond_p01.csv"
    if not src.exists():
        return
    with open(src, encoding="utf-8") as f:
        for row in csv.reader(f):
            ws.append(row)
    style_header(ws)
    autosize(ws, max_width=25)
    ws.freeze_panes = "B2"


def add_d01_vs_p01(wb):
    ws = wb.create_sheet("08 D01 vs P01")
    src = PHASE0 / "d01_vs_p01_inventory.csv"
    if not src.exists():
        return
    with open(src, encoding="utf-8") as f:
        for row in csv.reader(f):
            ws.append(row)
    style_header(ws)
    autosize(ws, max_width=40)
    ws.freeze_panes = "B2"
    # Highlight P01_ONLY rows red
    for r in ws.iter_rows(min_row=2):
        state_cell = r[2] if len(r) > 2 else None  # state column
        if state_cell and state_cell.value == "P01_ONLY":
            for c in r:
                c.fill = PatternFill("solid", fgColor="F5CACA")


def add_vendor_dq(wb):
    ws = wb.create_sheet("09 Vendor DQ")
    ws.append(["Metric", "Count", "% of active", "Notes"])
    ws.append(["Active LFA1 vendors (LOEVM blank)", 111241, "100.000%", "Source: Gold DB snapshot 2026-04"])
    ws.append(["Active vendors with ADRC join", 111238, "99.997%", ""])
    ws.append(["Missing CITY1 OR COUNTRY (MANDATORY)", 5, "0.005%", "HARD-FAIL post-Nov-2026 — manual cleanup required pre-go-live"])
    ws.append(["Missing CITY1", 1, "0.001%", ""])
    ws.append(["Missing COUNTRY", 4, "0.004%", ""])
    ws.append(["Missing STREET", 5570, "5.01%", "Recommended, not mandatory — Hybrid acceptable"])
    ws.append(["Missing POST_CODE1", 1153, "1.04%", "Recommended"])
    ws.append(["Missing HOUSE_NUM1 (BldgNb)", 76574, "68.84%", "Optional — no rejection risk"])
    ws.append([])
    ws.append(["The 5 vendors to fix — query to run in Phase 1:"])
    ws.append(["SELECT l.LIFNR, l.NAME1, a.CITY1, a.COUNTRY, a.STREET, a.POST_CODE1"])
    ws.append(["FROM LFA1 l JOIN ADRC a ON l.ADRNR = a.ADDRNUMBER"])
    ws.append(["WHERE (a.CITY1 IS NULL OR a.CITY1 = '' OR a.COUNTRY IS NULL OR a.COUNTRY = '')"])
    ws.append(["AND (l.LOEVM IS NULL OR l.LOEVM = '')"])
    ws.append(["ORDER BY l.LIFNR;"])
    style_header(ws)
    autosize(ws, max_width=80)
    # Highlight mandatory row
    for r in ws.iter_rows(min_row=4, max_row=4):
        for c in r:
            c.fill = PatternFill("solid", fgColor="F5CACA")
            c.font = Font(name="Segoe UI", size=10, bold=True)


def add_francesco_audit(wb):
    ws = wb.create_sheet("10 Francesco Audit")
    ws.append(["TRKORR", "Date", "Type", "Status", "Tree touched", "obj_count", "Classification", "Notes"])
    rows = [
        ("D01K9B0CZ0", "2025-03-20", "Workbench", "R", "/CGI_XML_CT_UNESCO", 3, "ASSIST", "PMF variant VC_TFPM042F config (Event 05 registration for CGI_XML_CT_UNESCO_BK bank variant)"),
        ("D01K9B0CWS", "2025-03-07", "Workbench", "R", "/CGI_XML_CT_UNESCO_1", 2, "IRRELEVANT", "Small variant config"),
        ("D01K9B0CUS", "2025-02-21", "Workbench", "R", "/CGI_XML_CT_UNESCO_1", 1, "IRRELEVANT", "Small variant config"),
        ("D01K9B0CUT", "2025-02-21", "Workbench", "R", "/CGI_XML_CT_UNESCO_1", 2, "IRRELEVANT", "Small variant config"),
        ("D01K9B0CTP", "2025-02-20", "Workbench", "R", "/CGI_XML_CT_UNESCO_1", 3, "ASSIST", "PMF variant VC_TFPM042F config"),
    ]
    for r in rows:
        ws.append(r)
    ws.append([])
    ws.append(["Summary"])
    ws.append(["Total FP_SPEZZANO transports 2017-2026: 70"])
    ws.append(["DMEE-touching transports: 5 (all 2025 Q1, all CGI trees)"])
    ws.append(["Conclusion: Francesco's work is PMF variant config (VC_TFPM042F), NOT tree node structure changes. User directive 'no sabía del proceso' preserved — structured-address migration NOT blocked by his work. Courtesy alignment call recommended before Phase 2 edits."])
    style_header(ws)
    autosize(ws, max_width=70)


def add_q1_q8(wb):
    ws = wb.create_sheet("11 Q1-Q8 Status")
    headers = ["Q#", "Question (handoff §11)", "Priority", "Status 2026-04-24", "Owner", "Resolution / notes"]
    ws.append(headers)
    rows = [
        ("Q1", "Exact DMEE tree name for CITI format in PRD vs DEV", "HIGH", "CLOSED", "Pablo", "Session #039 + Phase 0: /CITI/XML/UNESCO/DC_V3_01 confirmed in P01 via RFC"),
        ("Q2", "CITI pain.001 variant tags (proprietary?)", "HIGH", "OPEN", "Marlies + Citi TRM", "Phase 1 TRM outreach"),
        ("Q3", "Worldlink UltmtCdtr beneficiary address source", "HIGH", "BLOCKED", "Citi TRM + DBS", "Blocks Phase 2 Step 7 (CITI UltmtCdtr). Defer to V002 if unresolved by Phase 2 start."),
        ("Q4", "IIEP/UIL separate PMW assignments in FBZP?", "MEDIUM", "OPEN", "Pablo", "Phase 1 FBZP check"),
        ("Q5", "UBO/UIS separate OBPM4 Event 05 registrations?", "MEDIUM", "PARTIAL", "Pablo", "Phase 0 probe: TFPM042FB shows registration per FORMI, not per co code. Confirmed for CGI (FI_PAYMEDIUM_DMEE_CGI_05 serves all variants). To re-check CITI scoping."),
        ("Q6", "Vendor master DQ count", "HIGH", "CLOSED", "Phase 0", "5/111,241 missing mandatory (0.005%). Risk LOW."),
        ("Q7", "Formal bank deadline communications", "MEDIUM", "OPEN", "Treasury + Marlies", "Phase 1 written confirmations from each bank TRM"),
        ("Q8", "SEPA consolidation into CGI_XML_CT?", "LOW", "PARKED", "Pablo + Marlies", "Post-go-live decision, not in scope"),
    ]
    for r in rows:
        ws.append(r)
    style_header(ws)
    autosize(ws, max_width=60)
    # Color status col
    for r in ws.iter_rows(min_row=2):
        status_cell = r[3] if len(r) > 3 else None
        if status_cell and status_cell.value in STATUS_COLORS:
            status_cell.fill = PatternFill("solid", fgColor=STATUS_COLORS[status_cell.value])
            status_cell.font = Font(color="FFFFFF", bold=True)


def add_bank_spec(wb):
    ws = wb.create_sheet("12 Bank Spec Matrix")
    ws.append(["Bank", "Payment channel", "Tree", "Strictness required", "Hybrid acceptable?", "Test gateway available?", "TRM contact", "Contact method", "Response received", "Response date", "Notes"])
    rows = [
        ("Société Générale", "SEPA (EUR)", "/SEPA_CT_UNES", "TBD", "TBD", "TBD", "Marlies's SG TRM", "Phone/Teams", "PENDING", "", ""),
        ("Société Générale", "Non-SEPA (multi-currency)", "/CGI_XML_CT_UNESCO", "TBD", "TBD", "TBD", "Marlies's SG TRM (same)", "Phone/Teams", "PENDING", "", ""),
        ("Citibank", "USD Worldlink", "/CITI/XML/UNESCO/DC_V3_01", "TBD", "TBD", "CitiConnect test gateway likely available", "Marlies's Citi TRM", "CitiConnect portal / email", "PENDING", "", "Ask about Worldlink UltmtCdtr for Q3"),
        ("Shinhan Bank KR", "CGI KRW", "/CGI_XML_CT_UNESCO", "TBD", "TBD", "TBD", "via Marlies", "Email", "PENDING", "", ""),
        ("Metro Bank GB", "CGI GBP", "/CGI_XML_CT_UNESCO", "TBD", "TBD", "TBD", "via Marlies", "Email", "PENDING", "", ""),
        ("Guaranty Trust Bank NG", "CGI NGN", "/CGI_XML_CT_UNESCO", "TBD", "TBD", "TBD", "via Marlies", "Email", "PENDING", "", ""),
    ]
    for r in rows:
        ws.append(r)
    ws.append([])
    ws.append(["Questions to each TRM:"])
    ws.append(["1. Is fully-structured address required, or is Hybrid (TwnNm+Ctry + AdrLine) acceptable post-Nov-2026?"])
    ws.append(["2. Do you support a test gateway where we can send a sample file for validation without production impact?"])
    ws.append(["3. What XSD version does your parser expect? (pain.001.001.03 / .09)"])
    ws.append(["4. Any bank-specific extensions or proprietary tags in the PstlAdr block?"])
    ws.append(["5. Provide formal deadline confirmation for CBPR+ structured address enforcement."])
    style_header(ws)
    autosize(ws, max_width=40)


def add_sap_notes(wb):
    ws = wb.create_sheet("13 SAP Notes")
    ws.append(["Note #", "Title", "Priority", "Relevance", "Fetch owner", "Status"])
    rows = [
        ("1665873", "CGI_XML_CT format introduction guide (59 pages)", "P1 — CANONICAL", "Full DMEE CGI setup + PSTLADRMOR behavior", "Pablo", "To fetch from SAP for Me"),
        ("2795667", "ISO 20022 adoption — global harmonization (SEPA)", "P2", "SEPA compliance migration context", "Pablo", "To fetch"),
        ("2668719", "PMW format lifecycle", "P2", "Managing format versions — related to our V000/V001 strategy", "Pablo", "To fetch"),
        ("2819590", "Structured remittance info gap in CGI XML", "P3", "Check applicability — may explain RMTINF gaps", "Pablo", "To fetch if relevant post-probe"),
        ("2845063", "CGI identifications and organization ID config", "P3", "CGI Creditor ID configuration", "Pablo", "To fetch"),
        ("2944738", "pain.001.001.09 schema migration", "P4", "Parallel schema-version upgrade (out of current scope)", "Pablo", "Defer"),
        ("3208888", "CBPR+ structured address", "P2 (if exists)", "Direct mandate documentation", "Pablo", "Search + confirm"),
    ]
    for r in rows:
        ws.append(r)
    style_header(ws)
    autosize(ws, max_width=55)


def add_phase_plan(wb):
    ws = wb.create_sheet("14 Phase Plan")
    ws.append(["Phase", "Period", "Status", "Owner", "Key deliverables", "Gate criteria"])
    rows = [
        ("0 — Discovery", "2026-04-24 → 04-27", "COMPLETED", "Pablo",
         "100% code inventory · Findings A-L · Brain updates · Companion v1 · 4 commits",
         "xml_touch_points_complete.md signed ✅"),
        ("1 — Matrix + Specs", "2026-04-27 → 04-30", "PENDING", "Pablo + Marlies + TRM",
         "change_matrix.csv signed · bank spec per-bank · D01 retrofit plan · Q3 resolution · DQ 5 vendors list",
         "0 UNKNOWN in Change Matrix + bank specs received"),
        ("2 — Config D01", "May 2026 (4 weeks)", "PENDING", "Pablo + N_MENARD",
         "V001 created per tree · Pattern A BAdI fix · DMEE simulations pass",
         "Each tree transport reviewed by Marlies + N_MENARD (Pattern A)"),
        ("3 — Unit Test", "June 2026", "PENDING", "Pablo + DBS + Marlies",
         "UT-01 to UT-08 passed · XSD validation 100% · BAdI regression green",
         "test_results_june.md signed"),
        ("4 — UAT V01", "July 2026", "PENDING", "Marlies + Finance leads",
         "UAT-01 to UAT-12 passed · Bank test-gateway acceptance per bank",
         "All 3 banks + entity leads sign UAT close-out"),
        ("5 — Deploy P01", "Aug-Nov 2026", "PENDING", "DBS + Pablo + Marlies",
         "Staged V001 activation per tree · 14-day monitoring · V000 decommission at +30d",
         "CBPR+ Nov 2026 compliance · 0 rollbacks ideal"),
    ]
    for r in rows:
        ws.append(r)
    style_header(ws)
    autosize(ws, max_width=60)
    # Color status
    for r in ws.iter_rows(min_row=2):
        sc = r[2] if len(r) > 2 else None
        if sc and sc.value in STATUS_COLORS:
            sc.fill = PatternFill("solid", fgColor=STATUS_COLORS[sc.value])
            sc.font = Font(color="FFFFFF", bold=True)


def main():
    wb = Workbook()
    # Remove default sheet
    del wb["Sheet"]

    add_overview(wb)
    copy_marlies_analysis(wb)
    add_change_matrix(wb)
    add_dmee_headers(wb)
    add_address_nodes(wb)
    add_conditions(wb)
    add_d01_vs_p01(wb)
    add_vendor_dq(wb)
    add_francesco_audit(wb)
    add_q1_q8(wb)
    add_bank_spec(wb)
    add_sap_notes(wb)
    add_phase_plan(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Built: {OUT}")
    print(f"Size: {OUT.stat().st_size:,} bytes")
    print(f"Sheets: {len(wb.sheetnames)}")
    for i, s in enumerate(wb.sheetnames, 1):
        ws = wb[s]
        print(f"  {i:>2}. {s} ({ws.max_row} rows × {ws.max_column} cols)")


if __name__ == "__main__":
    main()
