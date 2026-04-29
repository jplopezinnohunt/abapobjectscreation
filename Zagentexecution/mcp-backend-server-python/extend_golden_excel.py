"""
Extend the realigned Golden Excel:
1. Add 5 new execution-tracking columns (Plan, NEEDED, DONE, IN CHARGE, WHO)
2. Append 29 additional items as 8 new phases (P21-P28) covering HR, AD,
   external systems, document management, etc.
"""
import sys
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from copy import copy

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

SRC = 'Zagentexecution/Golden_live_matrix_2026-04-16_realigned.xlsx'
DST = 'Zagentexecution/Golden_live_matrix_2026-04-16_realigned.xlsx'

wb = load_workbook(SRC)
ws = wb['Golden × Live']

# Current columns: A..S (19 cols)
# A=Phase, B=Item, C=System, D=Action, E=TCode, F=Type, G=Dep, H=Tables,
# I-R = 10 institutes, S=Comments
# Will add T=Plan, U=NEEDED, V=DONE, W=IN CHARGE, X=WHO

# Header styling
hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
hdr_font = Font(color='FFFFFF', bold=True)
phase_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
new_phase_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')

# Add new column headers
new_cols = ['Plan', 'NEEDED (Y/N)', 'DONE (Y/N)', 'IN CHARGE', 'WHO (Responsible + Backup)']
for i, name in enumerate(new_cols, start=20):  # T=20, U=21, V=22, W=23, X=24
    cell = ws.cell(row=1, column=i, value=name)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Set column widths
ws.column_dimensions['T'].width = 12
ws.column_dimensions['U'].width = 8
ws.column_dimensions['V'].width = 12
ws.column_dimensions['W'].width = 14
ws.column_dimensions['X'].width = 22

# Pre-fill existing 75 items with sensible defaults from current state
# Map (item_number) → (plan, needed, done, in_charge, who)
KNOWN_STATUS = {
    # Phase 1 Foundation
    1: (None, 'Y', 'Done', 'DBS', 'P.Lopez'),
    2: (None, 'Y', 'N', 'DBS', 'P.Lopez'),
    3: (None, 'Y', 'Done', 'DBS', 'P.Lopez'),
    4: (None, 'Y', 'Done', 'DBS', 'P.Lopez'),  # SKB1 aligned 2026-04-15
    5: (None, 'Y', 'Done', 'DBS', 'I. Konakov'),  # FM area
    8: (None, 'Y', 'In Progress', 'DBS', 'I. Konakov'),  # Commitment items
    9: (None, 'Y', 'Done', 'DBS', 'P.Lopez'),  # CO area
    10: (None, 'Y', 'Done', 'DBS', 'P.Lopez'),  # CO versions
    11: (None, 'Y', 'Done', 'DBS', 'P.Lopez'),  # Cost elements (CSKB aligned)
    12: (datetime.datetime(2026,4,10), 'Y', 'In Progress', 'DBS', 'R. Rios'),  # PS profiles
    # Phase 4 ValSub
    29: (None, 'Y', 'In Progress', 'DBS', 'I. Konakov'),  # Create subst rule
    30: (None, 'Y', 'In Progress', 'DBS', 'I. Konakov'),  # Assign OBBH
    33: (None, 'Y', 'Todo', 'DBS', 'I. Konakov'),  # RGUGBR00
    # Phase 17 Auth Roles
    64: (datetime.datetime(2026,4,10), 'Y', 'In Progress', 'DBS', 'P. Ikouna'),  # Clone Y_*_*
    # Phase 19 Activation
    69: (None, 'Y', 'Todo', 'FIN/FAS/FRA', 'J.La'),  # Open periods
    70: (None, 'Y', 'In Progress', 'DBS', 'P.Lopez'),  # Number ranges
}

# Walk all data rows and apply pre-fills
for r_idx in range(2, ws.max_row + 1):
    item = ws.cell(row=r_idx, column=2).value
    if item is None:
        continue  # group header
    if item in KNOWN_STATUS:
        plan, needed, done, who_dept, who_person = KNOWN_STATUS[item]
        ws.cell(row=r_idx, column=20, value=plan)
        ws.cell(row=r_idx, column=21, value=needed)
        ws.cell(row=r_idx, column=22, value=done)
        ws.cell(row=r_idx, column=23, value=who_dept)
        ws.cell(row=r_idx, column=24, value=who_person)
    else:
        # Default: Needed=Y, status blank
        ws.cell(row=r_idx, column=21, value='Y')

# Find next available row for appending new phases
next_row = ws.max_row + 2

# 29 NEW ITEMS (P21-P28) — based on user's "Current list" tab
# Format: (phase, system, action, tcode, type, dep, tables, comments, plan, needed, done, in_charge, who)
NEW_ITEMS = [
    # ---- P21 Master Data Conversion ----
    ('__HDR__', 'P21 Master Data Conversion'),
    ('P21 Conversion', 'Balance Conversion', 'Balance conversion (existing movements)', '', 'Conversion', '1', '',
     'Business Activity. If institute has historical movements, decide migration approach', None, 'Y', '', '', ''),
    ('P21 Conversion', 'AP/AR', 'Open items AP/AR migration', '', 'Conversion', '1,4', '',
     'Business Activity. Convert open AP/AR items', None, 'Y', '', '', ''),
    # ---- P22 External SAP Master Data Sync ----
    ('__HDR__', 'P22 External SAP Master Data Sync'),
    ('P22 ExtSync', 'MOUV', 'Synchronize SAP Master Data', '', 'Sync', '1', '',
     '1 day', None, 'Yes', '', '', 'S. Dartigolle, R.Rios'),
    ('P22 ExtSync', 'CMT', 'Synchronize SAP Master Data', '', 'Sync', '1', '',
     '1 day', None, 'Yes', '', 'DBS', 'D. Andros'),
    # ---- P23 HCM / HR / Travel ----
    ('__HDR__', 'P23 HCM / HR / Travel'),
    ('P23 HCM', 'SAP HCM', 'HCM customizing tables (T778A/V/U, T527O, T9SCT) + Infotypes 1000/1001/1081/00001', '', 'Config', '1', 'T778A,T778V,T778U,T527O,T9SCT',
     'Business should provide institute org structure', datetime.datetime(2026,4,3), 'Yes', '', 'DBS', 'A. Sefiani'),
    ('P23 HCM', 'SAP TV', 'TV customizing', '', 'Config', '1', '',
     '1 day', datetime.datetime(2026,4,3), 'Y', '', 'DBS', 'Laia'),
    ('P23 HCM', 'SAP-ROLE HR', 'Clone Y_*_HR roles + assign + block old', 'PFCG', 'Role', '4', '',
     'Explore copy from MGIE roles. 5 days', datetime.datetime(2026,4,10), 'Yes', '', 'DBS', 'P. Ikouna'),
    ('P23 HCM', 'Org Structure', 'Define Institute Org Structure', '', 'Setup', '1', '',
     'Create new or transfer HR Workflow', datetime.datetime(2026,4,15), 'Y', '', 'HRM', ''),
    ('P23 HCM', 'PA', 'Personal Administration — transfer staff to new co code', '', 'Master Data', '1', '',
     '*** MUST BE DONE BEFORE 17/04 ***', None, 'Y', '', 'HRM', ''),
    ('P23 HCM', 'HR-AP', 'Initial Vendor Creation (HR review)', '', 'Master Data', '1', '',
     'Business activity. Automatic HR review', None, 'Y', '', 'DBS FIN', ''),
    ('P23 HCM', 'Core HR+', 'Core HR setup', '', 'Setup', '1', '',
     'Not a blocker for Financial Execution', None, 'Yes', '', 'DBS', 'Adil, Laia'),
    # ---- P24 Identity & Access ----
    ('__HDR__', 'P24 Identity & Access'),
    ('P24 Identity', 'Active Directory', 'Create new accounts + AD groups', '', 'Setup', '1', '',
     'UNESDIR may auto-create following naming convention', None, 'Yes', '', 'DBS/OSM/NSO', 'C. Souvanthong'),
    ('P24 Identity', 'Role Mgmt', 'Create Places', '', 'Config', '1', '',
     '', None, 'Yes', '', 'DBS', 'Dimitri'),
    ('P24 Identity', 'Role Mgmt', 'Create users with appropriate roles under new places', '', 'Config', '1', '',
     'Verify Payment approval invoice roles', None, 'Yes', '', 'FIN', ''),
    # ---- P25 Field Office Planning ----
    ('__HDR__', 'P25 Field Office Planning'),
    ('P25 Planning', 'Core Planner', 'Set up Core Planner', '', 'Config', '1', '',
     'Once done, proceed with items 21,22 if needed', None, 'Yes', '', 'DBS', 'A. Mercea'),
    ('P25 Planning', 'Core Manager', 'Set up Core Manager', '', 'Config', '1', '',
     '', None, 'No', '', 'DBS', 'A. Mercea'),
    # ---- P26 Workflow / Sector Updates ----
    ('__HDR__', 'P26 Workflow / Sector Updates'),
    ('P26 Workflow', 'HR Workflow', 'Add new Sector', '', 'Config', '1', '',
     '', None, 'No', '', 'DBS', 'Naila'),
    ('P26 Workflow', 'TULIP', 'Add new Sector', '', 'Config', '1', '',
     '', None, 'No', '', 'DBS', 'Naila'),
    ('P26 Workflow', 'PROSPER', 'BFM impact actions', '', 'Config', '1', '',
     '', None, 'No', '', 'DBS', 'Naila'),
    ('P26 Workflow', 'INFOSERV', 'In UNESDIR, modify Cost Center associated to Sector/Service', '', 'Config', '1', '',
     'To be confirmed', None, 'TBD', '', 'DBS/IOP/TSU', 'Naila / Alessandra'),
    ('P26 Workflow', 'REC', 'Update & review Application Role name + republish web app', '', 'Config', '1', '',
     '', None, 'No', '', 'DBS', 'S. Dartigolle'),
    # ---- P27 Inventory & Procurement ----
    ('__HDR__', 'P27 Inventory & Procurement'),
    ('P27 Inventory', 'Inventory Management', 'Inventory Management setup', '', 'Setup', '1', '',
     'To conclude, need Org. STEM', None, 'Yes', 'In Progress', 'DBS', 'R. Rios'),
    # ---- P28 Communications & Documents ----
    ('__HDR__', 'P28 Communications & Documents'),
    ('P28 Comms', 'COUR(R)IER', 'Set exceptions to separate divisions for re-invoicing', '', 'Config', '1', '',
     '', None, 'Yes', '', 'DBS', 'C.Michiote'),
    ('P28 Comms', 'UNTERM', 'Update Terminology database', '', 'Config', '1', '',
     'Check if Reference is updated', None, 'Yes', '', 'CLD/T', ''),
    ('P28 Comms', 'Smart', 'Configure Institute options', '', 'Config', '1', '',
     '', None, 'Yes', '', 'DBS', 'Narmin'),
    ('P28 Comms', 'SHAREPOINT', 'Rename SharePoint group or create new', '', 'Config', '1', '',
     '', None, 'Yes', '', 'DBS', 'C.Michiote'),
    ('P28 Comms', 'UNESDOC', 'Check that replication job takes new name into account', '', 'Config', '1', '',
     'For FO/institute', None, 'No', '', 'DBS', '?'),
    ('P28 Comms', 'WATCHDOC', 'Implement specific groups for re-invoicing', '', 'Config', '1', '',
     '', None, 'Yes', '', 'DBS', 'R. Lévy'),
    ('P28 Comms', 'OURDRIVE', 'Implement the new Model Filing Plan', '', 'Config', '1', '',
     '', None, 'Yes', '', 'DBS', 'Imad'),
]

# Append to Excel
current_item = 76  # next item number after 75
for entry in NEW_ITEMS:
    if entry[0] == '__HDR__':
        # Group header row
        ws.cell(row=next_row, column=1, value=entry[1])
        for c in range(1, 25):
            ws.cell(row=next_row, column=c).fill = new_phase_fill
            ws.cell(row=next_row, column=c).font = Font(bold=True)
        next_row += 1
        continue
    phase, system, action, tcode, typ, dep, tables, comments, plan, needed, done, in_charge, who = entry
    ws.cell(row=next_row, column=1, value=phase)
    ws.cell(row=next_row, column=2, value=current_item)
    ws.cell(row=next_row, column=3, value=system)
    ws.cell(row=next_row, column=4, value=action)
    ws.cell(row=next_row, column=5, value=tcode)
    ws.cell(row=next_row, column=6, value=typ)
    ws.cell(row=next_row, column=7, value=dep)
    ws.cell(row=next_row, column=8, value=tables)
    # Per-institute columns I-R: leave blank (no transport signature for these new items)
    for c in range(9, 19):
        ws.cell(row=next_row, column=c, value='')
    ws.cell(row=next_row, column=19, value=comments)
    ws.cell(row=next_row, column=20, value=plan)
    ws.cell(row=next_row, column=21, value=needed)
    ws.cell(row=next_row, column=22, value=done)
    ws.cell(row=next_row, column=23, value=in_charge)
    ws.cell(row=next_row, column=24, value=who)
    current_item += 1
    next_row += 1

# Re-extend filter
ws.auto_filter.ref = f'A1:X{ws.max_row}'

wb.save(DST)
print(f'Saved: {DST}')
print(f'  Total rows: {ws.max_row}')
print(f'  Total columns: 24 (added Plan, NEEDED, DONE, IN CHARGE, WHO)')
print(f'  New items added: {current_item - 76}  (items 76..{current_item - 1})')
print(f'  New phases: P21-P28 (8 phases)')
