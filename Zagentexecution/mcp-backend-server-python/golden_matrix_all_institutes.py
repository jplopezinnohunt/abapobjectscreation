"""
Expand the 76-step Golden matrix to ALL 9 UNESCO institutes + STEM.

For each step, check if its PRIMARY tables have rows for each institute's
BUKRS/KOKRS/FIKRS key. This is LIVE STATE (not "was it created via transport")
which is actually more useful — answers "does this institute have this config?".

Data sources:
- Gold DB (fast) for tables already extracted
- D01 live RFC for the rest (when connected)
"""
import sys
import json
import sqlite3
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

HERE = Path(__file__).resolve().parent
GOLD_DB = HERE.parent / 'sap_data_extraction' / 'sqlite' / 'p01_gold_master_data.db'

INSTITUTES = ['UNES', 'IBE', 'IIEP', 'UBO', 'UIL', 'UIS', 'ICTP', 'ICBA', 'MGIE', 'STEM']


# --- Table → (key column, gold-db-table-name-or-None) ---
# For BUKRS-keyed tables at UNESCO KOKRS=BUKRS and FIKRS=BUKRS, so we can query on that
TABLE_META = {
    # Company code
    'T001': ('BUKRS', 'T001'),
    'T001Q': ('BUKRS', None),
    'T001A': ('BUKRS', None),
    'T001B': ('BUKRS', None),
    'T001W': ('BUKRS', None),
    # House banks
    'T012': ('BUKRS', 'T012'),
    'T012D': ('BUKRS', None),
    'T012K': ('BUKRS', 'T012K'),
    'T012T': ('BUKRS', 'T012T'),
    'T018V': ('BUKRS', 'T018V'),
    'T028B': ('BUKRS', 'T028B'),
    # Payment program
    'T042': ('BUKRS', 'T042'),
    'T042A': ('ZBUKR', 'T042A'),
    'T042B': ('ZBUKR', 'T042B'),
    'T042C': ('ZBUKR', None),
    'T042D': ('BUKRS', None),
    'T042E': ('LAND1', 'T042E'),
    'T042I': ('ZBUKR', 'T042I'),
    'T042T': ('ZBUKR', None),
    'T042V': ('BUKRS', None),
    'T042P': ('BUKRS', None),
    'T042S': ('ZBUKR', None),
    'T042W': ('BUKRS', None),
    'T042Y': ('BUKRS', None),
    'T042ZA': ('BUKRS', None),
    # Tolerances
    'T043G': ('BUKRS', None),
    'T043S': ('BUKRS', None),
    'T043T': ('BUKRS', None),
    'T043GT': ('BUKRS', None),
    'T043ST': ('BUKRS', None),
    'T169G': ('BUKRS', None),
    'T169P': ('BUKRS', None),
    'T169V': ('BUKRS', None),
    'T169B': ('BUKRS', None),
    'T169R': ('BUKRS', None),
    # Asset Accounting
    'T093': ('BUKRS', None),
    'T093A': ('AFAPL', None),
    'T093B': ('BUKRS', None),
    'T093C': ('BUKRS', None),
    'T093D': ('BUKRS', None),
    'T093U': ('BUKRS', None),
    'T093V': ('BUKRS', None),
    # GL accounts
    'SKB1': ('BUKRS', 'P01_SKB1'),
    # Cash planning
    'T035D': ('BUKRS', 'T035D'),
    'T035U': ('BUKRS', None),
    # Doc Types
    'T003D': ('BUKRS', None),
    # Dunning
    'T047': ('BUKRS', None),
    'T047E': ('BUKRS', None),
    # Consolidation
    'T882': ('RBUKRS', None),
    # ValSub
    'GB92': ('SUBSTID', None),
    'GB93': ('VALID', None),
    'GB90': ('BOOLID', None),
    'GGB1': (None, None),  # TDAT bundle, no direct key
    # Controlling
    'TKA01': ('KOKRS', None),
    'TKA02': ('BUKRS', None),
    'TKA07': ('KOKRS', None),
    'V_TKA3': ('KOKRS', None),
    'CSKA': ('KTOPL', 'P01_CSKA'),
    'CSKB': ('KOKRS', 'P01_CSKB'),
    'CSKS': ('KOKRS', None),
    'CSKU': ('KOKRS', 'P01_CSKU'),
    # FM
    'FM01': ('FIKRS', None),
    'FMCI': ('FIKRS', None),
    'FMFCTR': ('FIKRS', None),
    'FMFINT': ('FIKRS', None),
    'FMCFC0': ('FIKRS', None),
    'FMCFC1': ('FIKRS', None),
    'FMCFC2': ('FIKRS', None),
    # HR
    'T500P': ('WERKS', None),
    'T542A': ('SUBTY', None),
    'TPIR2': ('BUKRS', None),
    'TPIR2T': ('BUKRS', None),
    # MM
    'T024': ('EKORG', None),
    'T024E': ('EKORG', None),
    'T024W': ('WERKS', None),
    # Cash journal
    'TCJ_C_JOURNALS': ('BUKRS', None),
    'TCJ_CJ_NAMES': ('BUKRS', None),
    # Cross-company
    # F01U pairs - check by name pattern (not by key)
    # Tax
    'T007V': ('MWSKZ', None),
    'T030': ('BUKRS', None),
    'T030K': ('BUKRS', None),
    'T030H': ('BUKRS', 'T030H'),
    # Auth roles
    'AGR_TIMEB': (None, None),
    # Others
    'NRIV': ('SUBOBJECT', None),
    'ATPRA': ('BUKRS', None),
    'AAACC_OBJ': ('BUKRS', None),
    'MARV': ('BUKRS', None),
    'BNKA': ('BANKS', None),
    'TIBAN': ('BUKRS', 'TIBAN'),
    'T001D': ('BUKRS', None),
    'T077K': ('KTOKK', 'T077K'),
}


# Build presence matrix via Gold DB + sparse RFC fallback
gold_cache = {}


def check_gold(table, key_col, value):
    cache_key = (table, key_col, value)
    if cache_key in gold_cache:
        return gold_cache[cache_key]
    try:
        con = sqlite3.connect(GOLD_DB)
        q = f"SELECT COUNT(*) FROM {table} WHERE {key_col} = ?"
        result = con.execute(q, (value,)).fetchone()[0]
        con.close()
    except Exception as e:
        result = None
    gold_cache[cache_key] = result
    return result


def rfc_count(guard, table, key_col, value):
    """Count rows in a table where key_col=value via RFC_READ_TABLE."""
    try:
        r = guard.call('RFC_READ_TABLE',
                       QUERY_TABLE=table,
                       FIELDS=[{'FIELDNAME': key_col}],
                       OPTIONS=[{'TEXT': f"{key_col} = '{value}'"}],
                       DELIMITER='|', ROWCOUNT=5000)
        return len(r.get('DATA', []))
    except Exception:
        return None


def presence(table, institute, guard=None):
    """Return None (unknown) or int (count)."""
    meta = TABLE_META.get(table)
    if not meta:
        return None
    key_col, gold_tbl = meta
    if key_col is None:
        return None
    # Translate institute to the right key value
    # For CO and FM tables at UNESCO: KOKRS=BUKRS and FIKRS=BUKRS
    if key_col in ('BUKRS', 'ZBUKR', 'RBUKRS', 'KOKRS', 'FIKRS'):
        value = institute
    elif key_col == 'SUBSTID' or key_col == 'VALID' or key_col == 'BOOLID':
        # institute-named rule
        if institute == 'UNES':
            value = 'UNESCO'
        else:
            value = institute
    else:
        return None  # skip exotic key cols
    # Gold DB first
    if gold_tbl:
        c = check_gold(gold_tbl, key_col, value)
        if c is not None:
            return c
    # Fallback RFC
    if guard is not None:
        return rfc_count(guard, table, key_col, value)
    return None


# Load Golden steps from matrix.json
with open(HERE / 'golden_matrix.json') as f:
    golden = json.load(f)['rows']


# Initialize RFC (only used on demand)
guard = None
try:
    sys.path.insert(0, str(HERE))
    from rfc_helpers import get_connection
    guard = get_connection('D01')
    print('D01 connected for live fallback queries.')
except Exception as e:
    print(f'D01 not connected (Gold DB only mode): {str(e)[:100]}')


def evidence_for(row, institute):
    tables = [t.strip() for t in row['tables'].split(', ') if t.strip()]
    if not tables:
        return '—'
    found = 0
    unknown = 0
    for t in tables:
        c = presence(t, institute, guard)
        if c is None:
            unknown += 1
        elif c > 0:
            found += 1
    total = len(tables)
    if found == total:
        return '✓'
    if found == 0 and unknown == total:
        return '?'
    if found == 0:
        return '✗'
    return f'◐ {found}/{total}'


# Build rows
enriched = []
for row in golden:
    new_row = dict(row)
    for inst in INSTITUTES:
        new_row[inst] = evidence_for(row, inst)
    enriched.append(new_row)

if guard:
    guard.close()


# Save JSON
with open(HERE / 'golden_matrix_all_institutes.json', 'w') as f:
    json.dump({'rows': enriched, 'institutes': INSTITUTES,
               'generated': '2026-04-16'}, f, indent=2)
print(f'Wrote golden_matrix_all_institutes.json with {len(enriched)} rows × {len(INSTITUTES)} institutes')


# Excel
wb = Workbook()
ws = wb.active
ws.title = 'Golden × All Institutes'

hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
hdr_font = Font(color='FFFFFF', bold=True)
phase_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
done_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
partial_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
missing_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
unknown_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

headers = ['Phase', 'Item', 'System', 'Action', 'TCode', 'Type', 'Dep', 'Tables'] + INSTITUTES + ['Comments']
ws.append(headers)
for c in ws[1]:
    c.fill = hdr_fill; c.font = hdr_font
    c.alignment = Alignment(horizontal='center', wrap_text=True)

current_phase = None
for r in enriched:
    if r['phase'] != current_phase:
        ws.append([r['phase']] + [''] * (len(headers) - 1))
        for c in ws[ws.max_row]:
            c.fill = phase_fill; c.font = Font(bold=True)
        current_phase = r['phase']
    vals = [r['phase'], r['item'], r['system'], r['action'], r['tcode'],
            r['type'], r['dep'], r['tables']] + \
           [r[i] for i in INSTITUTES] + [r['comments']]
    ws.append(vals)
    row_idx = ws.max_row
    for i, inst in enumerate(INSTITUTES):
        col_letter = chr(ord('I') + i)  # I=9, J=10...R=18
        cell = ws[f'{col_letter}{row_idx}']
        status = r[inst]
        if status == '✓': cell.fill = done_fill
        elif status == '✗': cell.fill = missing_fill
        elif status.startswith('◐'): cell.fill = partial_fill
        elif status == '?': cell.fill = unknown_fill
        cell.alignment = Alignment(horizontal='center')

widths = [14, 5, 9, 36, 14, 9, 5, 22] + [5] * len(INSTITUTES) + [30]
for col_letter, w in zip('ABCDEFGHIJKLMNOPQRS', widths):
    ws.column_dimensions[col_letter].width = w
ws.row_dimensions[1].height = 32
ws.freeze_panes = 'B2'
ws.auto_filter.ref = f'A1:{chr(ord("A")+len(headers)-1)}{ws.max_row}'

out = Path('Zagentexecution/Golden_per_institute_matrix_ALL_2026-04-16.xlsx')
wb.save(out)
print(f'Wrote {out.resolve()}')

# Summary
print('\n=== Summary ===')
print(f'{"Institute":10s} ✓Full  ◐Partial  ✗Missing  ?Unknown  —NA')
for inst in INSTITUTES:
    s = [r[inst] for r in enriched]
    full = sum(1 for v in s if v == '✓')
    part = sum(1 for v in s if v.startswith('◐'))
    miss = sum(1 for v in s if v == '✗')
    unk = sum(1 for v in s if v == '?')
    na = sum(1 for v in s if v == '—')
    print(f'{inst:10s} {full:4d}   {part:6d}   {miss:7d}   {unk:7d}   {na:5d}')
