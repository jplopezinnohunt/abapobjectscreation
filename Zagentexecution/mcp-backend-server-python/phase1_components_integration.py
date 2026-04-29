"""phase1_components_integration.py

Build the full custom-components interaction map (CONFIG + CODE) and integrate
it into:
  - Excel BCM_StructuredAddress_Analysis.xlsx (new sheet 19)
  - Companion BCM_StructuredAddressChange.html (new tab "Components Map")
  - Brain v2 (new claims, annotations, objects)

The map answers: for every custom-or-touched component in the chain,
what is its type (CONFIG/CODE/DATA), what does it do, what we modify
in V001, and who reviews.
"""
from __future__ import annotations
import json, sys, csv
from pathlib import Path
from datetime import datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

REPO = Path(__file__).resolve().parents[2]
XLSX = REPO / "Zagentexecution" / "incidents" / "xml_payment_structured_address" / "BCM_StructuredAddress_Analysis.xlsx"
COMPONENTS_JSON = REPO / "knowledge" / "domains" / "Payment" / "phase0" / "components_map.json"

# ============================================================
# THE MAP — single source of truth, used by Excel + Companion + Brain
# ============================================================

COMPONENTS = [
    # === CONFIG (DMEE trees) ===
    {
        "id": "tree_sepa",
        "name": "/SEPA_CT_UNES",
        "type": "CONFIG", "layer": "DMEE tree",
        "owner": "Pablo + Marlies",
        "today_v000": "95 nodes · NO Event 05 registered · Dbtr Hybrid (Ctry+AdrLine PARIS) · Cdtr Hybrid · postal code missing per Marlies row 1",
        "v001_change": "Create V001 copy. Register Event 05 = FI_PAYMEDIUM_DMEE_CGI_05 (revised sub-opt A). Add structured Cdtr+Dbtr PstlAdr nodes (StrtNm/PstCd/TwnNm/Ctry minimum). Suppress AdrLine when structured populated. Test in D01 first.",
        "evidence": "Phase 0 GAP-006 + Marlies Excel rows 1-3 + D01 probe (FI_PAYMEDIUM_DMEE_CGI_05 EXISTS)",
        "reviewer": "Pablo (config) + Marlies (verification)",
        "abap_needed": False,
    },
    {
        "id": "tree_cgi",
        "name": "/CGI_XML_CT_UNESCO",
        "type": "CONFIG", "layer": "DMEE tree",
        "owner": "Pablo + Marlies",
        "today_v000": "631 nodes · Event 05 = FI_PAYMEDIUM_DMEE_CGI_05 (already) · Dbtr STRUCTURED OK · Cdtr STRUCTURED OK · CdtrAgt UNSTRUCTURED (3 cases NG/KR/GB)",
        "v001_change": "Create V001 copy. Fix CdtrAgt unstructured nodes — replace AdrLine with structured StrtNm/PstCd/TwnNm/Ctry sourced from BNKA standard bank DB.",
        "evidence": "Marlies Excel rows 8-10 CdtrAgt verdict 'Not accepted'. Probe: FPAYHX-UBSTR/UBORT/UBISO unstructured today.",
        "reviewer": "Pablo + Marlies",
        "abap_needed": False,
    },
    {
        "id": "tree_cgi_1",
        "name": "/CGI_XML_CT_UNESCO_1",
        "type": "CONFIG", "layer": "DMEE tree",
        "owner": "Pablo",
        "today_v000": "639 nodes · Same FIRSTNODE_ID as CGI_XML_CT_UNESCO (N_5649723900) — variant of parent. Last FP_SPEZZANO 2025-02-14",
        "v001_change": "SYNC pattern from V001 of /CGI_XML_CT_UNESCO. Same CdtrAgt fix.",
        "evidence": "Phase 0 Finding D — FIRSTNODE_ID identical confirms variant pattern",
        "reviewer": "Pablo",
        "abap_needed": False,
    },
    {
        "id": "tree_citi",
        "name": "/CITI/XML/UNESCO/DC_V3_01",
        "type": "CONFIG", "layer": "DMEE tree",
        "owner": "Pablo + Marlies + Citi TRM (for Q3)",
        "today_v000": "610 nodes · Event 05 = /CITIPMW/V3_PAYMEDIUM_DMEE_05 (CITIPMW industry solution) · Dbtr UNSTRUCTURED (no Ctry!) · Cdtr mostly OK · UltmtCdtr Hybrid (WHO/ICC case) · XSLT CGI_XML_CT_XSLT post-proc",
        "v001_change": "Create V001 copy. Add Dbtr structured nodes including missing Ctry. UltmtCdtr Worldlink BLOCKED by Q3 (Citi TRM data source confirmation needed) — defer to V002 if unresolved by Phase 2 start.",
        "evidence": "Marlies Excel rows 4,5,7 + Phase 0 Finding F",
        "reviewer": "Pablo + Marlies",
        "abap_needed": False,
    },
    # === CONFIG (OBPM4 events) ===
    {
        "id": "obpm4_event05",
        "name": "TFPM042FB (OBPM4 Event 05 registrations)",
        "type": "CONFIG", "layer": "Customizing table",
        "owner": "Pablo",
        "today_v000": "CGI=FI_PAYMEDIUM_DMEE_CGI_05 (SAP std) · CITI=/CITIPMW/V3_PAYMEDIUM_DMEE_05 · SEPA_CT_UNES=NONE",
        "v001_change": "ADD entry: FORMI=/SEPA_CT_UNES, EVENT=05, FNAME=FI_PAYMEDIUM_DMEE_CGI_05 (revised sub-opt A, test in D01 first). CGI/CITI registrations untouched.",
        "evidence": "Phase 1 batch D01 probe (FI_PAYMEDIUM_DMEE_CGI_05 exists, dispatches to CL_IDFI_CGI_CALL05_FR for French SEPA addresses)",
        "reviewer": "Pablo",
        "abap_needed": False,
    },
    # === CODE (UNESCO BAdI classes) ===
    {
        "id": "abap_fallback",
        "name": "YCL_IDFI_CGI_DMEE_FALLBACK_CM001::GET_CREDIT",
        "type": "CODE", "layer": "BAdI implementation",
        "owner": "N_MENARD (code owner)",
        "today_v000": "Stateful name-overflow-into-StrtNm logic at lines 13-31. Works in V000 (StrtNm empty). Would corrupt V001 structured StrtNm.",
        "v001_change": "Pattern A fix: 3-line guard 'AND c_value IS INITIAL' to prevent overflow corrupting V001 structured StrtNm. Backward-compatible.",
        "evidence": "Source extracted line-by-line · Phase 0 Finding L · Pattern A in Excel sheet 16",
        "reviewer": "N_MENARD (alignment call required)",
        "abap_needed": True,
        "transport": "D01K-BADI-FIX-01",
    },
    {
        "id": "abap_fr",
        "name": "YCL_IDFI_CGI_DMEE_FR (6 includes)",
        "type": "CODE", "layer": "BAdI impl (country=FR)",
        "owner": "N_MENARD",
        "today_v000": "France-specific tag overrides. Implements IF_IDFI_CGI_DMEE_COUNTRIES.",
        "v001_change": "NO CHANGE — works for both V000 and V001 (no stateful address bug)",
        "evidence": "Source extracted, no name-overflow logic present",
        "reviewer": "—",
        "abap_needed": False,
    },
    {
        "id": "abap_util",
        "name": "YCL_IDFI_CGI_DMEE_UTIL (6 includes + CM001-4)",
        "type": "CODE", "layer": "BAdI impl (utilities)",
        "owner": "N_MENARD",
        "today_v000": "Helper class — central tag resolution + customization lookup. GET_TAG_VALUE_FROM_CUSTO method.",
        "v001_change": "NO CHANGE",
        "evidence": "Source extracted, no address logic touched",
        "reviewer": "—",
        "abap_needed": False,
    },
    {
        "id": "abap_de",
        "name": "YCL_IDFI_CGI_DMEE_DE (7 includes)",
        "type": "CODE", "layer": "BAdI impl (country=DE)",
        "owner": "N_MENARD",
        "today_v000": "SHELL — overrides ONLY <CdtrAgt><FinInstnId><ClrSysMmbId><MmbId> from i_fpayh-zbnkl. Delegates everything else to FALLBACK via super->get_credit.",
        "v001_change": "NO CHANGE — but D01 retrofit needed (P01-only object, see Finding I)",
        "evidence": "Phase 1 batch extraction · CM001 confirmed shell pattern",
        "reviewer": "N_MENARD (D01 retrofit approval)",
        "abap_needed": False,
    },
    {
        "id": "abap_it",
        "name": "YCL_IDFI_CGI_DMEE_IT (7 includes)",
        "type": "CODE", "layer": "BAdI impl (country=IT)",
        "owner": "N_MENARD",
        "today_v000": "SHELL — same pattern as DE, only overrides <CdtrAgt><FinInstnId><ClrSysMmbId><MmbId>",
        "v001_change": "NO CHANGE — but D01 retrofit needed (P01-only)",
        "evidence": "Phase 1 batch extraction",
        "reviewer": "N_MENARD",
        "abap_needed": False,
    },
    {
        "id": "abap_enho_de",
        "name": "Y_IDFI_CGI_DMEE_COUNTRIES_DE (ENHO)",
        "type": "CODE", "layer": "Enhancement Implementation",
        "owner": "N_MENARD",
        "today_v000": "Enhancement spot impl — registers DE class with BAdI dispatcher",
        "v001_change": "NO CHANGE — D01 retrofit only",
        "evidence": "Phase 0 Finding I (P01-ONLY)",
        "reviewer": "N_MENARD",
        "abap_needed": False,
    },
    {
        "id": "abap_enho_fr",
        "name": "Y_IDFI_CGI_DMEE_COUNTRIES_FR (ENHO)",
        "type": "CODE", "layer": "Enhancement Implementation",
        "owner": "N_MENARD",
        "today_v000": "Enhancement spot impl — registers FR class with BAdI dispatcher",
        "v001_change": "NO CHANGE — D01 retrofit only (P01-ONLY despite class being in D01)",
        "evidence": "Phase 0 Finding I (P01-ONLY ENHO)",
        "reviewer": "N_MENARD",
        "abap_needed": False,
    },
    {
        "id": "abap_enho_it",
        "name": "Y_IDFI_CGI_DMEE_COUNTRIES_IT (ENHO)",
        "type": "CODE", "layer": "Enhancement Implementation",
        "owner": "N_MENARD",
        "today_v000": "Enhancement spot impl — registers IT class",
        "v001_change": "NO CHANGE — D01 retrofit only",
        "evidence": "Phase 0 Finding I",
        "reviewer": "N_MENARD",
        "abap_needed": False,
    },
    # === SAP-STANDARD (we use, do not modify) ===
    {
        "id": "sap_fm_event05_cgi",
        "name": "FI_PAYMEDIUM_DMEE_CGI_05",
        "type": "CODE (SAP std)", "layer": "Function module — Event 05 wrapper",
        "owner": "SAP (do not modify)",
        "today_v000": "Thin wrapper. Calls cl_idfi_cgi_call05_factory=>get_instance, then instance->fill_fpay_fref. Populates FPAYHX_FREF + FPAYP_FREF buffers from country-specific class.",
        "v001_change": "NO CHANGE — REUSE as-is. Will be registered for /SEPA_CT_UNES Event 05 (revised sub-opt A).",
        "evidence": "Source extracted · 41 lines",
        "reviewer": "—",
        "abap_needed": False,
    },
    {
        "id": "sap_factory",
        "name": "CL_IDFI_CGI_CALL05_FACTORY",
        "type": "CODE (SAP std)", "layer": "Factory class",
        "owner": "SAP",
        "today_v000": "Dispatches by country to: AT, BE, CH, CZ, DE, DK, ES, FI, FR, GB, HR, IT, LU, PT, RO, SK, TH, GENERIC",
        "v001_change": "NO CHANGE",
        "evidence": "TADIR probe in P01 — 20 country variants confirmed",
        "reviewer": "—",
        "abap_needed": False,
    },
    {
        "id": "sap_fm_event05_citi",
        "name": "/CITIPMW/V3_PAYMEDIUM_DMEE_05",
        "type": "CODE (SAP std)", "layer": "FM (Citibank PMW Industry Solution)",
        "owner": "SAP / Citi PMW",
        "today_v000": "Event 05 for /CITI/XML/UNESCO/DC_V3_01 — populates Citi-specific REF fields.",
        "v001_change": "NO CHANGE — keep using",
        "evidence": "TFPM042FB probe",
        "reviewer": "—",
        "abap_needed": False,
    },
    {
        "id": "sap_xslt",
        "name": "CGI_XML_CT_XSLT (XSLT post-processor)",
        "type": "CODE (SAP std)", "layer": "XSLT transformation",
        "owner": "SAP (devclass=ID-DMEE)",
        "today_v000": "Post-processes CITI tree XML output before file write.",
        "v001_change": "NO CHANGE — but Phase 4 UAT must verify it doesn't mangle V001 structured nodes.",
        "evidence": "TADIR probe · source extraction blocked via RFC (SAP-std)",
        "reviewer": "Phase 4 UAT validation",
        "abap_needed": False,
    },
    # === DDIC structures ===
    {
        "id": "ddic_fpayhx_fref",
        "name": "FPAYHX_FREF",
        "type": "DDIC", "layer": "Structure (buffer)",
        "owner": "SAP std",
        "today_v000": "16 fields: CURNO + REF01..REF15 (each CHAR 264). Populated by Event 05 country class. Source for DMEE Dbtr-level structured nodes.",
        "v001_change": "NO CHANGE — we read from this buffer in V001 nodes",
        "evidence": "DDIF_FIELDINFO_GET in P01",
        "reviewer": "—",
        "abap_needed": False,
    },
    {
        "id": "ddic_fpayp_fref",
        "name": "FPAYP_FREF",
        "type": "DDIC", "layer": "Structure (buffer)",
        "owner": "SAP std",
        "today_v000": "5 fields: REF01..REF05 (CHAR 264). Line-level buffer. Source for Cdtr per-line structured nodes.",
        "v001_change": "NO CHANGE",
        "evidence": "DDIF_FIELDINFO_GET",
        "reviewer": "—",
        "abap_needed": False,
    },
    {
        "id": "ddic_fpayhx",
        "name": "FPAYHX (with CI_FPAYHX append)",
        "type": "DDIC", "layer": "Structure (header extension)",
        "owner": "SAP std + UNESCO append",
        "today_v000": "432 fields total · 25 Z-fields incl. ZPFST, ZPLOR, ZLISO, ZNM1S, ZBANK, ZREF01..ZREF10",
        "v001_change": "NO CHANGE — we use existing fields",
        "evidence": "GAP-004 closure",
        "reviewer": "—",
        "abap_needed": False,
    },
    # === DATA (master + customizing) ===
    {
        "id": "data_lfa1_adrc",
        "name": "LFA1 + ADRC (vendor master)",
        "type": "DATA", "layer": "Master data",
        "owner": "Master Data team",
        "today_v000": "111,241 active vendors. 5 missing CITY1 or COUNTRY (mandatory CBPR+).",
        "v001_change": "Manual fix or KILL the 5 LIFNRs (0000020171, 0000020731, 0000020815, 0000020843, 0000059828)",
        "evidence": "Phase 1 batch SQL on Gold DB · sheet 09 + vendor_dq_5_specific.csv",
        "reviewer": "Master Data team (Marlies coordinates)",
        "abap_needed": False,
    },
    {
        "id": "data_t012_t012k",
        "name": "T012 + T012K (house bank)",
        "type": "DATA", "layer": "Customizing",
        "owner": "Pablo (config) + Marlies",
        "today_v000": "House bank addresses (UNESCO own banks). Source for Dbtr Ctry. Some banks have full address, others minimal.",
        "v001_change": "Review for completeness. No structural change expected.",
        "evidence": "Probe context",
        "reviewer": "Pablo + Marlies",
        "abap_needed": False,
    },
    {
        "id": "data_bnka",
        "name": "BNKA (bank master, standard bank DB)",
        "type": "DATA", "layer": "Customizing",
        "owner": "SAP std + UNESCO bank master maintenance",
        "today_v000": "Standard bank database with structured address per bank.",
        "v001_change": "NO CHANGE — used by V001 CGI CdtrAgt nodes as structured source.",
        "evidence": "Marlies Excel CGI rows 8-10 'standard bank database'",
        "reviewer": "—",
        "abap_needed": False,
    },
    {
        "id": "data_t042z",
        "name": "T042Z (payment method per country → tree)",
        "type": "DATA", "layer": "Customizing",
        "owner": "Pablo (config)",
        "today_v000": "263 rows. UNES/IIEP/UIL → /SEPA_CT_UNES (FR+S, DE+S). UNES/UIS/UBO → /CITI/XML/UNESCO/DC_V3_01. UNES/IIEP → /CGI_XML_CT_UNESCO. XSTRA='X' for most, XSTRA='' only for FR+A Treasury Transfer.",
        "v001_change": "Possibly add Z-payment methods routing to V001 trees during Phase 3 testing window. For Phase 5 cutover: standard payment methods point to V001 atomically.",
        "evidence": "GAP-008 closure",
        "reviewer": "Pablo",
        "abap_needed": False,
    },
    # === OUT OF SCOPE (referenced but not modified) ===
    {
        "id": "tfpm042f_variants",
        "name": "TFPM042F* (PMF variants — Francesco's work)",
        "type": "CONFIG", "layer": "OBPM customizing",
        "owner": "Francesco (FP_SPEZZANO)",
        "today_v000": "5 transports 2025 Q1 — bank variant configs (e.g., /CGI_XML_CT_UNESCO_BK)",
        "v001_change": "NO CHANGE in our scope. Courtesy heads-up to Francesco that we are creating V001 of the parent trees.",
        "evidence": "Phase 0 Finding D · Francesco audit · sheet 10",
        "reviewer": "Francesco (courtesy email)",
        "abap_needed": False,
    },
    {
        "id": "wf_90000003",
        "name": "Workflow 90000003 (BCM dual approval)",
        "type": "CONFIG", "layer": "Workflow",
        "owner": "Treasury workflow team",
        "today_v000": "Dual control approval for BCM batches. PD object groups OTYPE='RY'.",
        "v001_change": "NO CHANGE — out of scope. UltmtCdtr structured may render slightly different in BNK_APP UI but test in Phase 4.",
        "evidence": "Brain Claim #65 · feedback_bcm_signatory_ry_otype",
        "reviewer": "—",
        "abap_needed": False,
    },
    {
        "id": "bcm_batch_badis",
        "name": "BCM_BATCH_* BAdIs (BCM_BATCH_ENHANCE/GROUP/APPROVAL/STATUS/SUBMISSION)",
        "type": "CODE", "layer": "BCM BAdI hooks",
        "owner": "?? (not yet identified)",
        "today_v000": "Not extracted. Operate post-DMEE on batch level. May or may not have UNESCO impls.",
        "v001_change": "NO CHANGE in V001 scope. Address change is XML-content, not batch-level.",
        "evidence": "Phase 0 follow-up — extract if BCM-level surprises emerge in Phase 4",
        "reviewer": "—",
        "abap_needed": False,
    },
]


def save_json():
    print(f"Saving components map JSON: {COMPONENTS_JSON}")
    COMPONENTS_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(COMPONENTS_JSON, "w", encoding="utf-8") as f:
        json.dump({"generated_at": datetime.now().isoformat(),
                   "total": len(COMPONENTS),
                   "components": COMPONENTS}, f, indent=2, ensure_ascii=False)


def update_excel():
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    print(f"Updating Excel: {XLSX}")
    wb = load_workbook(XLSX)
    if "19 Components Map" in wb.sheetnames:
        del wb["19 Components Map"]
    ws = wb.create_sheet("19 Components Map")

    headers = ["ID", "Component", "Type", "Layer", "Owner", "Today (V000)", "V001 change", "Evidence", "Reviewer", "ABAP needed?"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="1E3A4A")
        c.font = Font(name="Segoe UI", color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(vertical="center", wrap_text=True)

    type_colors = {
        "CONFIG": "CFE2F3",
        "CODE": "FCE5CD",
        "CODE (SAP std)": "EAD1DC",
        "DDIC": "D9EAD3",
        "DATA": "FFF2CC",
    }
    for c in COMPONENTS:
        row = [
            c["id"], c["name"], c["type"], c["layer"], c["owner"],
            c["today_v000"], c["v001_change"], c["evidence"], c["reviewer"],
            "YES" if c.get("abap_needed") else "no",
        ]
        ws.append(row)
        rownum = ws.max_row
        # Color by type
        fill = PatternFill("solid", fgColor=type_colors.get(c["type"], "F5F5F5"))
        for cell in ws[rownum]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Segoe UI", size=9)
        # Highlight ABAP-needed rows
        if c.get("abap_needed"):
            ws.cell(rownum, 10).fill = PatternFill("solid", fgColor="E74C3C")
            ws.cell(rownum, 10).font = Font(color="FFFFFF", bold=True)
        ws.row_dimensions[rownum].height = 80

    # Column widths
    widths = {1: 18, 2: 32, 3: 14, 4: 28, 5: 22, 6: 55, 7: 55, 8: 35, 9: 25, 10: 12}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "C2"
    wb.save(XLSX)
    print(f"  Saved {len(COMPONENTS)} components to sheet 19")


def main():
    save_json()
    update_excel()
    print("\nSummary:")
    by_type = {}
    abap = 0
    for c in COMPONENTS:
        by_type[c["type"]] = by_type.get(c["type"], 0) + 1
        if c.get("abap_needed"):
            abap += 1
    print(f"  Total components: {len(COMPONENTS)}")
    for t, n in sorted(by_type.items()):
        print(f"    {t}: {n}")
    print(f"  ABAP changes needed: {abap}")
    print(f"  CONFIG changes needed: {sum(1 for c in COMPONENTS if c['type'] == 'CONFIG' and 'NO CHANGE' not in c['v001_change'])}")


if __name__ == "__main__":
    main()
