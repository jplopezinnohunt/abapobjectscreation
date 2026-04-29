"""
Build per-institute evidence matrix for each Golden List step.
For each of the 76 steps, check which tables/objects belong to it and
report whether MGIE / ICBA / ICTP transported them.
Outputs Excel + JSON + HTML snippet for the companion.
"""
import sys
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

HERE = Path(__file__).resolve().parent

# Load pre-computed per-institute object sets
with open(HERE / 'all3_institutes_analysis.json') as f:
    analysis = json.load(f)

mgie = {c: set(v) for c, v in analysis['per_institute']['MGIE'].items()}
icba = {c: set(v) for c, v in analysis['per_institute']['ICBA'].items()}
ictp = {c: set(v) for c, v in analysis['per_institute']['ICTP'].items()}


def all_objs(d):
    s = set()
    for v in d.values():
        s |= v
    return s


mgie_all = all_objs(mgie)
icba_all = all_objs(icba)
ictp_all = all_objs(ictp)

# Load STEM captured state
with open(HERE / 'mgie_vs_stem_raw.json') as f:
    stem_raw = json.load(f)
stem_all = set()
for tr, pl in stem_raw['STEM'].items():
    for o in pl.get('objects', []):
        n = o.get('OBJ_NAME', '').strip()
        if n:
            stem_all.add(n)
    for k in pl.get('keys', []):
        n = k.get('OBJNAME', '').strip()
        if n:
            stem_all.add(n)

# Golden List: 76 steps, each mapped to "evidence tables" (set of SAP tables
# that would appear in transports if the step was done via transport)
GOLDEN = [
    # (phase, item, system, action, tcode, type, dep, tables, comments)
    ('P1 Foundation', 1, 'SAP FI', 'Create / Copy Company Code', 'EC01 / OX02', 'Config', '', {'T001', 'T001Q', 'T001A'}, 'Base legal entity record'),
    ('P1 Foundation', 2, 'SAP FI', 'Add Group Currency', 'OB22', 'Config', '1', {'T001A'}, 'Parallel currency activation'),
    ('P1 Foundation', 3, 'SAP FI', 'Company code global parameters', 'OBY6', 'Config', '1', {'T001'}, 'KTOPL, PERIV, FSTVA, WAERS'),
    ('P1 Foundation', 4, 'SAP FI', 'Open posting periods', 'OB52', 'Config', '1', {'T001B'}, 'Not transported — per env'),
    ('P1 Foundation', 5, 'SAP FI', 'Document + object number ranges', 'FBN1 / OBH2', 'Config', '1', {'NRIV'}, 'Per env'),
    ('P2 Banking/Payment', 6, 'SAP FI', 'Create house banks', 'FI12', 'Config', '1', {'T012', 'T012D', 'T012T'}, 'Country-specific'),
    ('P2 Banking/Payment', 7, 'SAP FI', 'Bank directory entries', 'FI01', 'Master Data', '6', {'T012A', 'T012B', 'T012C', 'T018V', 'T018C', 'T018D', 'T028', 'T028B', 'BNKA'}, 'Country bank directory + SWIFT'),
    ('P2 Banking/Payment', 8, 'SAP FI', 'Bank accounts (T012K)', 'FI12', 'Config', '6', {'T012K'}, 'GL per bank account'),
    ('P2 Banking/Payment', 9, 'SAP FI', 'Electronic Bank Statement', 'OT83 / FF.5', 'Config', '6,8', {'T028G', 'T028I', 'T028S'}, 'Transaction types, determination'),
    ('P2 Banking/Payment', 10, 'SAP FI', 'FBZP — All company codes', 'FBZP', 'Config', '1', {'T042'}, 'Base co code payment setup'),
    ('P2 Banking/Payment', 11, 'SAP FI', 'FBZP — Paying company codes', 'FBZP', 'Config', '10', {'T042C', 'T042T'}, 'Methods + signatory'),
    ('P2 Banking/Payment', 12, 'SAP FI', 'Payment methods per country', 'FBZP', 'Config', '10', {'T042E'}, 'Country-level'),
    ('P2 Banking/Payment', 13, 'SAP FI', 'Payment methods per co code', 'FBZP', 'Config', '11,12', {'T042B'}, 'Method limits'),
    ('P2 Banking/Payment', 14, 'SAP FI', 'Bank determination / ranking', 'FBZP', 'Config', '8,13', {'T042I'}, 'Critical for F110'),
    ('P2 Banking/Payment', 15, 'SAP FI', 'FBZP extras (T042A/D/V/P/S/W/Y/ZA)', 'FBZP', 'Config', '8,14', {'T042A', 'T042D', 'T042V', 'T042P', 'T042S', 'T042W', 'T042Y', 'T042ZA'}, 'Full F110 variants (ICTP-style)'),
    ('P2 Banking/Payment', 16, 'SAP FI', 'DMEE payment format tree', 'DMEE / OBPM1', 'Config', '10', {'TRGD'}, 'Country format'),
    ('P2 Banking/Payment', 17, 'SAP FI', 'OBPM4 selection variants', 'OBPM4', 'Config', '16', set(), 'Never transported'),
    ('P2 Banking/Payment', 18, 'SAP FI', 'Cash management planning groups', 'OT42', 'Config', '1', {'T035D', 'T035U'}, 'Per currency'),
    ('P3 GL Accounts', 19, 'SAP FI', 'Extend GL per co code (SKB1)', 'FS00', 'Master Data', '1', {'SKB1'}, 'Aligned STEM 2026-04-15'),
    ('P3 GL Accounts', 20, 'SAP FI', 'GL automatic account determination', 'OBYC / OBYA', 'Config', '19', {'T030', 'T030K'}, 'GR/IR, FX, tax'),
    ('P3 GL Accounts', 21, 'SAP FI', 'Exchange rate diff accounts', 'OBA1', 'Config', '19,20', {'T030H', 'T030HB'}, 'Per valuation method'),
    ('P3 GL Accounts', 22, 'SAP FI', 'Retained earnings account', 'OB53', 'Config', '19', {'T030'}, 'Per co code'),
    ('P4 ValSub', 23, 'SAP FI', 'Create substitution rule (GGB1)', 'GGB1', 'Config', '19', {'GGB1', 'GB90', 'GB92', 'GB921', 'GB922'}, 'Mirror MGIE 8-step'),
    ('P4 ValSub', 24, 'SAP FI', 'Assign substitution (OBBH)', 'OBBH', 'Config', '23', set(), 'T001D-family or OBBH table'),
    ('P4 ValSub', 25, 'SAP FI', 'Create validation rule (GGB0) — opt', 'GGB0', 'Config', '19', {'GB93', 'GGB0'}, 'Optional'),
    ('P4 ValSub', 26, 'SAP FI', 'Assign validation (OB28) — opt', 'OB28', 'Config', '25', {'T001D'}, 'T001D row'),
    ('P4 ValSub', 27, 'SAP FI', '** RGUGBR00 regenerate **', 'SE38 RGUGBR00', 'Activate', '23,24', set(), 'MANDATORY per env'),
    ('P4 ValSub', 28, 'SAP FI', 'YRGGBS00 ABAP exit per-BUKRS branch', 'SE38', 'Dev', '23', set(), 'Source code custom'),
    ('P5 Tolerances', 29, 'SAP FI', 'Tolerance groups (FI)', 'OBA4 / OBA0', 'Config', '1', {'T043G', 'T043GT', 'T043S', 'T043ST', 'T043T'}, 'Employee + posting'),
    ('P5 Tolerances', 30, 'SAP FI', 'Assign users to tolerance groups', 'OB57', 'Config', '29', set(), 'Per user, per env'),
    ('P5 Tolerances', 31, 'SAP MM', 'MM/IV tolerances', 'OMR6', 'Config', '1', {'T169G', 'T169P', 'T169V', 'T169B', 'T169R'}, 'Invoice verification'),
    ('P6 Tax', 32, 'SAP FI', 'Create tax codes per country', 'FTXP', 'Config', '1', {'T007V', 'T007A', 'T683'}, 'NEVER copy'),
    ('P6 Tax', 33, 'SAP FI', 'Assign tax procedure to country', 'OBBG', 'Config', '32', {'T005'}, 'Country → procedure'),
    ('P6 Tax', 34, 'SAP FI', 'Tax GL account determination', 'OBCN', 'Config', '32,19', {'T030K'}, 'Per tax code'),
    ('P6 Tax', 35, 'SAP FI', 'Withholding tax (optional)', 'OBWP / OBWQ', 'Config', '32', {'T059Z'}, 'Country dependent'),
    ('P7 Dunning', 36, 'SAP FI-AR', 'Dunning procedure', 'FBMP / OB96', 'Config', '1', {'T047', 'T047E'}, 'Per co code'),
    ('P8 Asset Accounting', 37, 'SAP FI-AA', 'Chart of depreciation assignment', 'OAOB', 'Config', '1', {'T093C', 'T093C_RSL', 'T093CE'}, ''),
    ('P8 Asset Accounting', 38, 'SAP FI-AA', 'Asset class → GL determination', 'AO90', 'Config', '19,37', {'T093', 'T093A', 'T093B', 'T093B_RSL'}, ''),
    ('P8 Asset Accounting', 39, 'SAP FI-AA', 'Depreciation keys + useful lives', 'AFAMS', 'Config', '37', {'T093S', 'T093T', 'T093Y', 'T096'}, 'Country tax'),
    ('P8 Asset Accounting', 40, 'SAP FI-AA', 'Asset transfer date', 'OAAQ', 'Config', '37', {'T093D'}, 'Legacy asset go-live'),
    ('P8 Asset Accounting', 41, 'SAP FI-AA', 'Open AA period', 'OAAQ', 'Config', '40', {'T093D'}, 'Per fiscal year'),
    ('P9 Controlling', 42, 'SAP FI-CO', 'Create CO area + hierarchy', 'OKKP', 'Config', '1', {'TKA00', 'TKA01', 'TKA02', 'TKA07', 'TKA09'}, 'KOKRS=BUKRS per UNESCO'),
    ('P9 Controlling', 43, 'SAP FI-CO', 'CO versions + plan', 'OKEQ', 'Config', '42', {'V_TKA3', 'TKA30', 'TKA31'}, 'Plan settings'),
    ('P9 Controlling', 44, 'SAP FI-CO', 'Create main cost centers', 'KS01', 'Master Data', '42', {'CSKS', 'CSKT'}, 'Initial hierarchy'),
    ('P9 Controlling', 45, 'SAP FI-CO', 'Create cost elements', 'KA01 / OKTZ', 'Master Data', '42,19', {'CSKA', 'CSKB', 'CSKU'}, 'STEM aligned 2026-04-15'),
    ('P9 Controlling', 46, 'SAP FI-CO', 'Default cost center per GL', 'OKB9', 'Config', '44,19', {'FINB_TR_DERIVATION'}, 'FINB'),
    ('P9 Controlling', 47, 'SAP FI-CO', 'Profit centers (N/A at UNESCO)', 'KE51', 'Master Data', '42', {'CEPC', 'CEPCT'}, 'Not used at UNESCO'),
    ('P9 Controlling', 48, 'SAP FI-CO', 'CO allocation cycles', 'KSV5 / KSU5', 'Config', '44', set(), 'Optional'),
    ('P10 FM', 49, 'SAP FM', 'Create FM area + config', 'OF01', 'Config', '1', {'FM01'}, 'FIKRS=BUKRS'),
    ('P10 FM', 50, 'SAP FM', 'Fund types', 'FM5D', 'Config', '49', {'V_FMFUNDTYPE', 'FMFUNDTYPE'}, ''),
    ('P10 FM', 51, 'SAP FM', 'Activate FM components', 'OFUP', 'Config', '49', {'FM01'}, 'XFMCO, XFMCB, XFMCA'),
    ('P10 FM', 52, 'SAP FM', 'Create commitment items', 'FMCIA / FMCI', 'Master Data', '49', {'FMCI', 'FMCIT'}, '22-26 items at UNESCO'),
    ('P10 FM', 53, 'SAP FM', 'Create fund centers', 'FMSA', 'Master Data', '49', {'FMFCTR', 'FMFCTRT'}, ''),
    ('P10 FM', 54, 'SAP FM', 'Create funds', 'FM5I', 'Master Data', '49,50', {'FMFINT'}, ''),
    ('P10 FM', 55, 'SAP FM', 'FM derivation strategy', 'FMDERIVE', 'Config', '52', set(), 'GL→CI bridge'),
    ('P10 FM', 56, 'SAP FM', 'AVC activation', 'FMBB / FM5T', 'Config', '53,54', set(), 'Budget control'),
    ('P10 FM', 57, 'SAP FM', 'Carry-forward + reassignment (FMCFC)', 'FMBN / 2KEC', 'Config', '54', {'FMCFC0', 'FMCFC1', 'FMCFC2', 'FMIUR', 'FMREAS_STRATS', 'FMZUCO'}, 'ICTP-pattern full carry-forward'),
    ('P11 PS', 58, 'SAP FI-PS', 'PS integration (profiles)', 'OPSA / OPSB', 'Config', '42', set(), 'Project profiles'),
    ('P12 Consolidation', 59, 'SAP FI-LC', 'Consolidation unit assignment', 'OC13', 'Config', '1', {'T882'}, ''),
    ('P13 Cash Journal', 60, 'SAP FI', 'Cash journal setup', 'FBCJC0', 'Config', '19', {'TCJ_C_JOURNALS', 'TCJ_CJ_NAMES', 'TCJ_PRINT', 'TCJ_TRANSACTIONS', 'TCJ_TRANS_NAMES'}, 'Optional petty cash'),
    ('P14 Cross-Company', 61, 'SAP FI', 'Cross-company clearing pairs', 'OBYA', 'Config', '19', set(), 'F01U pairs bilateral'),
    ('P15 HR/Travel', 62, 'SAP HR', 'Personnel area + subarea', 'OOMO', 'Config', '1', {'T500P', 'T542A'}, 'Country + subarea'),
    ('P15 HR/Travel', 63, 'SAP HR-TV', 'Travel integration', 'PR05', 'Config', '62', {'TPIR2', 'TPIR2T'}, 'Travel personnel link'),
    ('P16 MM', 64, 'SAP MM', 'Plant creation + assignment', 'OX10 / OX18', 'Config', '1', {'T001W', 'T001K'}, 'Only if procures'),
    ('P16 MM', 65, 'SAP MM', 'Purchasing organization', 'OX17', 'Config', '64', {'T024', 'T024E', 'T024W'}, ''),
    ('P16 MM', 66, 'SAP MM', 'MM Account assignment + GBB/WRX', 'OME9 / OBYC', 'Config', '64,19', {'UNESGBB', 'UNESWRX'}, 'UNESCO-custom keys'),
    ('P16 MM', 67, 'SAP MM', 'Release strategy', 'OMGQ', 'Config', '64,65', {'T16F', 'T16FG_2', 'T16FG_3'}, 'PO approval workflow'),
    ('P17 Auth Roles', 68, 'SAP ROLE', 'Clone Y_*_* roles (~48)', 'PFCG / SUPC', 'Role', '19', {'AGR_TIMEB', 'SPERS_OBJ', 'USR12', 'USR13', 'UST12'}, 'Y_XXXX_* pattern'),
    ('P17 Auth Roles', 69, 'SAP ROLE', 'Assign roles to users', 'SU01 / SU10', 'Role', '68', set(), 'Per env'),
    ('P17 Auth Roles', 70, 'SAP ROLE', 'Update F_BKPF_BUK values', 'PFCG', 'Role', '68', set(), 'Multi-BUKRS roles'),
    ('P18 Master Data', 71, 'SAP FI-AP', 'Extend vendors to co code', 'FK01 / XK01', 'Master Data', '1,19', set(), 'LFB1 per BUKRS'),
    ('P18 Master Data', 72, 'SAP FI-AR', 'Extend customers to co code', 'FD01 / XD01', 'Master Data', '1,19', set(), 'KNB1 per BUKRS'),
    ('P19 Testing', 73, 'SAP FI', 'Test FI posting', 'FB50', 'Test', 'all', set(), 'FM + substitution check'),
    ('P19 Testing', 74, 'SAP FI', 'Test F110 payment', 'F110', 'Test', '14,15', set(), 'Bank ranking check'),
    ('P19 Testing', 75, 'SAP FI', 'Test EBS import', 'FF.5', 'Test', '9', set(), 'MT940 format'),
    ('P19 Testing', 76, 'SAP FM', 'Test budget + commitment', 'FMX1 / FMBB', 'Test', '52-56', set(), 'AVC verify'),
]


def evidence(tables, inst_objs):
    """Return tuple (status, matched_tables)."""
    if not tables:
        return ('—', [])
    matched = [t for t in tables if t in inst_objs]
    if not matched:
        return ('✗', [])
    elif len(matched) == len(tables):
        return ('✓', matched)
    else:
        return (f'◐ {len(matched)}/{len(tables)}', matched)


# Build matrix
rows = []
for row in GOLDEN:
    phase, item, system, action, tcode, typ, dep, tables, comments = row
    mgie_st, mgie_m = evidence(tables, mgie_all)
    icba_st, icba_m = evidence(tables, icba_all)
    ictp_st, ictp_m = evidence(tables, ictp_all)
    stem_st, stem_m = evidence(tables, stem_all)
    rows.append({
        'phase': phase, 'item': item, 'system': system, 'action': action,
        'tcode': tcode, 'type': typ, 'dep': dep,
        'tables': ', '.join(sorted(tables)) if tables else '',
        'MGIE': mgie_st, 'MGIE_matched': mgie_m,
        'ICBA': icba_st, 'ICBA_matched': icba_m,
        'ICTP': ictp_st, 'ICTP_matched': ictp_m,
        'STEM': stem_st, 'STEM_matched': stem_m,
        'comments': comments,
    })


# Excel
wb = Workbook()
ws = wb.active
ws.title = 'Golden × Institute'

hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
hdr_font = Font(color='FFFFFF', bold=True)
phase_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
done_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
partial_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
missing_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')

headers = ['Phase', 'Item', 'System', 'Action', 'TCode', 'Type', 'Dep', 'Tables',
           'MGIE', 'ICBA', 'ICTP', 'STEM', 'Comments']
ws.append(headers)
for c in ws[1]:
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = Alignment(horizontal='center', wrap_text=True)

current_phase = None
for r in rows:
    if r['phase'] != current_phase:
        ws.append([r['phase']] + [''] * (len(headers) - 1))
        for c in ws[ws.max_row]:
            c.fill = phase_fill
            c.font = Font(bold=True)
        current_phase = r['phase']

    ws.append([
        r['phase'], r['item'], r['system'], r['action'], r['tcode'],
        r['type'], r['dep'], r['tables'],
        r['MGIE'], r['ICBA'], r['ICTP'], r['STEM'], r['comments'],
    ])
    row_idx = ws.max_row
    # Color MGIE/ICBA/ICTP/STEM cells
    for col_letter, status in [('I', r['MGIE']), ('J', r['ICBA']), ('K', r['ICTP']), ('L', r['STEM'])]:
        cell = ws[f'{col_letter}{row_idx}']
        if status == '✓':
            cell.fill = done_fill
        elif status == '✗':
            cell.fill = missing_fill
        elif status.startswith('◐'):
            cell.fill = partial_fill

widths = [14, 5, 9, 36, 15, 9, 5, 28, 7, 7, 7, 7, 30]
for col_letter, w in zip('ABCDEFGHIJKLM', widths):
    ws.column_dimensions[col_letter].width = w
ws.row_dimensions[1].height = 32
ws.freeze_panes = 'B2'
ws.auto_filter.ref = f'A1:M{ws.max_row}'

out_xlsx = Path('Zagentexecution/Golden_per_institute_matrix_2026-04-16.xlsx')
wb.save(out_xlsx)
print(f'Wrote: {out_xlsx.resolve()}')

# JSON for companion
with open(HERE / 'golden_matrix.json', 'w') as f:
    json.dump({
        'rows': [{k: v for k, v in r.items() if k != 'MGIE_matched' and k != 'ICBA_matched'
                  and k != 'ICTP_matched' and k != 'STEM_matched'} for r in rows],
        'generated': '2026-04-16',
    }, f, indent=2)
print(f'Wrote: golden_matrix.json')

# Quick summary
print('\n=== Per-institute coverage summary ===')
print(f'{"Institute":10s} {"Full (✓)":10s} {"Partial (◐)":12s} {"Missing (✗)":12s} {"N/A (—)":10s}')
for inst in ['MGIE', 'ICBA', 'ICTP', 'STEM']:
    statuses = [r[inst] for r in rows]
    full = sum(1 for s in statuses if s == '✓')
    partial = sum(1 for s in statuses if s.startswith('◐'))
    missing = sum(1 for s in statuses if s == '✗')
    na = sum(1 for s in statuses if s == '—')
    print(f'{inst:10s} {full:10d} {partial:12d} {missing:12d} {na:10d}')
