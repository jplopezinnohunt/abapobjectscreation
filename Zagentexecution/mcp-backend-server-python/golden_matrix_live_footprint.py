"""
Rebuild Golden matrix with LIVE full-history footprint per institute.

For each of the 76 Golden steps, check whether any of the step's tables
has BUKRS-keyed entries for each institute across ALL transports in D01's
history (not just the creation cluster).

This answers the real question: "Does this institute HAVE this config today?"
"""
import sys
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

HERE = Path(__file__).resolve().parent

with open(HERE / 'all_institutes_footprint_live.json') as f:
    footprints = json.load(f)

with open(HERE / 'golden_matrix.json') as f:
    golden = json.load(f)['rows']

INSTITUTES = ['UNES', 'IBE', 'IIEP', 'UBO', 'UIL', 'UIS', 'ICTP', 'ICBA', 'MGIE', 'STEM']

# Known-migrated institutes: their config exists but transport history is missing
MIGRATED = {'IBE', 'UBO', 'UIL', 'UIS'}


def evidence(tables, footprint, is_migrated):
    if not tables:
        return '—'
    ts = [t.strip() for t in (tables.split(', ') if isinstance(tables, str) else tables) if t.strip()]
    if not ts:
        return '—'
    matched = [t for t in ts if t in footprint]
    if matched and len(matched) == len(ts):
        return '✓'
    if matched:
        return f'◐ {len(matched)}/{len(ts)}'
    if is_migrated:
        return '∅'  # migrated-institute: no transport trail
    return '✗'


rows = []
for g in golden:
    new = dict(g)
    for inst in INSTITUTES:
        fp = footprints.get(inst, {})
        new[inst] = evidence(g['tables'], fp, inst in MIGRATED)
    rows.append(new)

# Save JSON
with open(HERE / 'golden_matrix_live.json', 'w') as f:
    json.dump({'rows': rows, 'institutes': INSTITUTES}, f, indent=2)

# Excel
wb = Workbook()
ws = wb.active
ws.title = 'Golden × Live'
hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
hdr_font = Font(color='FFFFFF', bold=True)
phase_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
done_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
partial_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
missing_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
migrated_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
unknown_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

headers = ['Phase', 'Item', 'System', 'Action', 'TCode', 'Type', 'Dep', 'Tables'] + INSTITUTES + ['Comments']
ws.append(headers)
for c in ws[1]:
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = Alignment(horizontal='center', wrap_text=True)

current_phase = None
for r in rows:
    if r['phase'] != current_phase:
        ws.append([r['phase']] + [''] * (len(headers) - 1))
        for c in ws[ws.max_row]:
            c.fill = phase_fill
            c.font = Font(bold=True)
        current_phase = r['phase']
    vals = [r['phase'], r['item'], r['system'], r['action'], r['tcode'],
            r['type'], r['dep'], r['tables']] + [r[i] for i in INSTITUTES] + [r['comments']]
    ws.append(vals)
    row_idx = ws.max_row
    for i, inst in enumerate(INSTITUTES):
        col_letter = chr(ord('I') + i)
        cell = ws[f'{col_letter}{row_idx}']
        cell.alignment = Alignment(horizontal='center')
        status = r[inst]
        if status == '✓':
            cell.fill = done_fill
        elif status == '✗':
            cell.fill = missing_fill
        elif status.startswith('◐'):
            cell.fill = partial_fill
        elif status == '∅':
            cell.fill = migrated_fill

widths = [14, 5, 9, 36, 14, 8, 5, 22] + [5] * len(INSTITUTES) + [30]
for col_letter, w in zip('ABCDEFGHIJKLMNOPQRS', widths):
    ws.column_dimensions[col_letter].width = w
ws.row_dimensions[1].height = 32
ws.freeze_panes = 'B2'
ws.auto_filter.ref = f'A1:{chr(ord("A")+len(headers)-1)}{ws.max_row}'

out = Path('Zagentexecution/Golden_live_matrix_2026-04-16.xlsx')
wb.save(out)
print(f'Wrote: {out.resolve()}')

# Summary
print('\n=== Live-footprint coverage per institute ===')
print(f'{"Institute":10s} ✓Full  ◐Partial  ✗Missing  ∅Migrated  —NA')
for inst in INSTITUTES:
    s = [r[inst] for r in rows]
    full = sum(1 for v in s if v == '✓')
    part = sum(1 for v in s if v.startswith('◐'))
    miss = sum(1 for v in s if v == '✗')
    mig = sum(1 for v in s if v == '∅')
    na = sum(1 for v in s if v == '—')
    print(f'{inst:10s} {full:4d}   {part:6d}   {miss:7d}   {mig:8d}   {na:5d}')
