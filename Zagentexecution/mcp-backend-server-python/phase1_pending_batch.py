"""phase1_pending_batch.py — execute Phase 1 pending items batch.

1. Extract YCL_IDFI_CGI_DMEE_DE + _IT class source via class include reading
2. Extract XSLT CGI_XML_CT_XSLT source via RPY_PROGRAM_READ
3. Probe D01 for FVD_SEPA_OL_CT_DMEE_05 + alternatives existence (read-only)
4. Read T012/T012K to verify house bank / vendor master DQ link
"""
from __future__ import annotations
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
from rfc_helpers import get_connection

REPO = HERE.parent.parent
OUT = REPO / "extracted_code" / "FI" / "DMEE_full_inventory"
OUT.mkdir(parents=True, exist_ok=True)


def read_program(conn, prog_name):
    """Read source via RPY_PROGRAM_READ."""
    try:
        r = conn.call("RPY_PROGRAM_READ", PROGRAM_NAME=prog_name, WITH_LOWERCASE="X")
        lines = r.get("SOURCE_EXTENDED", []) or r.get("SOURCE", [])
        return "\n".join(l.get("LINE", "") if isinstance(l, dict) else l for l in lines)
    except Exception as e:
        return f"[ERR {prog_name}: {e}]"


def step1_de_it_classes():
    """Extract YCL_IDFI_CGI_DMEE_DE + _IT via class includes (SAP naming convention)."""
    print("\n[1] Extract YCL_IDFI_CGI_DMEE_DE + _IT classes via class includes")
    conn = get_connection("P01")
    try:
        for cls in ["YCL_IDFI_CGI_DMEE_DE", "YCL_IDFI_CGI_DMEE_IT"]:
            print(f"\n  -- {cls} --")
            # Class includes: CLASSNAME (30 chars) + CCDEF/CCIMP/CCMAC/CP/CO/CU/CM001..CMnnn
            # SAP naming pads class name to 30 chars with =
            padded = cls.ljust(30, "=")
            for suffix in ["CCDEF", "CCIMP", "CCMAC", "CP", "CO", "CU", "CM001", "CM002", "CM003", "CM004"]:
                inc_name = f"{padded}{suffix}"
                try:
                    src = read_program(conn, inc_name)
                    if src and not src.startswith("[ERR"):
                        out_path = OUT / f"{cls}_{suffix}.abap"
                        out_path.write_text(src, encoding="utf-8")
                        non_empty = sum(1 for l in src.split("\n") if l.strip() and not l.strip().startswith("*"))
                        print(f"    {suffix:<6}: {len(src):>5} chars, {non_empty} non-comment lines -> {out_path.name}")
                except Exception as e:
                    pass
    finally:
        conn.close()


def step2_xslt():
    """Extract XSLT CGI_XML_CT_XSLT source."""
    print("\n[2] Extract XSLT CGI_XML_CT_XSLT")
    conn = get_connection("P01")
    try:
        # XSLT in SAP is stored as a special program type; try various reads
        for prog in ["CGI_XML_CT_XSLT", "CGI_XML_DD_XSLT"]:
            print(f"\n  -- {prog} --")
            # Try RPY_PROGRAM_READ
            src = read_program(conn, prog)
            if src and not src.startswith("[ERR"):
                out_path = OUT / f"{prog}.xsl"
                out_path.write_text(src, encoding="utf-8")
                print(f"    saved: {out_path.name} ({len(src):,} chars)")
                continue

            # Try as XSLT object via O2_API_GET_OBJECT
            try:
                r = conn.call("O2_API_GET_XSLT", XSLT_DESC=prog)
                print(f"    O2_API result keys: {list(r.keys())}")
            except Exception as e:
                print(f"    O2_API_GET_XSLT err: {e}")

            # Try via TADIR-based RPY_OBJECT_READ
            try:
                r = conn.call("OBJECT_LOAD", OBJ_TYPE="XSLT", OBJ_NAME=prog)
                print(f"    OBJECT_LOAD: {r}")
            except Exception as e:
                print(f"    OBJECT_LOAD err: {e}")

            # Try reading the underlying include — XSLTs sometimes have a virtual program
            try:
                r = conn.call("TR_OBJECT_LOAD_TLOGOTAB", OBJECT="XSLT", OBJ_NAME=prog)
                print(f"    TR_OBJECT_LOAD: {r}")
            except Exception as e:
                pass

    finally:
        conn.close()


def step3_d01_sepa_event05():
    """Read-only probe in D01 for SEPA Event 05 candidates."""
    print("\n[3] D01 read-only probe for SEPA Event 05 candidates")
    try:
        conn = get_connection("D01")
    except Exception as e:
        print(f"  D01 connection failed: {e}")
        return
    try:
        # Check if FM exists in D01
        for fm in ["FVD_SEPA_OL_CT_DMEE_05", "FI_PAYMEDIUM_DMEE_CGI_05",
                   "FI_PAYMEDIUM_DMEE_CN_05", "FI_PAYMEDIUM_SEPA_05"]:
            try:
                r = conn.call("RFC_READ_TABLE",
                              QUERY_TABLE="TFDIR",
                              DELIMITER="|",
                              OPTIONS=[{"TEXT": f"FUNCNAME = '{fm}'"}],
                              FIELDS=[{"FIELDNAME": "FUNCNAME"}, {"FIELDNAME": "PNAME"}])
                if r.get("DATA"):
                    vals = r["DATA"][0].get("WA", "").split("|")
                    print(f"  ✅ {fm} EXISTS in D01 (PNAME={vals[1].strip() if len(vals)>1 else '?'})")
                else:
                    print(f"  ❌ {fm} NOT in D01")
            except Exception as e:
                print(f"  ?? {fm}: {e}")

        # Check D01's TFPM042FB for what's registered for SEPA-like trees
        print("\n  D01 SEPA-related Event registrations:")
        try:
            r = conn.call("RFC_READ_TABLE",
                          QUERY_TABLE="TFPM042FB",
                          DELIMITER="|",
                          ROWCOUNT=20,
                          OPTIONS=[{"TEXT": "FORMI LIKE '%SEPA%'"}],
                          FIELDS=[{"FIELDNAME": "FORMI"}, {"FIELDNAME": "EVENT"}, {"FIELDNAME": "FNAME"}])
            for d in r.get("DATA", [])[:10]:
                vals = d.get("WA", "").split("|")
                print(f"    {dict(zip(['FORMI','EVENT','FNAME'], [v.strip() for v in vals]))}")
        except Exception as e:
            print(f"    err: {e}")
    finally:
        conn.close()


def step4_inspect_marlies_excel_dme_changes():
    """Re-read Marlies's Excel and extract any specific text we missed about implementation pattern."""
    print("\n[4] Re-inspect Marlies's Excel for implementation hints")
    from openpyxl import load_workbook
    p = REPO / "Zagentexecution" / "incidents" / "xml_payment_structured_address" / "original_marlies" / "XML Address un structured.xlsx"
    if not p.exists():
        print(f"  not found: {p}")
        return
    wb = load_workbook(p, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n  === {sn} ===")
        # Look for "general remark" column or any free text
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    s = cell.strip()
                    if any(kw in s.lower() for kw in ["recommend", "remark", "test", "verify", "ensure"]) and len(s) > 30:
                        print(f"    → {s[:180]}")


def main():
    step1_de_it_classes()
    step2_xslt()
    step3_d01_sepa_event05()
    step4_inspect_marlies_excel_dme_changes()


if __name__ == "__main__":
    main()
