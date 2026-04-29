"""
Build gap analysis Excel + markdown from mgie_vs_stem_raw.json.
Run: python stem_align_to_mgie_cskb_dryrun_helper.py
"""
import json
from collections import Counter, defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

HERE = Path(__file__).resolve().parent
with open(HERE / 'mgie_vs_stem_raw.json') as f:
    raw = json.load(f)
with open(HERE / 'mgie_creation_cluster_2013.json') as f:
    cluster = json.load(f)


def cat(n):
    n = n.strip().upper()
    if not n: return '99_empty'
    if n in ('T001', 'T001Q', 'T001A', 'T001B', 'T001V', 'T001Z', 'T001O'): return '01_CoCode'
    if n.startswith('T012') or n in ('TIBAN', 'BNKA') or n.startswith('T028') or n.startswith('T030H'): return '02_HouseBanks'
    if n.startswith('T042') or n.startswith('T033'): return '03_PaymentProgram'
    if n.startswith('T043') or n.startswith('T169'): return '04_Tolerances'
    if n.startswith('T093') or n in ('ATPRA', 'AAACC_OBJ', 'T094', 'T095'): return '05_AssetAccounting'
    if n == 'SKB1': return '06_GL_CoCode'
    if n in ('SKA1', 'SKAT'): return '06b_GLMaster'
    if n == 'MARV': return '07_MMPeriod'
    if n in ('T035D', 'T035U', 'T003D', 'T003T'): return '08_CashPlanning_DocType'
    if n.startswith('T047'): return '09_Dunning'
    if n == 'T882': return '10_Consolidation'
    if n.startswith('GB') or n in ('GGB1', 'GGB0'): return '11_ValSub_Rules'
    if n in ('TKA00', 'TKA01', 'TKA02', 'TKA07', 'TKA09', 'TKT09', 'TKA3J', 'V_TKA3'): return '12_ControllingArea'
    if n in ('CSKA', 'CSKU', 'CSKB', 'CSKS', 'CSKT', 'CEPC', 'CEPCT', 'AUFK'): return '12b_COMasterData'
    if n.startswith('FM') or 'FUND' in n or n in ('FMCI', 'FMFCTR', 'FMFINT') or n.startswith('FMUP'): return '13_FM'
    if n.startswith('V_T001') or 'CAREAMAINT' in n or 'ADDRESS' in n: return '14_ViewMetadata'
    if n.startswith('TPIR') or n in ('T542A',) or n.startswith('T500') or n.startswith('T527'): return '15_HR_Travel'
    if n.startswith('T030'): return '16_GLAutoPost'
    if n.startswith('TCJ'): return '17_CashJournal'
    if n.startswith('TBPF') or n.startswith('FMHIV') or n.startswith('FMFUND'): return '18_BudgetPlanning'
    if n.startswith('T004'): return '19_ChartOfAccounts'
    if n.startswith('Y_'): return '20a_AuthRoles_Custom'
    pair_names = ['MGIE', 'ICTP', 'IIEP', 'UBO', 'UIL', 'UIS', 'UNES', 'IBE', 'ICBA', 'STEM']
    if any(p in n for p in pair_names) and len(n.replace(' ', '')) <= 10 and \
       not n.startswith(('V_', 'VC_', 'T', 'F', 'Y', 'Z', 'VV_', 'W_', 'CS', 'AQ')):
        return '20b_CrossCompanyPair'
    if n == 'GRW_SET' or n.startswith('KE34'): return '21_COPA_Sets'
    if n == 'FINB_TR_DERIVATION': return '22_CO_Derivation'
    if n.startswith('AQ') or n in ('USR12', 'USR13', 'UST12', 'AGR_TIMEB', 'SPERS_OBJ', 'SUSPR', 'USMD2001', 'USMD2003'): return '23_SystemTech'
    if n.startswith('T007') or 'TAXIN' in n or n.startswith('T683'): return '24_Tax'
    if n.startswith('T024') or n.startswith('T16F') or n.startswith('T027'): return '25_MMorg'
    if n in ('UNESGBB', 'UNESWRX'): return '26_AccountGroups'
    return '99_Other'


def summarize(cluster_data):
    obj_by_cat = defaultdict(set)
    keys_by_cat = defaultdict(int)
    for tr, pl in cluster_data.items():
        if pl.get('error'): continue
        for o in pl['objects']:
            n = o.get('OBJ_NAME', '').strip()
            if n and not n.startswith('D01K9'):
                obj_by_cat[cat(n)].add(n)
        for k in pl['keys']:
            keys_by_cat[cat(k.get('OBJNAME', '').strip())] += 1
    return obj_by_cat, keys_by_cat


mgie_obj, mgie_keys = summarize(raw['MGIE'])
stem_obj, stem_keys = summarize(raw['STEM'])
all_cats = sorted(set(mgie_obj) | set(stem_obj))

cat_labels = {
    '01_CoCode': 'Company Code core',
    '02_HouseBanks': 'House Banks + Bank Master',
    '03_PaymentProgram': 'Payment Program (FBZP/F110)',
    '04_Tolerances': 'Tolerances',
    '05_AssetAccounting': 'Asset Accounting',
    '06_GL_CoCode': 'GL Accounts (co code)',
    '06b_GLMaster': 'GL Master (chart)',
    '07_MMPeriod': 'MM Period',
    '08_CashPlanning_DocType': 'Cash Planning / Doc Types',
    '09_Dunning': 'Dunning',
    '10_Consolidation': 'Consolidation',
    '11_ValSub_Rules': 'Validations/Substitutions (GGB1)',
    '12_ControllingArea': 'Controlling Area',
    '12b_COMasterData': 'CO Master Data',
    '13_FM': 'Funds Management',
    '14_ViewMetadata': 'View metadata',
    '15_HR_Travel': 'HR / Travel',
    '16_GLAutoPost': 'GL Auto Post',
    '17_CashJournal': 'Cash Journal',
    '18_BudgetPlanning': 'Budget Planning',
    '19_ChartOfAccounts': 'Chart of Accounts',
    '20a_AuthRoles_Custom': 'Custom Auth Roles (Y_*)',
    '20b_CrossCompanyPair': 'Cross-Company Pairs (F01U)',
    '21_COPA_Sets': 'CO-PA / Sets',
    '22_CO_Derivation': 'CO Derivation (OKB9)',
    '23_SystemTech': 'System/Tech',
    '24_Tax': 'Tax codes',
    '25_MMorg': 'MM Org structures',
    '26_AccountGroups': 'Account Groups',
    '99_Other': 'Other',
    '99_empty': 'empty',
}
crit_map = {
    '03_PaymentProgram': 'CRITICAL', '11_ValSub_Rules': 'CRITICAL',
    '12_ControllingArea': 'CRITICAL', '13_FM': 'CRITICAL',
    '15_HR_Travel': 'HIGH', '17_CashJournal': 'HIGH',
    '20a_AuthRoles_Custom': 'HIGH', '20b_CrossCompanyPair': 'HIGH',
    '24_Tax': 'HIGH', '22_CO_Derivation': 'MEDIUM',
    '08_CashPlanning_DocType': 'MEDIUM', '25_MMorg': 'MEDIUM',
}
rec_map = {
    '03_PaymentProgram': 'MANDATORY: Configure FBZP payment methods/bank ranking for STEM',
    '11_ValSub_Rules': 'MANDATORY: Create STEM substitution rule in GGB1 + OBBH assignment',
    '12_ControllingArea': 'MANDATORY: Complete CO area config (V_TKA3 - CO versions) for STEM',
    '13_FM': 'MANDATORY: Complete FM area config (FM01, V_FMFUNDTYPE) for STEM',
    '15_HR_Travel': 'HIGH: Add STEM personnel area (T500P, T542A)',
    '17_CashJournal': 'HIGH: Configure cash journal if STEM uses it',
    '20a_AuthRoles_Custom': 'HIGH: Clone Y_MGIE_* roles as Y_STEM_*',
    '20b_CrossCompanyPair': 'HIGH: Create F01U pairs STEM<->each other institute',
    '24_Tax': 'HIGH: Create Chinese tax codes (STEM) - India-specific MGIE tax must not be reused',
    '22_CO_Derivation': 'MEDIUM: FINB_TR_DERIVATION (OKB9 defaults) for STEM',
    '08_CashPlanning_DocType': 'MEDIUM: T003D document type defaults',
    '25_MMorg': 'MEDIUM: MM org (plant, purchasing org) if STEM procures',
}

wb = Workbook()
hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
hdr_font = Font(color='FFFFFF', bold=True)
crit_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

ws = wb.active; ws.title = '1_Gap_Overview'
ws.append(['Category', 'MGIE obj', 'MGIE keys', 'STEM obj', 'STEM keys',
           'Gap obj', 'Gap keys', 'Criticality'])
for c in ws[1]:
    c.fill = hdr_fill; c.font = hdr_font
    c.alignment = Alignment(horizontal='center', wrap_text=True)

for c in all_cats:
    mo = len(mgie_obj.get(c, set())); mk = mgie_keys.get(c, 0)
    so = len(stem_obj.get(c, set())); sk = stem_keys.get(c, 0)
    gap_o = mo - so; gap_k = mk - sk
    crit = crit_map.get(c, '')
    label = cat_labels.get(c, c)
    ws.append([label, mo, mk, so, sk, gap_o, gap_k, crit])
    rn = ws.max_row
    if crit == 'CRITICAL':
        for col in range(1, 9): ws.cell(row=rn, column=col).fill = crit_fill
    elif gap_o <= 0 and gap_k <= 0 and mo > 0:
        for col in range(1, 9): ws.cell(row=rn, column=col).fill = ok_fill

for col, w in zip('ABCDEFGH', [40, 12, 13, 12, 13, 10, 11, 14]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:H{ws.max_row}'

# Sheet 2: MGIE-only objects
ws2 = wb.create_sheet('2_MGIE_only_objects')
ws2.append(['Category', 'Object Name', 'Recommendation for STEM'])
for c in ws2[1]:
    c.fill = hdr_fill; c.font = hdr_font
    c.alignment = Alignment(horizontal='center')
for c in all_cats:
    only_mgie = sorted(mgie_obj.get(c, set()) - stem_obj.get(c, set()))
    if only_mgie:
        rec = rec_map.get(c, '')
        label = cat_labels.get(c, c)
        for n in only_mgie:
            ws2.append([label, n, rec])
for col, w in zip('ABC', [32, 40, 80]):
    ws2.column_dimensions[col].width = w
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f'A1:C{ws2.max_row}'

# Sheet 3: MGIE transports
ws3 = wb.create_sheet('3_MGIE_Transports')
ws3.append(['Date', 'TRKORR', 'User', 'FN', 'Text', 'Objects', 'Keys'])
for c in ws3[1]:
    c.fill = hdr_fill; c.font = hdr_font
    c.alignment = Alignment(horizontal='center')
for r in cluster['main_transports']:
    pl = raw['MGIE'].get(r['TRKORR'], {})
    obj_count = len(pl.get('objects', []))
    key_count = len(pl.get('keys', []))
    ws3.append([r['DATE'], r['TRKORR'], r['USER'], r['FN'], r['TEXT'], obj_count, key_count])
for col, w in zip('ABCDEFG', [12, 15, 14, 6, 60, 10, 10]):
    ws3.column_dimensions[col].width = w
ws3.freeze_panes = 'A2'

# Sheet 4: STEM transports
ws4 = wb.create_sheet('4_STEM_Transports')
ws4.append(['TRKORR', 'Objects', 'Keys', 'Notes'])
for c in ws4[1]:
    c.fill = hdr_fill; c.font = hdr_font
stem_notes = {
    'D01K9B0CBF': 'FI core (JP_LOPEZ) - empty on read',
    'D01K9B0CBG': 'FI task (JP_LOPEZ) - SKB1 aligned 2026-04-15 (639 keys) + CSKB added (359 keys)',
    'D01K9B0F3V': 'FM (I_KONAKOV) - empty on read',
    'D01K9B0F4I': 'Posting periods (R_RIOS)',
    'D01K9B0F40': 'Logistics (R_RIOS) - empty on read',
}
for tr, pl in raw['STEM'].items():
    ws4.append([tr, len(pl.get('objects', [])), len(pl.get('keys', [])), stem_notes.get(tr, '')])
for col, w in zip('ABCD', [15, 10, 10, 70]):
    ws4.column_dimensions[col].width = w
ws4.freeze_panes = 'A2'

out = Path('Zagentexecution/MGIE_vs_STEM_gap_analysis_2026-04-15.xlsx')
wb.save(out)
print(f'Wrote: {out.resolve()}')

# Markdown summary
md_lines = [
    '# MGIE -> STEM Gap Analysis',
    '',
    '**Generated:** 2026-04-15 - live RFC on E071/E071K',
    '**MGIE source:** 16 transports from 2013 creation cluster',
    '**STEM source:** 5 transports currently in progress',
    '',
    '## Executive summary',
    '',
]
mgie_total_obj = sum(len(p.get('objects', [])) for p in raw['MGIE'].values())
mgie_total_key = sum(len(p.get('keys', [])) for p in raw['MGIE'].values())
stem_total_obj = sum(len(p.get('objects', [])) for p in raw['STEM'].values())
stem_total_key = sum(len(p.get('keys', [])) for p in raw['STEM'].values())
md_lines += [
    f'- MGIE creation cluster (2013): **16 transports, {mgie_total_obj} objects, {mgie_total_key} keys** over 5 months (Jul->Dec 2013)',
    f'- STEM creation cluster (2026 in-progress): **5 transports, {stem_total_obj} objects, {stem_total_key} keys**',
    '',
    '## Gap by category',
    '',
    '| Category | MGIE (obj/keys) | STEM (obj/keys) | Gap | Criticality |',
    '|---|---|---|---|---|',
]
for c in all_cats:
    mo = len(mgie_obj.get(c, set())); mk = mgie_keys.get(c, 0)
    so = len(stem_obj.get(c, set())); sk = stem_keys.get(c, 0)
    crit = crit_map.get(c, '')
    label = cat_labels.get(c, c)
    md_lines.append(f'| {label} | {mo}/{mk} | {so}/{sk} | {mo - so} obj, {mk - sk} keys | {crit} |')

md_lines += ['', '## Critical gaps (STEM missing):', '']
for c in ['03_PaymentProgram', '11_ValSub_Rules', '12_ControllingArea', '13_FM']:
    only_mgie = sorted(mgie_obj.get(c, set()) - stem_obj.get(c, set()))
    if only_mgie:
        md_lines.append(f'### {cat_labels.get(c, c)}')
        md_lines.append(f'**Rec:** {rec_map.get(c, "")}')
        md_lines.append(f'**Missing objects:** `{", ".join(only_mgie)}`')
        md_lines.append('')

md_lines += ['', '## All MGIE 2013 creation transports:', '',
             '| Date | TRKORR | User | Text | Objects | Keys |',
             '|---|---|---|---|---|---|']
for r in cluster['main_transports']:
    pl = raw['MGIE'].get(r['TRKORR'], {})
    md_lines.append(f'| {r["DATE"]} | `{r["TRKORR"]}` | {r["USER"]} | {r["TEXT"]} | {len(pl.get("objects", []))} | {len(pl.get("keys", []))} |')

md_path = Path('knowledge/configuration_retros/MGIE_vs_STEM_gap_2026-04-15.md')
md_path.parent.mkdir(parents=True, exist_ok=True)
md_path.write_text('\n'.join(md_lines), encoding='utf-8')
print(f'Wrote: {md_path.resolve()}')
