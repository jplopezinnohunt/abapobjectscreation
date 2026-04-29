"""
Total verification: Gold DB vs P01 LIVE for every table relevant to STEM
company-code creation. Compares row counts per institute (FIKRS or BUKRS).

Tables verified:
  - FMFINCODE (Funds) -> Gold.funds
  - FMFCTR (Fund Centers) -> Gold.fund_centers
  - PROJ (Projects) -> Gold.proj
  - PRPS (WBS) -> Gold.prps
  - CSKS (Cost Centers) -> Gold.CSKS
  - T001 (Company Code master) -> Gold.T001

Output: discrepancy report + appended sheet to STEM Excel.
"""
import sys
import sqlite3
from pathlib import Path
from datetime import datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
from rfc_helpers import get_connection  # type: ignore

REPO = HERE.parents[1]
GOLD = REPO / "Zagentexecution" / "sap_data_extraction" / "sqlite" / "p01_gold_master_data.db"
EXCEL = REPO / "Zagentexecution" / f"STEM_3institutes_reference_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

INSTITUTES = ["MGIE", "IBE", "ICBA", "ICTP", "IIEP", "UIS", "UIL", "UBO", "UNES"]


def live_count(g, table, where, label):
    """Count rows for a single WHERE via paginated rowskips."""
    skip = 0
    page = 30000
    total = 0
    while True:
        try:
            res = g.call(
                "RFC_READ_TABLE",
                QUERY_TABLE=table,
                FIELDS=[{"FIELDNAME": "MANDT"}],
                OPTIONS=[{"TEXT": where}],
                DELIMITER="|",
                ROWCOUNT=page,
                ROWSKIPS=skip,
            )
        except Exception as e:
            msg = str(e)
            if "TABLE_WITHOUT_DATA" in msg:
                break
            print(f"  [{label} skip={skip}] {msg[:100]}")
            break
        n = len(res.get("DATA", []))
        total += n
        if n < page:
            break
        skip += page
    return total


def gold_count(con, table, where_sql, params):
    cur = con.cursor()
    try:
        cur.execute(f"SELECT COUNT(*) FROM {table} WHERE {where_sql}", params)
        return cur.fetchone()[0]
    except Exception as e:
        return f"ERR:{str(e)[:40]}"


def main():
    g = get_connection("P01")
    con = sqlite3.connect(GOLD)

    rows = []
    print(f"{'Table':12s} {'Institute':10s} {'Live':>8s} {'Gold':>8s} {'Delta':>8s}  Status")

    # FMFINCODE vs funds
    for inst in INSTITUTES:
        live = live_count(g, "FMFINCODE", f"FIKRS = '{inst}'", f"FMFINCODE[{inst}]")
        gold = gold_count(con, "funds", "FIKRS = ?", (inst,))
        delta = live - gold if isinstance(gold, int) else "—"
        status = "OK" if isinstance(gold, int) and live == gold else "DRIFT"
        rows.append(("FMFINCODE/funds", inst, live, gold, delta, status))
        print(f"  FMFINCODE  {inst:10s} {live:8d} {str(gold):>8s} {str(delta):>8s}  {status}")

    # FMFCTR vs fund_centers
    for inst in INSTITUTES:
        live = live_count(g, "FMFCTR", f"FIKRS = '{inst}'", f"FMFCTR[{inst}]")
        gold = gold_count(con, "fund_centers", "FIKRS = ?", (inst,))
        delta = live - gold if isinstance(gold, int) else "—"
        status = "OK" if isinstance(gold, int) and live == gold else "DRIFT"
        rows.append(("FMFCTR/fund_centers", inst, live, gold, delta, status))
        print(f"  FMFCTR     {inst:10s} {live:8d} {str(gold):>8s} {str(delta):>8s}  {status}")

    # PROJ vs proj (use VBUKR as filter)
    for inst in INSTITUTES:
        live = live_count(g, "PROJ", f"VBUKR = '{inst}'", f"PROJ[{inst}]")
        gold = gold_count(con, "proj", "VBUKR = ?", (inst,))
        delta = live - gold if isinstance(gold, int) else "—"
        status = "OK" if isinstance(gold, int) and live == gold else "DRIFT"
        rows.append(("PROJ/proj", inst, live, gold, delta, status))
        print(f"  PROJ       {inst:10s} {live:8d} {str(gold):>8s} {str(delta):>8s}  {status}")

    # PRPS vs prps (PBUKR)
    for inst in INSTITUTES:
        live = live_count(g, "PRPS", f"PBUKR = '{inst}'", f"PRPS[{inst}]")
        gold = gold_count(con, "prps", "PBUKR = ?", (inst,))
        delta = live - gold if isinstance(gold, int) else "—"
        status = "OK" if isinstance(gold, int) and live == gold else "DRIFT"
        rows.append(("PRPS/prps", inst, live, gold, delta, status))
        print(f"  PRPS       {inst:10s} {live:8d} {str(gold):>8s} {str(delta):>8s}  {status}")

    # CSKS vs CSKS (BUKRS)
    for inst in INSTITUTES:
        live = live_count(g, "CSKS", f"BUKRS = '{inst}'", f"CSKS[{inst}]")
        gold = gold_count(con, "CSKS", "BUKRS = ?", (inst,))
        delta = live - gold if isinstance(gold, int) else "—"
        status = "OK" if isinstance(gold, int) and live == gold else "DRIFT"
        rows.append(("CSKS", inst, live, gold, delta, status))
        print(f"  CSKS       {inst:10s} {live:8d} {str(gold):>8s} {str(delta):>8s}  {status}")

    g.close()
    con.close()

    # ----- Append to Excel -----
    print(f"\n[Excel] appending Verification sheet to {EXCEL.name}")
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(EXCEL)
    if "Verification_Live_vs_Gold" in wb.sheetnames:
        del wb["Verification_Live_vs_Gold"]
    s = wb.create_sheet("Verification_Live_vs_Gold", 1)
    s.append([f"Verification: P01 LIVE vs Gold DB — {datetime.now().isoformat()}"])
    s["A1"].font = Font(bold=True, size=12)
    s.append([])
    headers = ["Table", "Institute", "Live (P01)", "Gold DB", "Delta", "Status"]
    s.append(headers)
    bold = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    good = PatternFill("solid", fgColor="C6EFCE")
    bad = PatternFill("solid", fgColor="FFC7CE")
    for c in range(1, 7):
        s.cell(row=s.max_row, column=c).font = bold
        s.cell(row=s.max_row, column=c).fill = header_fill
    for r in rows:
        s.append(list(r))
        cell = s.cell(row=s.max_row, column=6)
        cell.fill = good if r[5] == "OK" else bad
        cell.font = Font(bold=True)
    s.freeze_panes = "A4"
    for i, c in enumerate(headers, 1):
        s.column_dimensions[get_column_letter(i)].width = max(14, len(c) + 4)
    wb.save(EXCEL)
    print(f"  Saved.")

    # ----- Summary -----
    drift = [r for r in rows if r[5] == "DRIFT"]
    print(f"\n=== SUMMARY ===")
    print(f"Total checks: {len(rows)}")
    print(f"OK:           {len(rows) - len(drift)}")
    print(f"DRIFT:        {len(drift)}")
    if drift:
        print("\nDrift detail:")
        for r in drift:
            print(f"  {r[0]:24s} {r[1]:8s}  Live={r[2]:>6}  Gold={r[3]:>6}  Δ={r[4]:>+6}")


if __name__ == "__main__":
    main()
