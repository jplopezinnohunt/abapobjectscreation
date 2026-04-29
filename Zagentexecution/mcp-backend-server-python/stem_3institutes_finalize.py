"""
Finalize STEM 3-institutes Excel with corrected understanding:
- All 3 institutes (MGIE, IBE, ICBA) HAVE their own FM area + CO area
- ICBA simply has empty/sparse master data (1 cost center, 0 funds, 0 fund_centers)
- T036FT does NOT exist — fund TYPE codes are free-text per UNESCO convention
- Infer fund-type descriptions by clustering FMFINT (fund descriptions) per TYPE
"""
import sys
import sqlite3
from pathlib import Path
from datetime import datetime
from collections import Counter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
from rfc_helpers import get_connection  # type: ignore

REPO = HERE.parents[1]
GOLD = REPO / "Zagentexecution" / "sap_data_extraction" / "sqlite" / "p01_gold_master_data.db"
INSTITUTES = ["MGIE", "IBE", "ICBA"]
EXCEL = REPO / "Zagentexecution" / f"STEM_3institutes_reference_{datetime.now().strftime('%Y-%m-%d')}.xlsx"


def rfc_table(guard, table, fields, where, label, rowcount=80000):
    try:
        res = guard.call(
            "RFC_READ_TABLE",
            QUERY_TABLE=table,
            FIELDS=[{"FIELDNAME": f} for f in fields],
            OPTIONS=[{"TEXT": where}],
            DELIMITER="|",
            ROWCOUNT=rowcount,
        )
    except Exception as e:
        msg = str(e)
        if "TABLE_WITHOUT_DATA" in msg:
            print(f"  [{label}] no rows")
            return []
        print(f"  [{label}] ERR: {msg[:120]}")
        return []
    rows = []
    for d in res.get("DATA", []):
        wa = d["WA"].split("|")
        rows.append({f: wa[i].strip() if i < len(wa) else "" for i, f in enumerate(fields)})
    return rows


def main():
    g = get_connection("P01")
    print("Connected P01.")

    # ----- 1. T001 with FIKRS for the 3 institutes -----
    t001 = rfc_table(g, "T001", ["BUKRS", "FIKRS", "BUTXT", "LAND1", "WAERS", "KTOPL", "KOKFI", "FMHRDATE"],
                     " OR ".join([f"BUKRS = '{i}'" for i in INSTITUTES]), "T001")
    print(f"T001 rows: {len(t001)}")

    # ----- 2. TKA02 (CO area assignment) -----
    tka02 = rfc_table(g, "TKA02", ["KOKRS", "BUKRS"],
                      " OR ".join([f"BUKRS = '{i}'" for i in INSTITUTES]), "TKA02")
    print(f"TKA02 rows: {len(tka02)}")

    # ----- 3. FMFINT (Fund texts) for the 3 institutes -----
    fmfint = []
    for i in INSTITUTES:
        rows = rfc_table(g, "FMFINT", ["FIKRS", "FINCODE", "SPRAS", "BEZEICH", "BESCHR"],
                         f"FIKRS = '{i}' AND SPRAS = 'E'", f"FMFINT[{i}]")
        fmfint.extend(rows)
    print(f"FMFINT rows: {len(fmfint)}")

    # ----- 4. CSKS already in Gold DB (from previous run) -----
    con = sqlite3.connect(GOLD)
    cur = con.cursor()
    cur.execute(f"SELECT BUKRS,KOKRS,KOSTL,DATBI,DATAB,KOSAR,VERAK,ERSDA,USNAM,KTEXT,LTEXT FROM CSKS WHERE BUKRS IN ('MGIE','IBE','ICBA') ORDER BY BUKRS,KOSTL,DATBI")
    csks_cols = [d[0] for d in cur.description]
    csks = [dict(zip(csks_cols, r)) for r in cur.fetchall()]
    print(f"CSKS rows from Gold DB: {len(csks)}")

    # ----- 5. funds + fund_centers from Gold DB -----
    cur.execute(f"SELECT FIKRS,FINCODE,TYPE,ERFDAT,ERFNAME FROM funds WHERE FIKRS IN ('MGIE','IBE','ICBA') ORDER BY FIKRS,FINCODE")
    funds = [dict(zip([d[0] for d in cur.description], r)) for r in cur.fetchall()]
    cur.execute(f"SELECT FIKRS,FICTR,ERFDAT,ERFNAME FROM fund_centers WHERE FIKRS IN ('MGIE','IBE','ICBA') ORDER BY FIKRS,FICTR")
    fc = [dict(zip([d[0] for d in cur.description], r)) for r in cur.fetchall()]
    cur.execute(f"SELECT VBUKR,PSPID,POST1,VERNR,ERDAT,PSPNR FROM proj WHERE VBUKR IN ('MGIE','IBE','ICBA') ORDER BY VBUKR,PSPID")
    proj = [dict(zip([d[0] for d in cur.description], r)) for r in cur.fetchall()]
    cur.execute(f"SELECT PBUKR,POSID,POST1,VERNR,ERDAT,PSPHI,PSPNR,OBJNR FROM prps WHERE PBUKR IN ('MGIE','IBE','ICBA') ORDER BY PBUKR,POSID")
    prps = [dict(zip([d[0] for d in cur.description], r)) for r in cur.fetchall()]
    print(f"Gold DB: funds={len(funds)} fc={len(fc)} proj={len(proj)} prps={len(prps)}")
    con.close()

    # ----- 6. Enrich funds with text + cluster TYPE descriptions -----
    text_map = {(t["FIKRS"], t["FINCODE"]): (t.get("BEZEICH", ""), t.get("BESCHR", "")) for t in fmfint}
    for f in funds:
        bz, bs = text_map.get((f["FIKRS"], f["FINCODE"]), ("", ""))
        f["BEZEICH"] = bz
        f["BESCHR"] = bs

    # Cluster descriptions per (FIKRS, TYPE)
    cluster = {}
    for f in funds:
        key = (f["FIKRS"], f["TYPE"])
        cluster.setdefault(key, []).append((f.get("BEZEICH", ""), f.get("BESCHR", "")))

    fund_type_summary = []
    for (fikrs, ftype), descs in sorted(cluster.items()):
        words = []
        for bz, bs in descs:
            words.extend(w for w in (bz + " " + bs).split() if len(w) > 3)
        common = [w for w, _ in Counter(words).most_common(8)]
        sample = descs[0][0] if descs[0][0] else descs[0][1]
        fund_type_summary.append({
            "FIKRS": fikrs,
            "FONDSART": ftype,
            "num_funds": len(descs),
            "sample_desc": sample[:80],
            "frequent_terms": " | ".join(common[:6]),
        })

    g.close()

    # ----- 7. Build/Update Excel -----
    print(f"\nUpdating Excel: {EXCEL}")
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(EXCEL)
    bold = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    bad_fill = PatternFill("solid", fgColor="FFC7CE")

    # Drop old/redundant sheets
    for old in ["ICBA_Verification", "FundTypes_T036FT_live", "FundTypes_distinct_codes",
                "Funds_FMFINCODE", "FundCenters_FMFCTR", "Projects_PROJ", "WBS_PRPS",
                "CostCenters_CSKS"]:
        if old in wb.sheetnames:
            del wb[old]

    # ===== Summary (rebuild) =====
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    s = wb.create_sheet("Summary", 0)
    s.append(["STEM Company-Code Copy — Reference Data: MGIE / IBE / ICBA"])
    s["A1"].font = Font(bold=True, size=14)
    s.append([f"Generated: {datetime.now().isoformat()} — Source: P01 RFC + Gold DB"])
    s.append([])

    # Configuration matrix
    s.append(["1. Configuration footprint — does the company code have its own FM/CO area?"])
    s.cell(row=s.max_row, column=1).font = Font(bold=True, size=12)
    s.append(["Institute", "BUKRS", "FIKRS (FM area)", "KOKRS (CO area)", "KTOPL (Chart)", "Currency", "Country"])
    for c in range(1, 8):
        s.cell(row=s.max_row, column=c).font = bold
        s.cell(row=s.max_row, column=c).fill = header_fill
    t001_map = {r["BUKRS"]: r for r in t001}
    tka02_map = {r["BUKRS"]: r["KOKRS"] for r in tka02}
    for inst in INSTITUTES:
        t = t001_map.get(inst, {})
        s.append([
            t.get("BUTXT", inst), inst, t.get("FIKRS", ""),
            tka02_map.get(inst, ""), t.get("KTOPL", ""), t.get("WAERS", ""), t.get("LAND1", "")
        ])

    s.append([])
    s.append(["2. Master data volumes per institute"])
    s.cell(row=s.max_row, column=1).font = Font(bold=True, size=12)
    s.append(["Object", "MGIE", "IBE", "ICBA", "Total"])
    for c in range(1, 6):
        s.cell(row=s.max_row, column=c).font = bold
        s.cell(row=s.max_row, column=c).fill = header_fill

    def by(rows, key, code):
        return sum(1 for r in rows if r[key] == code)

    rows_def = [
        ("Funds (FMFINCODE)", funds, "FIKRS"),
        ("Fund Centers (FMFCTR)", fc, "FIKRS"),
        ("Cost Centers (CSKS)", csks, "BUKRS"),
        ("Projects (PROJ)", proj, "VBUKR"),
        ("WBS Elements (PRPS)", prps, "PBUKR"),
    ]
    for label, data, key in rows_def:
        counts = [by(data, key, i) for i in INSTITUTES]
        s.append([label, *counts, sum(counts)])
        for col_idx, val in enumerate(counts, 2):
            cell = s.cell(row=s.max_row, column=col_idx)
            cell.fill = good_fill if val > 0 else bad_fill

    s.append([])
    s.append(["3. Distinct fund TYPE codes per institute (FMFINCODE.TYPE)"])
    s.cell(row=s.max_row, column=1).font = Font(bold=True, size=12)
    s.append(["FIKRS", "FONDSART", "# funds using it", "Sample description", "Frequent terms"])
    for c in range(1, 6):
        s.cell(row=s.max_row, column=c).font = bold
        s.cell(row=s.max_row, column=c).fill = header_fill
    for r in fund_type_summary:
        s.append([r["FIKRS"], r["FONDSART"], r["num_funds"], r["sample_desc"], r["frequent_terms"]])

    s.append([])
    s.append(["KEY FINDINGS — for STEM creation"])
    s.cell(row=s.max_row, column=1).font = Font(bold=True, size=12)
    findings = [
        "1. ALL 3 institutes have their OWN FM area (FIKRS=BUKRS) and CO area (KOKRS=BUKRS). STEM must follow the same pattern.",
        "2. ICBA is a 'skeleton' company code: FM area + CO area + 1 cost center exist, but 0 funds and 0 fund centers. ICBA likely posts via UNES funds in practice.",
        "3. T036FT (Fund Type catalog) does NOT exist in this system — domain FM_FUNDTYPE has no fixed values. Fund TYPE codes (100/200/300...) are UNESCO free-text convention, descriptions inferable only via FMFINT (fund-level texts).",
        "4. MGIE is the most diverse: 9 distinct fund types (100, 200, 300, 301, 302, 303, 304, 305, 306). IBE only uses 4 (100, 200, 300, 306).",
        "5. The recommended copy template for STEM is MGIE — most complete fund-type spectrum and full project structure (7 projects, 7 WBS).",
        "6. The MGIE fund-type spectrum (301..306) suggests sub-categorization of institute-specific operations that IBE/ICBA flattened into 100/200/300.",
    ]
    for f in findings:
        s.append([f])
        s.cell(row=s.max_row, column=1).fill = note_fill

    for col in range(1, 8):
        s.column_dimensions[get_column_letter(col)].width = [60, 14, 18, 18, 16, 12, 14][col - 1] if col <= 7 else 14

    # ===== Detail sheets =====
    def add_sheet(title, cols, rows):
        s = wb.create_sheet(title)
        s.append(cols)
        for c in range(1, len(cols) + 1):
            s.cell(row=1, column=c).font = bold
            s.cell(row=1, column=c).fill = header_fill
        for r in rows:
            s.append([r.get(c, "") for c in cols])
        s.freeze_panes = "A2"
        for i, c in enumerate(cols, 1):
            s.column_dimensions[get_column_letter(i)].width = max(12, min(45, len(c) + 4))

    add_sheet("FundTypes_Inferred", ["FIKRS", "FONDSART", "num_funds", "sample_desc", "frequent_terms"], fund_type_summary)
    add_sheet("Funds_FMFINCODE", ["FIKRS", "FINCODE", "TYPE", "BEZEICH", "BESCHR", "ERFDAT", "ERFNAME"], funds)
    add_sheet("FundCenters_FMFCTR", ["FIKRS", "FICTR", "ERFDAT", "ERFNAME"], fc)
    add_sheet("CostCenters_CSKS", ["BUKRS", "KOKRS", "KOSTL", "KTEXT", "LTEXT", "KOSAR", "VERAK", "DATAB", "DATBI", "ERSDA", "USNAM"], csks)
    add_sheet("Projects_PROJ", ["VBUKR", "PSPID", "POST1", "VERNR", "ERDAT", "PSPNR"], proj)
    add_sheet("WBS_PRPS", ["PBUKR", "POSID", "POST1", "VERNR", "ERDAT", "PSPHI", "PSPNR", "OBJNR"], prps)

    wb.save(EXCEL)
    print(f"Excel saved: {EXCEL}")

    # ----- 8. Persist FMFINT to Gold DB -----
    con = sqlite3.connect(GOLD)
    cur = con.cursor()
    cur.execute("DROP TABLE IF EXISTS FMFINT")
    cur.execute("CREATE TABLE FMFINT (FIKRS TEXT, FINCODE TEXT, SPRAS TEXT, BEZEICH TEXT, BESCHR TEXT, PRIMARY KEY (FIKRS, FINCODE, SPRAS))")
    cur.executemany("INSERT OR REPLACE INTO FMFINT VALUES (?,?,?,?,?)",
                    [(r["FIKRS"], r["FINCODE"], r["SPRAS"], r.get("BEZEICH", ""), r.get("BESCHR", "")) for r in fmfint])
    cur.execute("DROP TABLE IF EXISTS TKA02")
    cur.execute("CREATE TABLE TKA02 (KOKRS TEXT, BUKRS TEXT, PRIMARY KEY (KOKRS, BUKRS))")
    cur.executemany("INSERT OR REPLACE INTO TKA02 VALUES (?,?)", [(r["KOKRS"], r["BUKRS"]) for r in tka02])
    con.commit()
    con.close()
    print(f"Gold DB: FMFINT + TKA02 added")


if __name__ == "__main__":
    main()
