"""
Build per-institute transport inventory tabs in Excel — same concept as STEM.

For each institute, pull the creation-cluster transports' E071/E071K and
explain every object with Golden step mapping.

Institutes with full creation traceability:
  - MGIE (16 transports, 2013)
  - ICBA (~20 transports, 2016)
  - ICTP (~40 transports, 2011)
  - UNES (200+ transports, 2001 — limit to FI/CO/FM core)
  - IIEP (100+ transports, 2006 — limit similarly)

Migrated institutes (no creation trail):
  - IBE, UBO, UIL, UIS — annotated as 'pre-D01'
"""
import sys
import json
from collections import Counter, defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

sys.path.insert(0, 'Zagentexecution/mcp-backend-server-python')
from rfc_helpers import get_connection
g = get_connection('D01')

# Per-institute creation-cluster transport lists (from earlier discovery)
INSTITUTE_CLUSTERS = {
    'MGIE': [
        'D01K9A00FB','D01K9A011G','D01K9A0150','D01K9A01AG','D01K9A01AS',
        'D01K9A01B4','D01K9A01BE','D01K9A01JA','D01K9A01J2','D01K9A01IQ',
        'D01K9A01IE','D01K9A01KK','D01K9A01HS','D01K9A01IG','D01K9A01LE',
        'D01K9A020G',
    ],
    # For ICBA, ICTP, UNES, IIEP: discover dynamically below
}

# Golden step mapping — same as STEM script
GOLDEN_MAP = {
    'T001':[(1,'Create / Copy Company Code')], 'T001Q':[(1,'Create / Copy Company Code')],
    'T001A':[(2,'Add Group Currency')], 'T001B':[(69,'Open posting periods')],
    'T010O':[(69,'Open posting periods')], 'T010P':[(69,'Open posting periods')],
    'SKB1':[(4,'Extend GL per co code')], 'SKA1':[(4,'GL master')],
    'T012':[(16,'Create house banks')], 'T012D':[(16,'Create house banks')],
    'T012K':[(18,'Bank accounts')], 'T012T':[(16,'Create house banks')],
    'T018V':[(17,'Bank directory entries')], 'T028':[(17,'Bank directory')],'T028B':[(17,'Bank directory')],
    'T030H':[(14,'OBA1 / Exchange rate diff')],
    'T035D':[(28,'Cash management planning groups')],'T035U':[(28,'Cash management planning groups')],
    'T042':[(20,'FBZP — All co codes')],'T042B':[(23,'Payment methods per co code')],'T042E':[(22,'Payment methods per country')],
    'T042I':[(24,'Bank determination/ranking')],'T042T':[(21,'FBZP Paying co codes')],'T042V':[(25,'FBZP extras')],
    'T042A':[(25,'FBZP extras')],'T042D':[(25,'FBZP extras')],'T042P':[(25,'FBZP extras')],
    'T043G':[(35,'Tolerance groups')],'T043GT':[(35,'Tolerance groups')],
    'T043S':[(35,'Tolerance groups')],'T043ST':[(35,'Tolerance groups')],'T043T':[(35,'Tolerance groups')],
    'T169G':[(37,'MM/IV tolerances')],'T169P':[(37,'MM/IV tolerances')],'T169V':[(37,'MM/IV tolerances')],
    'T093B':[(44,'Asset class → GL determination')],'T093C':[(43,'Chart of depreciation assignment')],
    'T093D':[(46,'Asset transfer date')],'T093U':[(45,'Depreciation keys')],'T093V':[(43,'Chart of depreciation')],
    'AAACC_OBJ':[(43,'Account assignment objects')], 'ATPRA':[(46,'Asset transfer')], 'MARV':[(46,'MM/AA period')],
    'T047':[(42,'Dunning procedure')],
    'T882':[(55,'Consolidation unit')],
    'GB90':[(29,'GGB1 substitution')],'GB901':[(29,'GGB1 substitution')],'GB903':[(29,'GGB1 substitution')],
    'GB90T':[(29,'GGB1 substitution')],'GB92':[(29,'GGB1 substitution')],'GB921':[(29,'GGB1 substitution')],
    'GB922':[(29,'GGB1 substitution')],'GB92T':[(29,'GGB1 substitution')],'GB921T':[(29,'GGB1 substitution')],
    'GB93':[(31,'GGB0 validation')],
    'GGB1':[(29,'GGB1 substitution bundle')], 'GGB0':[(31,'GGB0 validation bundle')],
    'TKA00':[(9,'CO area')],'TKA01':[(9,'CO area')],'TKA02':[(9,'CO area assignment')],
    'TKA07':[(9,'CO area')],'TKA09':[(9,'CO area')],'TKT09':[(9,'CO area text')],
    'V_TKA3':[(10,'CO versions')],
    'CSKA':[(11,'Cost elements chart')], 'CSKB':[(11,'Cost elements per CO area')], 'CSKU':[(11,'Cost elements text')],
    'CSKS':[(47,'Cost centers')], 'CSKT':[(47,'Cost centers text')],
    'FM01':[(5,'FM area')], 'FMCI':[(8,'Commitment items')], 'FMCIT':[(8,'Commitment items text')],
    'FMFCTR':[(50,'Fund centers')], 'FMFCTRT':[(50,'Fund centers text')],
    'FMFINT':[(51,'Funds')],
    'FMFUNDTYPE':[(6,'Fund types')], 'V_FMFUNDTYPE':[(6,'Fund types maintenance')],
    'TPIR2':[(59,'Travel integration')], 'TPIR2T':[(59,'Travel integration text')],
    'T542A':[(58,'Personnel area subarea')], 'T500P':[(58,'Personnel area')],
    'T527O':[(80,'HCM customizing')], 'T7UNPAD_DS0P':[(80,'HR custom')],
    'T024':[(61,'Purchasing org')], 'T024E':[(61,'Purchasing org')], 'T024W':[(60,'Plant assignment')],
    'T001W':[(60,'Plant')], 'T001K':[(60,'Valuation area')],
    'T077K':[(0,'Vendor account groups')],
    'TCJ_C_JOURNALS':[(56,'Cash journal setup')],
    'TCJ_CJ_NAMES':[(56,'Cash journal setup')],
    'TCJ_PRINT':[(56,'Cash journal setup')],
    'TCJ_TRANSACTIONS':[(56,'Cash journal setup')],
    'TCJ_TRANS_NAMES':[(56,'Cash journal setup')],
    'TPIR2T':[(59,'Travel integration text')],
    'FMCFC0':[(54,'FM Carry-forward')],
    'FMCFC1':[(54,'FM Carry-forward')],
    'FMCFC2':[(54,'FM Carry-forward')],
    'T007V':[(38,'Tax codes')],
    'TAXIN':[(38,'Tax codes (India)')],
}

def explain(obj, name):
    name_u = (name or '').upper().strip()
    if obj == 'TABU':
        if name_u in ('T001','T001Q','T001A'): return 'Company code core record (T001 family)'
        if name_u in ('T012','T012D','T012K','T012T'): return 'House bank / bank account / bank text'
        if name_u in ('T018V','T018C','T018D','T028','T028B'): return 'Bank directory + bank type'
        if name_u.startswith('T042'): return 'FBZP payment program config'
        if name_u.startswith('T043') or name_u.startswith('T169'): return 'Tolerance group config'
        if name_u.startswith('T093') or name_u in ('AAACC_OBJ','ATPRA'): return 'Asset accounting config'
        if name_u == 'SKB1': return 'GL accounts at company-code level'
        if name_u in ('SKA1','SKAT'): return 'GL master / GL text'
        if name_u in ('T035D','T035U'): return 'Cash management planning groups'
        if name_u == 'T047': return 'Dunning procedure assignment'
        if name_u == 'T882': return 'Consolidation unit'
        if name_u.startswith('GB'): return 'Validation/Substitution rule (GGB0/GGB1 family)'
        if name_u.startswith('TKA'): return 'Controlling area config'
        if name_u in ('CSKA','CSKB','CSKS','CSKT','CSKU'): return 'CO master data'
        if name_u.startswith('FM') or 'FUND' in name_u: return 'Funds Management config'
        if name_u.startswith('TPIR') or name_u in ('T500P','T542A'): return 'HR / Travel personnel config'
        if name_u.startswith('TCJ'): return 'Cash journal config'
        if name_u in ('T010O','T010P','T001B'): return 'Posting periods config'
        if name_u.startswith('T024'): return 'Purchasing org config'
        if name_u.startswith('T001W') or name_u == 'T001K': return 'Plant / valuation area config'
        if name_u.startswith('T527') or name_u.startswith('T7UN'): return 'HR personnel customizing'
        if name_u in ('UNESGBB','UNESWRX'): return 'UNESCO custom MM account determination keys'
        if name_u.startswith('T007') or 'TAXIN' in name_u: return 'Tax code config'
        if name_u.startswith('T030'): return 'GL automatic account determination'
        return f'Customizing table {name}'
    if obj == 'TDAT':
        if name_u == 'GGB1': return 'GGB1 substitution bundle (carries GB90/921/922/T)'
        if name_u == 'GGB0': return 'GGB0 validation bundle'
        if name_u == 'ADDRESS' or 'ADDRESS' in name_u: return 'Address tables bundle (ADR2..ADR13, ADRC...)'
        if name_u == 'GRW_SET': return 'GS02 sets (BSET) bundle'
        if name_u.startswith('KE34'): return 'CO-PA operating concern data'
        return f'Transport data bundle {name}'
    if obj == 'VDAT': return f'View data {name}'
    if obj == 'CDAT': return f'Customer table data {name}'
    if obj == 'F01U': return f'Cross-company clearing pair {name}'
    if obj == 'CORR': return 'Sub-task release record'
    if obj == 'RELE': return 'Released sub-task record'
    if obj == 'ACGR': return f'Authorization role {name}'
    if obj.startswith('CL') or obj == 'CLAS': return f'ABAP class {name}'
    if obj == 'PROG': return f'ABAP program {name}'
    if obj == 'FUGR': return f'Function group {name}'
    if obj == 'TABL' or obj == 'DTEL': return f'DDIC structure {name}'
    return f'{obj} {name}'

# Discover other institutes' creation clusters (limit per institute)
def discover_cluster(institute, max_trs=30):
    # Get T001 transports for this BUKRS
    res = g.call('RFC_READ_TABLE', QUERY_TABLE='E071K',
                 FIELDS=[{'FIELDNAME':'TRKORR'},{'FIELDNAME':'TABKEY'}],
                 OPTIONS=[{'TEXT':f"OBJNAME = 'T001' AND TABKEY LIKE '%{institute}%'"}],
                 DELIMITER='|', ROWCOUNT=200)
    trs = set()
    for d in res.get('DATA', []):
        wa = d['WA'].split('|')
        tk = wa[1].strip() if len(wa) > 1 else ''
        if len(tk) >= 7 and tk[3:7] == institute:
            trs.add(wa[0].strip())
    if not trs:
        return []
    # Get earliest date and find all transports in ±90 days touching BUKRS-keyed entries
    md = {}
    for tr in trs:
        r = g.call('RFC_READ_TABLE', QUERY_TABLE='E070',
                   FIELDS=[{'FIELDNAME':'AS4DATE'}],
                   OPTIONS=[{'TEXT':f"TRKORR = '{tr}'"}], DELIMITER='|', ROWCOUNT=1)
        if r.get('DATA'):
            md[tr] = r['DATA'][0]['WA'].strip()
    if not md:
        return []
    earliest = min(md.values())
    # Get all transports with BUKRS-keyed entries near that date
    res = g.call('RFC_READ_TABLE', QUERY_TABLE='E071K',
                 FIELDS=[{'FIELDNAME':'TRKORR'}],
                 OPTIONS=[{'TEXT':f"TABKEY LIKE '%{institute}%'"}],
                 DELIMITER='|', ROWCOUNT=20000)
    candidate = set(d['WA'].strip() for d in res.get('DATA', []))
    # Get dates and filter to earliest year
    enriched = {}
    for tr in candidate:
        r = g.call('RFC_READ_TABLE', QUERY_TABLE='E070',
                   FIELDS=[{'FIELDNAME':'AS4DATE'},{'FIELDNAME':'AS4USER'}],
                   OPTIONS=[{'TEXT':f"TRKORR = '{tr}'"}], DELIMITER='|', ROWCOUNT=1)
        if r.get('DATA'):
            wa = r['DATA'][0]['WA'].split('|')
            d = wa[0].strip()
            if d and d[:4] == earliest[:4]:  # same year as creation
                enriched[tr] = (d, wa[1].strip())
    sorted_trs = sorted(enriched.items(), key=lambda x: x[1][0])
    return sorted_trs[:max_trs]

# Discover ICBA / ICTP / UNES / IIEP clusters
print('\n=== Discovering creation clusters ===')
for inst in ['ICBA','ICTP','UNES','IIEP']:
    cluster = discover_cluster(inst, max_trs=30)
    INSTITUTE_CLUSTERS[inst] = [t[0] for t in cluster]
    print(f'  {inst}: {len(INSTITUTE_CLUSTERS[inst])} transports in {cluster[0][1][0][:4] if cluster else "?"}')

# Pull E071/E071K per transport per institute
def pull_inventory(institute, trs):
    print(f'\n  Pulling {institute} inventory ({len(trs)} transports)...')
    rows = []
    for tr in trs:
        # E070 metadata
        r = g.call('RFC_READ_TABLE', QUERY_TABLE='E070',
                   FIELDS=[{'FIELDNAME':'AS4DATE'},{'FIELDNAME':'AS4USER'},
                           {'FIELDNAME':'TRFUNCTION'},{'FIELDNAME':'TRSTATUS'}],
                   OPTIONS=[{'TEXT':f"TRKORR = '{tr}'"}], DELIMITER='|', ROWCOUNT=1)
        if not r.get('DATA'): continue
        m = [x.strip() for x in r['DATA'][0]['WA'].split('|')]
        # E07T text
        rt = g.call('RFC_READ_TABLE', QUERY_TABLE='E07T',
                    FIELDS=[{'FIELDNAME':'AS4TEXT'}],
                    OPTIONS=[{'TEXT':f"TRKORR = '{tr}' AND LANGU = 'E'"}],
                    DELIMITER='|', ROWCOUNT=1)
        text = rt['DATA'][0]['WA'].strip() if rt.get('DATA') else ''
        # E071 objects
        ro = g.call('RFC_READ_TABLE', QUERY_TABLE='E071',
                    FIELDS=[{'FIELDNAME':'PGMID'},{'FIELDNAME':'OBJECT'},{'FIELDNAME':'OBJ_NAME'}],
                    OPTIONS=[{'TEXT':f"TRKORR = '{tr}'"}], DELIMITER='|', ROWCOUNT=500)
        objs = [d['WA'].split('|') for d in ro.get('DATA', [])]
        # E071K key counts
        rk = g.call('RFC_READ_TABLE', QUERY_TABLE='E071K',
                    FIELDS=[{'FIELDNAME':'OBJECT'},{'FIELDNAME':'OBJNAME'}],
                    OPTIONS=[{'TEXT':f"TRKORR = '{tr}'"}], DELIMITER='|', ROWCOUNT=20000)
        kcnt = Counter()
        for d in rk.get('DATA', []):
            wa = d['WA'].split('|')
            kcnt[(wa[0].strip(), wa[1].strip())] += 1
        for o in objs:
            pgmid = o[0].strip(); obj = o[1].strip(); name = o[2].strip()
            cnt = kcnt.get((obj, name), 0)
            plain = explain(obj, name)
            gold = GOLDEN_MAP.get(name.upper(), [])
            gold_str = '; '.join(f'#{g[0]} {g[1]}' for g in gold) if gold else ''
            rows.append({
                'transport': tr, 'date': m[0], 'user': m[1], 'fn': m[2], 'st': m[3],
                'tr_desc': text,
                'pgmid': pgmid, 'object': obj, 'obj_name': name,
                'keys': cnt, 'explanation': plain, 'maps_to_golden': gold_str,
            })
    return rows

# Build per-institute inventories
all_inventories = {}
for inst, trs in INSTITUTE_CLUSTERS.items():
    if trs:
        all_inventories[inst] = pull_inventory(inst, trs)

# Save JSON
with open('Zagentexecution/mcp-backend-server-python/all_institutes_inventory_2026-04-16.json','w') as f:
    json.dump(all_inventories, f, indent=2)

# Add tabs to Excel
SRC = 'Zagentexecution/Golden_live_matrix_2026-04-16_realigned_v3.xlsx'
DST = 'Zagentexecution/Golden_live_matrix_2026-04-16_realigned_v3.xlsx'
wb = load_workbook(SRC)

hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
hdr_font = Font(color='FFFFFF', bold=True)
tr_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')

for inst, inv in all_inventories.items():
    sheet_name = f'{inst} Inventory'
    if sheet_name in wb.sheetnames: del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    headers = ['Transport','Date','Owner','FN','ST','Desc','PGMID','OBJECT','OBJ_NAME','Keys','Explanation','Maps to Golden Step']
    ws.append(headers)
    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    cur_tr = None
    for r in inv:
        if r['transport'] != cur_tr:
            ws.append([f"== {r['transport']} == ({r['date']} {r['user']} fn={r['fn']} st={r['st']}) {r['tr_desc']}"] + ['']*11)
            for c in ws[ws.max_row]:
                c.fill = tr_fill; c.font = Font(bold=True)
            cur_tr = r['transport']
        ws.append([
            r['transport'], r['date'], r['user'], r['fn'], r['st'], r['tr_desc'],
            r['pgmid'], r['object'], r['obj_name'], r['keys'],
            r['explanation'], r['maps_to_golden'],
        ])
    widths = [16, 12, 12, 4, 4, 35, 6, 7, 22, 7, 50, 30]
    for col_letter, w in zip('ABCDEFGHIJKL', widths):
        ws.column_dimensions[col_letter].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:L{ws.max_row}'
    print(f'  Tab "{sheet_name}": {ws.max_row - 1} rows')

# Add a "Migrated Institutes" tab for IBE/UBO/UIL/UIS
sheet = 'Migrated Institutes'
if sheet in wb.sheetnames: del wb[sheet]
ws = wb.create_sheet(sheet)
ws.append(['Institute','Status','Note'])
for c in ws[1]:
    c.fill = hdr_fill; c.font = hdr_font
for inst in ['IBE','UBO','UIL','UIS']:
    ws.append([inst,'Pre-D01 / Migrated','No T001 creation transport found in D01 E071K. Config exists in live tables (T001, T012, etc.) but the creation trail is in a predecessor SAP system that was migrated.'])
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 100

wb.save(DST)
print(f'\nSaved: {DST}')
print(f'Total tabs in workbook: {len(wb.sheetnames)} = {wb.sheetnames}')
g.close()
