"""
Extend STEM 3-institutes Excel with live RFC data:
  1. Verify ICBA FM area mapping (T001/T001K/FM01)
  2. T036FT + T036FTT (Fund Type texts)
  3. CSKS + CSKT (Cost Centers + texts) for BUKRS in MGIE/IBE/ICBA
  4. Append to existing Excel + persist new tables to Gold DB
"""
import sys
import sqlite3
from pathlib import Path
from datetime import datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
from rfc_helpers import get_connection  # type: ignore

REPO = HERE.parents[1]
GOLD = REPO / "Zagentexecution" / "sap_data_extraction" / "sqlite" / "p01_gold_master_data.db"
INSTITUTES = ["MGIE", "IBE", "ICBA"]
EXCEL = REPO / "Zagentexecution" / f"STEM_3institutes_reference_{datetime.now().strftime('%Y-%m-%d')}.xlsx"


def rfc_table(guard, table, fields, where_list, label, rowcount=80000):
    rows = []
    for w in where_list:
        try:
            res = guard.call(
                "RFC_READ_TABLE",
                QUERY_TABLE=table,
                FIELDS=[{"FIELDNAME": f} for f in fields],
                OPTIONS=[{"TEXT": w}],
                DELIMITER="|",
                ROWCOUNT=rowcount,
            )
        except Exception as e:
            msg = str(e)
            if "TABLE_WITHOUT_DATA" in msg:
                continue
            print(f"  [{label}] WHERE='{w}' ERR: {msg[:120]}")
            continue
        for d in res.get("DATA", []):
            wa = d["WA"].split("|")
            rows.append({f: wa[i].strip() if i < len(wa) else "" for i, f in enumerate(fields)})
    return rows


def main():
    guard = get_connection("P01")
    print("Connected P01.")

    # ----- 1. Verify ICBA FM area -----
    print("\n[1] Verify ICBA FM area in live data")
    # T001K = company code -> FM area
    t001k = rfc_table(guard, "T001K", ["BUKRS", "FIKRS", "PERIV", "BWMOD"],
                      [f"BUKRS = '{i}'" for i in INSTITUTES], "T001K")
    print(f"  T001K rows: {len(t001k)}")
    for r in t001k:
        print(f"    BUKRS={r['BUKRS']} -> FIKRS={r['FIKRS']}")

    # FM01 = FM area master (table FM01 / TFKBVZ)
    # Try FM01 directly
    fm01 = rfc_table(guard, "FM01", ["FIKRS", "FYEAR_VARIANT", "WAERS"],
                     [f"FIKRS = '{i}'" for i in INSTITUTES], "FM01")
    print(f"  FM01 rows for institutes: {len(fm01)}")
    for r in fm01:
        print(f"    FM area exists: FIKRS={r['FIKRS']}")

    # All FM areas with text
    fm01t = rfc_table(guard, "FM01T", ["FIKRS", "SPRAS", "BEZEI"],
                      ["SPRAS = 'E'"], "FM01T", rowcount=100)
    print(f"  All FM areas (FM01T English): {len(fm01t)}")
    for r in fm01t:
        print(f"    FIKRS={r['FIKRS']} ({r['BEZEI']})")

    # Identify which FIKRS each institute uses
    icba_fikrs = next((r["FIKRS"] for r in t001k if r["BUKRS"] == "ICBA"), None)
    mgie_fikrs = next((r["FIKRS"] for r in t001k if r["BUKRS"] == "MGIE"), None)
    ibe_fikrs = next((r["FIKRS"] for r in t001k if r["BUKRS"] == "IBE"), None)
    print(f"\n  ICBA FM area = {icba_fikrs}")
    print(f"  MGIE FM area = {mgie_fikrs}")
    print(f"  IBE  FM area = {ibe_fikrs}")

    # ----- 2. T036FT + T036FTT -----
    print("\n[2] T036FT (Fund Types) + T036FTT (texts)")
    fikrs_to_query = sorted(set([f for f in [icba_fikrs, mgie_fikrs, ibe_fikrs] if f] + INSTITUTES))
    print(f"  Querying FIKRS: {fikrs_to_query}")
    t036ft = rfc_table(guard, "T036FT", ["FIKRS", "FONDSART"],
                       [f"FIKRS = '{f}'" for f in fikrs_to_query], "T036FT")
    print(f"  T036FT rows: {len(t036ft)}")
    t036ftt = rfc_table(guard, "T036FTT", ["FIKRS", "SPRAS", "FONDSART", "BEZEICH"],
                        [f"FIKRS = '{f}' AND SPRAS = 'E'" for f in fikrs_to_query], "T036FTT")
    text_map = {(r["FIKRS"], r["FONDSART"]): r["BEZEICH"] for r in t036ftt}
    for r in t036ft:
        r["BEZEICH"] = text_map.get((r["FIKRS"], r["FONDSART"]), "")
    print(f"  T036FTT rows (EN): {len(t036ftt)}")

    # ----- 3. CSKS + CSKT (Cost Centers) -----
    print("\n[3] CSKS (Cost Centers) for BUKRS in MGIE/IBE/ICBA")
    csks = rfc_table(guard, "CSKS",
                     ["KOKRS", "KOSTL", "DATBI", "DATAB", "BUKRS", "KOSAR", "VERAK", "ERSDA", "USNAM"],
                     [f"BUKRS = '{i}'" for i in INSTITUTES], "CSKS", rowcount=100000)
    print(f"  CSKS rows: {len(csks)}")
    by_bukrs = {}
    for r in csks:
        by_bukrs[r["BUKRS"]] = by_bukrs.get(r["BUKRS"], 0) + 1
    print(f"  by BUKRS: {by_bukrs}")

    # CSKT keyed by KOKRS+KOSTL+DATBI+SPRAS — pull per KOKRS encountered
    kokrs_set = sorted({r["KOKRS"] for r in csks if r["KOKRS"]})
    print(f"  Distinct KOKRS in result: {kokrs_set}")
    cskt = []
    for kk in kokrs_set:
        cskt.extend(rfc_table(guard, "CSKT",
                              ["KOKRS", "KOSTL", "DATBI", "KTEXT", "LTEXT"],
                              [f"KOKRS = '{kk}' AND SPRAS = 'E'"],
                              f"CSKT[{kk}]", rowcount=200000))
    print(f"  CSKT rows (EN): {len(cskt)}")
    txt_map = {(r["KOKRS"], r["KOSTL"], r["DATBI"]): r for r in cskt}
    for r in csks:
        t = txt_map.get((r["KOKRS"], r["KOSTL"], r["DATBI"]))
        r["KTEXT"] = t.get("KTEXT", "") if t else ""
        r["LTEXT"] = t.get("LTEXT", "") if t else ""

    guard.close()

    # ----- 4. Persist to Gold DB -----
    print("\n[4] Persist T036FT, CSKS, CSKT to Gold DB")
    con = sqlite3.connect(GOLD)
    cur = con.cursor()
    cur.execute("DROP TABLE IF EXISTS T036FT")
    cur.execute("CREATE TABLE T036FT (FIKRS TEXT, FONDSART TEXT, BEZEICH TEXT, PRIMARY KEY (FIKRS, FONDSART))")
    cur.executemany("INSERT OR REPLACE INTO T036FT VALUES (?,?,?)",
                    [(r["FIKRS"], r["FONDSART"], r.get("BEZEICH", "")) for r in t036ft])
    cur.execute("DROP TABLE IF EXISTS CSKS")
    cur.execute("""CREATE TABLE CSKS (KOKRS TEXT, KOSTL TEXT, DATBI TEXT, DATAB TEXT,
                   BUKRS TEXT, KOSAR TEXT, VERAK TEXT, ERSDA TEXT, USNAM TEXT,
                   KTEXT TEXT, LTEXT TEXT,
                   PRIMARY KEY (KOKRS, KOSTL, DATBI))""")
    cur.executemany("INSERT OR REPLACE INTO CSKS VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    [(r["KOKRS"], r["KOSTL"], r["DATBI"], r["DATAB"], r["BUKRS"],
                      r["KOSAR"], r["VERAK"], r["ERSDA"], r["USNAM"],
                      r.get("KTEXT", ""), r.get("LTEXT", "")) for r in csks])
    cur.execute("DROP TABLE IF EXISTS T001K")
    cur.execute("CREATE TABLE T001K (BUKRS TEXT PRIMARY KEY, FIKRS TEXT, PERIV TEXT, BWMOD TEXT)")
    cur.executemany("INSERT OR REPLACE INTO T001K VALUES (?,?,?,?)",
                    [(r["BUKRS"], r["FIKRS"], r.get("PERIV", ""), r.get("BWMOD", "")) for r in t001k])
    con.commit()
    con.close()
    print(f"  Gold DB updated: T036FT, CSKS, T001K")

    # ----- 5. Update Excel -----
    print("\n[5] Update Excel")
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(EXCEL)
    bold = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    note_fill = PatternFill("solid", fgColor="FFF2CC")

    # Add new sheet: ICBA_FMArea_Verification (insights)
    if "ICBA_Verification" in wb.sheetnames:
        del wb["ICBA_Verification"]
    s = wb.create_sheet("ICBA_Verification", 1)
    s.append([f"ICBA FM Area Verification — live P01 RFC at {datetime.now().isoformat()}"])
    s["A1"].font = Font(bold=True, size=12)
    s.append([])
    s.append(["T001K — Company Code -> FM Area mapping"])
    s.cell(row=s.max_row, column=1).font = Font(bold=True)
    s.append(["BUKRS", "FIKRS", "PERIV", "BWMOD"])
    for c in range(1, 5):
        s.cell(row=s.max_row, column=c).font = bold
        s.cell(row=s.max_row, column=c).fill = header_fill
    for r in t001k:
        s.append([r["BUKRS"], r["FIKRS"], r.get("PERIV", ""), r.get("BWMOD", "")])
    s.append([])
    s.append(["FM01T — All FM areas in P01 (English)"])
    s.cell(row=s.max_row, column=1).font = Font(bold=True)
    s.append(["FIKRS", "BEZEI"])
    for c in range(1, 3):
        s.cell(row=s.max_row, column=c).font = bold
        s.cell(row=s.max_row, column=c).fill = header_fill
    for r in fm01t:
        s.append([r["FIKRS"], r["BEZEI"]])
    s.append([])
    s.append(["Conclusion:"])
    s.cell(row=s.max_row, column=1).font = Font(bold=True)
    icba_t001k = next((r for r in t001k if r["BUKRS"] == "ICBA"), None)
    if icba_t001k and icba_t001k["FIKRS"] != "ICBA":
        s.append([f"  ICBA company code is mapped to FM area '{icba_t001k['FIKRS']}' (not its own area)."])
        s.append([f"  This explains why funds/fund_centers are 0 for FIKRS='ICBA' in Gold DB."])
        s.append([f"  ICBA reuses the funds and fund centers of FIKRS={icba_t001k['FIKRS']}."])
    elif icba_t001k:
        s.append([f"  ICBA has its own FM area but no funds extracted — investigate further."])
    else:
        s.append([f"  ICBA not found in T001K — anomaly, investigate."])
    for r in range(s.max_row - 2, s.max_row + 1):
        s.cell(row=r, column=1).fill = note_fill
    s.column_dimensions["A"].width = 80

    # Update FundTypes sheet — append a new sheet with live texts
    if "FundTypes_T036FT_live" in wb.sheetnames:
        del wb["FundTypes_T036FT_live"]
    s = wb.create_sheet("FundTypes_T036FT_live")
    s.append(["FIKRS", "FONDSART", "BEZEICH (Description)"])
    for c in range(1, 4):
        s.cell(row=1, column=c).font = bold
        s.cell(row=1, column=c).fill = header_fill
    for r in sorted(t036ft, key=lambda x: (x["FIKRS"], x["FONDSART"])):
        s.append([r["FIKRS"], r["FONDSART"], r.get("BEZEICH", "")])
    s.freeze_panes = "A2"
    s.column_dimensions["A"].width = 12
    s.column_dimensions["B"].width = 12
    s.column_dimensions["C"].width = 50

    # CostCenters live
    if "CostCenters_CSKS" in wb.sheetnames:
        del wb["CostCenters_CSKS"]
    s = wb.create_sheet("CostCenters_CSKS")
    cols = ["BUKRS", "KOKRS", "KOSTL", "KTEXT", "LTEXT", "KOSAR", "VERAK", "DATAB", "DATBI", "ERSDA", "USNAM"]
    s.append(cols)
    for c in range(1, len(cols) + 1):
        s.cell(row=1, column=c).font = bold
        s.cell(row=1, column=c).fill = header_fill
    for r in sorted(csks, key=lambda x: (x["BUKRS"], x["KOSTL"], x["DATBI"])):
        s.append([r.get(c, "") for c in cols])
    s.freeze_panes = "A2"
    for i, c in enumerate(cols, 1):
        s.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(c) + 4))

    # Update Summary
    s = wb["Summary"]
    # Find row "Cost Centers (CSKS) — RFC required" and update
    for row in s.iter_rows(min_row=1, max_row=s.max_row):
        if row[0].value and "Cost Centers" in str(row[0].value):
            row[0].value = "Cost Centers (CSKS)"
            for i, inst in enumerate(INSTITUTES, 2):
                row[i - 1].value = by_bukrs.get(inst, 0)
            row[4].value = sum(by_bukrs.values())
            break
    s.append([])
    s.append([f"Updated with RFC live data at {datetime.now().isoformat()}"])
    s.cell(row=s.max_row, column=1).font = Font(italic=True)

    wb.save(EXCEL)
    print(f"  Excel updated: {EXCEL}")
    print(f"\nSummary by BUKRS in CSKS: {by_bukrs}")
    print(f"T036FT total: {len(t036ft)}")


if __name__ == "__main__":
    main()
