"""
build_full_census.py — CENSO COMPLETO de la revaluacion FX de una sociedad, en Excel.

Sin filtros: TODAS las cuentas de balance del plan, esten dentro o fuera de una variante, tengan
o no divisa, bloqueadas o no. La idea es control total en una sola hoja: para cada cuenta se ve
donde se presenta, quien la revalua, de que moneda es, cuanto tiene, si hay algo que revaluar y
si esta bloqueada.

Nace de la sesion 102: las tablas anteriores filtraban por "tiene divisa hoy" y eso, leido como
censo, engana — de la posicion 1.1.2.1 solo aparecia una cuenta de ocho.

Columnas, en el orden pedido: VARIANTE -> posicion FS10 -> cuenta -> el resto. Ordena por
variante y dentro de cada una por posicion: es el eje por el que se decide y se actua.

SOLO LECTURA. La exposicion se pregunta cuenta a cuenta a BSIS (una tabla de 3,3 M de filas no
se deja leer entera), asi que la corrida tarda unos minutos.

Uso:
    python build_full_census.py [--bukrs UNES] [--out <fichero.xlsx>]
"""
import argparse
import collections
import os
import sys
from datetime import datetime, timezone

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", "..", ".."))
sys.path.insert(0, os.path.join(REPO, "Zagentexecution", "mcp-backend-server-python"))
sys.path.insert(0, os.path.join(REPO, "Zagentexecution", "quality_checks"))
from rfc_helpers import get_connection                                   # noqa: E402
from ob09_vs_variant_check import parse, rd, variant_selection, PROGRAM  # noqa: E402
from openpyxl import Workbook                                            # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill                 # noqa: E402
from openpyxl.utils import get_column_letter                             # noqa: E402

HDR = PatternFill("solid", fgColor="08305F")
RED = PatternFill("solid", fgColor="FADDE1")
AMB = PatternFill("solid", fgColor="FFECCC")
GRN = PatternFill("solid", fgColor="E3F5EC")
GRY = PatternFill("solid", fgColor="EEF1F5")


def num(s):
    s = (s or "").strip()
    return 0.0 if not s else (-float(s[:-1]) if s.endswith("-") else float(s))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--bukrs", default="UNES")
    ap.add_argument("--ktopl", default="UNES")
    ap.add_argument("--versn", default="FS10")
    ap.add_argument("--year", default="2026")
    ap.add_argument("--out", default=os.path.join(HERE, "fx_revaluation_full_census_UNES.xlsx"))
    a = ap.parse_args()

    c = get_connection("P01")
    try:
        variants = sorted({r["VARIANT"] for r in
                           rd(c, "VARID", ["VARIANT"], "REPORT = '%s'" % PROGRAM)
                           if r["VARIANT"].startswith(a.bukrs)})
        vtext = {r["VARIANT"]: r.get("VTEXT") for r in
                 rd(c, "VARIT", ["REPORT", "VARIANT", "VTEXT"], "REPORT = '%s'" % PROGRAM)}
        sel = {v: variant_selection(c, v) for v in variants}
        # el mecanismo (saldo / partidas abiertas) vive en los parametros, no en la seleccion
        sel_raw = {}
        for v in variants:
            try:
                sel_raw[v] = (c.call("RS_VARIANT_CONTENTS_RFC", REPORT=PROGRAM, VARIANT=v,
                                     VALUTAB=[]).get("VALUTAB") or [])
            except Exception:
                sel_raw[v] = []
        ska1 = {r["SAKNR"]: r for r in rd(c, "SKA1", ["SAKNR", "XBILK"],
                                          "KTOPL = '%s'" % a.ktopl)}
        skb1 = {r["SAKNR"]: r for r in rd(c, "SKB1", ["SAKNR", "XSPEB", "MITKZ", "WAERS",
                                                      "XOPVW"], "BUKRS = '%s'" % a.bukrs)}
        txt = {r["SAKNR"]: r["TXT50"] for r in
               rd(c, "SKAT", ["SAKNR", "TXT50"], "KTOPL = '%s' AND SPRAS = 'E'" % a.ktopl)}
        qt = {r["ERGSL"]: r["TXT45"] for r in
              rd(c, "FAGL_011QT", ["VERSN", "ERGSL", "TXT45", "SPRAS"],
                 "VERSN = '%s' AND SPRAS = 'E'" % a.versn)}
        zc = rd(c, "FAGL_011ZC", ["ERGSL", "VONKT", "BISKT"],
                "KTOPL = '%s' AND VERSN = '%s'" % (a.ktopl, a.versn))
        t030h = {r["HKONT"] for r in rd(c, "T030H", ["HKONT"],
                                        "KTOPL = '%s' AND CURTP = '10'" % a.ktopl)}
        local = next((r["WAERS"] for r in rd(c, "T001", ["BUKRS", "WAERS"],
                                             "BUKRS = '%s'" % a.bukrs)), "USD")
        iv = [(r["VONKT"].rjust(10, "0"), (r["BISKT"] or r["VONKT"]).rjust(10, "0"), r["ERGSL"])
              for r in zc]
        cols = ["RACCT", "HSLVT"] + ["HSL%02d" % i for i in range(1, 13)]
        bal = {}
        for r in rd(c, "GLT0", cols, "BUKRS = '%s' AND RYEAR = '%s'" % (a.bukrs, a.year)):
            v = sum(num(r.get(k)) for k in cols[1:])
            if abs(v) > 0.005:
                bal[r["RACCT"]] = bal.get(r["RACCT"], 0.0) + v

        todas = sorted(x for x in skb1 if ska1.get(x, {}).get("XBILK") == "X")
        activas = [x for x in todas if skb1[x].get("XSPEB") != "X"]

        def pos(x):
            h = [e for lo, hi, e in iv if lo <= x <= hi]
            return sorted(set(h))[0] if h else ""

        def field(x):
            return "AKONTO" if (skb1[x].get("MITKZ") or "") else "SKONTO"

        # COMO entra cada cuenta, no solo SI entra. Es la columna que explica el comportamiento:
        #   RANGE      -> la coge un intervalo. Se mantiene SOLA: una cuenta nueva dentro del
        #                 rango queda cubierta sin que nadie haga nada.
        #   INDIVIDUAL -> esta puesta a mano, una a una. Cada alta exige una accion humana, y
        #                 por eso es donde se acumulan los huecos.
        #   ALL-BUT    -> el campo solo tiene exclusiones: cubre TODO menos lo excluido. Tambien
        #                 se mantiene solo, pero al reves.
        member = collections.defaultdict(list)
        how = collections.defaultdict(set)
        excl = collections.defaultdict(list)
        for v in variants:
            for f in ("SKONTO", "AKONTO"):
                s = sel[v][f]
                pool = [x for x in todas if field(x) == f]
                ex = {x for x in pool
                      if x in s["exc"] or any(lo <= x <= hi for lo, hi in s["rex"])}
                if not s["inc"] and not s["rin"] and (s["exc"] or s["rex"]):
                    for x in set(pool) - ex:
                        member[x].append(v)
                        how[x].add("ALL-BUT")
                else:
                    for x in pool:
                        if x in ex:
                            continue
                        if x in s["inc"]:
                            member[x].append(v)
                            how[x].add("INDIVIDUAL")
                        elif any(lo <= x <= hi for lo, hi in s["rin"]):
                            member[x].append(v)
                            how[x].add("RANGE")
                for x in ex:
                    excl[x].append(v)

        print("midiendo divisa abierta cuenta a cuenta en %d cuentas activas..." % len(activas))
        fc = {}
        for i, x in enumerate(activas, 1):
            try:
                rows = parse(c.call("RFC_READ_TABLE", QUERY_TABLE="BSIS", DELIMITER="|",
                                    FIELDS=[{"FIELDNAME": y} for y in
                                            ("WAERS", "WRBTR", "DMBTR", "SHKZG")],
                                    OPTIONS=[{"TEXT": "BUKRS = '%s' AND HKONT = '%s'"
                                              % (a.bukrs, x)}], ROWCOUNT=0))
            except Exception:
                continue
            d, eq = {}, 0.0
            for r in rows:
                if r["WAERS"] and r["WAERS"] != local:
                    sg = -1 if r["SHKZG"] == "H" else 1
                    d[r["WAERS"]] = d.get(r["WAERS"], 0.0) + sg * num(r["WRBTR"])
                    eq += sg * num(r["DMBTR"])
            d = {k: v for k, v in d.items() if abs(v) > 0.005}
            if d:
                fc[x] = (d, eq)
            if i % 150 == 0:
                print("   %d/%d" % (i, len(activas)), flush=True)
    finally:
        c.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Full census"
    # La VARIANTE va primera y ordena: es el eje por el que se decide y se actua.
    head = ["F.05 variant", "FS10 position", "Position text", "G/L account", "Description",
            "Selected by", "Account currency", "Currency is fixed?", "Managed as", "Blocked?",
            "Balance %s (%s)" % (a.year, local), "Anything to revalue?",
            "FX currency", "Balance in FX currency", "All open FX",
            "FX in %s" % local, "OB09 (T030H)", "Verdict"]
    ws.append(head)
    for x in sorted(todas, key=lambda z: (", ".join(sorted(member[z])) or "zzz NONE",
                                          pos(z), z)):
        r = skb1[x]
        blocked = r.get("XSPEB") == "X"
        vs = ", ".join(sorted(member[x]))
        d, eq = fc.get(x, ({}, 0.0))
        fixed = bool(r.get("WAERS")) and r["WAERS"] != local
        if blocked:
            verdict = "Blocked - out of scope"
        elif d and not vs:
            verdict = "GAP - FX open, no variant"
        elif vs:
            verdict = "Revalued"
        elif x in t030h:
            verdict = "Latent - OB09 set, no variant, no FX today"
        elif excl[x]:
            verdict = "Excluded on purpose (%s)" % ", ".join(sorted(excl[x]))
        else:
            verdict = "Out - nothing to revalue"
        # La exposicion EN SU PROPIA MONEDA: la moneda dominante y su importe. El contravalor
        # va aparte, porque son dos preguntas distintas y mezclarlas confunde.
        cur, amt = "", None
        if d:
            cur = max(d, key=lambda k: abs(d[k]))
            amt = round(d[cur], 2)
            if len(d) > 1:
                cur += " (+%d)" % (len(d) - 1)
        ws.append([vs or "NONE", pos(x), qt.get(pos(x), ""), x, txt.get(x, ""),
                   " + ".join(sorted(how[x])) or ("EXCLUDED" if excl[x] else "not selected"),
                   r.get("WAERS") or "", "YES" if fixed else "no",
                   "Open items" if r.get("XOPVW") == "X" else "Balance",
                   "YES" if blocked else "", round(bal.get(x, 0.0), 2),
                   "YES" if d else "no", cur, amt,
                   " | ".join("%s %s" % (k, "{:,.0f}".format(v)) for k, v in sorted(d.items())),
                   round(eq, 2) if eq else None,
                   "YES" if x in t030h else "", verdict])
        i = ws.max_row
        fill = (GRY if blocked else RED if verdict.startswith("GAP") else
                AMB if verdict.startswith("Latent") else GRN if vs else None)
        for cn in range(1, len(head) + 1):
            if fill:
                ws.cell(row=i, column=cn).fill = fill
            if cn in (11, 14, 16):
                ws.cell(row=i, column=cn).number_format = "#,##0"
    for cell in ws[1]:
        cell.fill = HDR
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(head)), ws.max_row)
    for j, w in enumerate([17, 13, 26, 13, 40, 20, 12, 12, 11, 9, 17, 12,
                           13, 20, 34, 15, 10, 34], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    mech = {}
    for v in variants:
        dd = {x["SELNAME"]: (x.get("LOW") or "").strip() for x in (sel_raw.get(v) or [])}
        mech[v] = {"bal": dd.get("X_SALBEW") == "X",
                   "oi": any(dd.get(k) == "X" for k in ("X_GL", "X_AP", "X_AR"))}

    # ---- hoja resumen: VARIANTE x POSICION. Una posicion puede salir en varias filas,
    # una por variante, porque contiene cuentas con comportamientos distintos.
    ws2 = wb.create_sheet("Summary")
    # OJO con la columna de lo que falta: "Missing in the position" es un numero de la POSICION
    # y se repite en cada fila de variante; repetirlo sin decirlo hace pensar que a cada variante
    # le faltan esas. "Missing that FITS this variant" atribuye cada cuenta que falta a la
    # variante cuyo MECANISMO le encaja (SKB1-XOPVW): en 1.1.1.1 las 19 que faltan son todas de
    # partidas abiertas, asi que a UNES_UNBA le faltan CERO. Cuando dos variantes comparten
    # mecanismo, la cuenta aparece en las dos: la ambiguedad es real y se muestra.
    ws2.append(["F.05 variant", "Selected by", "FS10 position", "Position text",
                "Accounts this variant covers here", "Missing that FITS this variant",
                "Accounts in the position (active)", "Covered by all variants",
                "Missing in the position", "Blocked", "With FX open",
                "GAPS (FX + no variant)", "Balance (%s)" % local])
    filas = []
    for p in sorted({pos(x) for x in todas}):
        g = [x for x in todas if pos(x) == p]
        act_p = [x for x in g if skb1[x].get("XSPEB") != "X"]
        inv = [x for x in act_p if member[x]]
        out = [x for x in act_p if not member[x]]
        gaps = [x for x in out if x in fc]
        vs = sorted({v for x in inv for v in member[x]})
        base = [len(act_p), len(inv), len(out), len(g) - len(act_p),
                len([x for x in act_p if x in fc]), len(gaps),
                round(sum(bal.get(x, 0.0) for x in g), 2)]
        if not vs:
            filas.append(("zzz NO VARIANT", "-", p, qt.get(p, ""), 0, len(out),
                          base, gaps, inv, act_p))
            continue
        for v in vs:
            mine = [x for x in inv if v in member[x]]
            modo = " + ".join(sorted({m for x in mine for m in how[x]})) or "-"
            # Un rango con exclusiones individuales sigue siendo una REGLA, no un inventario:
            # los recortes son higiene (cuentas cerradas que caen dentro del bloque) y se
            # declaran uno a uno. Medido en UNES_OI_G/L: 3 rangos + 3 exclusiones, y las tres
            # excluidas son cuentas CLOSED, bloqueadas, sin saldo y sin divisa.
            # sobre TODAS las cuentas de la posicion, no solo las activas: los recortes de
            # UNES_OI_G/L son las tres cuentas CLOSED y bloqueadas, que ya salieron de act_p.
            n_carve = len([x for x in g if v in excl[x]])
            if n_carve and "RANGE" in modo:
                modo += " + %d carve-out" % n_carve + ("s" if n_carve > 1 else "")
            encajan = [x for x in out
                       if (mech[v]["oi"] if skb1[x].get("XOPVW") == "X" else mech[v]["bal"])]
            filas.append((v, modo, p, qt.get(p, ""), len(mine), len(encajan),
                          base, gaps, inv, act_p))
    for v, modo, p, ptxt, n_mine, n_fit, base, gaps, inv, act_p in sorted(filas):
        ws2.append([v.replace("zzz ", ""), modo, p, ptxt, n_mine, n_fit] + base)
        i = ws2.max_row
        for cn in range(1, 14):
            if gaps:
                ws2.cell(row=i, column=cn).fill = RED
            elif act_p and not inv:
                ws2.cell(row=i, column=cn).fill = AMB
            elif not base[2]:
                ws2.cell(row=i, column=cn).fill = GRN
            if cn == 13:
                ws2.cell(row=i, column=cn).number_format = "#,##0"
    for cell in ws2[1]:
        cell.fill = HDR
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = "A1:M%d" % ws2.max_row
    for j, w in enumerate([17, 20, 13, 30, 15, 17, 15, 13, 12, 9, 11, 13, 18], 1):
        ws2.column_dimensions[get_column_letter(j)].width = w

    # ---- hoja 3: que falta en cada posicion, y a que variante deberia ir
    ws3 = wb.create_sheet("Missing by position")
    ws3.append(["FS10 position", "Position text", "Variants working it", "Accounts",
                "Revalued", "Missing", "G/L account", "Description", "Managed as",
                "Account currency", "Blocked?", "Balance (%s)" % local,
                "Anything to revalue?", "Open FX", "OB09", "TARGET variant"])
    for p in sorted({pos(x) for x in todas}):
        g = [x for x in todas if pos(x) == p and skb1[x].get("XSPEB") != "X"]
        if not g:
            continue
        inv = [x for x in g if member[x]]
        out = [x for x in g if not member[x]]
        vs = sorted({v for x in inv for v in member[x]})
        for x in sorted(out):
            oi = skb1[x].get("XOPVW") == "X"
            target = [v for v in vs if (mech[v]["oi"] if oi else mech[v]["bal"])]
            if excl[x]:
                tgt = "EXCLUDED on purpose (%s)" % ", ".join(sorted(excl[x]))
            elif not vs:
                tgt = "NONE - no variant works this position"
            elif not target:
                tgt = "NO MATCH - %s work it but mechanism does not fit" % ", ".join(vs)
            else:
                tgt = ", ".join(target)
            d, eq = fc.get(x, ({}, 0.0))
            ws3.append([p, qt.get(p, ""), ", ".join(vs) or "NONE", len(g), len(inv), len(out),
                        x, txt.get(x, ""), "Open items" if oi else "Balance",
                        skb1[x].get("WAERS") or "", "", round(bal.get(x, 0.0), 2),
                        "YES" if d else "no",
                        " | ".join("%s %s" % (k, "{:,.0f}".format(val))
                                   for k, val in sorted(d.items())),
                        "YES" if x in t030h else "", tgt])
            i = ws3.max_row
            fill = (RED if d else AMB if x in t030h else
                    GRY if tgt.startswith("EXCLUDED") else None)
            for cn in range(1, 17):
                if fill:
                    ws3.cell(row=i, column=cn).fill = fill
                if cn == 12:
                    ws3.cell(row=i, column=cn).number_format = "#,##0"
    for cell in ws3[1]:
        cell.fill = HDR
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = "A1:P%d" % ws3.max_row
    for j, w in enumerate([13, 28, 26, 9, 9, 8, 13, 40, 11, 10, 9, 17, 12, 30, 8, 34], 1):
        ws3.column_dimensions[get_column_letter(j)].width = w

    # Summary primero, luego el detalle, luego lo que falta.
    wb.move_sheet("Summary", offset=-wb.sheetnames.index("Summary"))
    wb.save(a.out)
    gaps = [x for x in activas if x in fc and not member[x]]
    print("\nescrito: %s" % a.out)
    print("  %d cuentas de balance (%d activas, %d bloqueadas) · %d con divisa abierta · %d GAPS"
          % (len(todas), len(activas), len(todas) - len(activas), len(fc), len(gaps)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
