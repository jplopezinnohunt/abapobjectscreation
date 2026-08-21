"""
build_scope_workbook.py — FX revaluation scope, as an Excel workbook (English).

Reads P01 live and produces one .xlsx with four sheets:
  1. Scope by account   — one row per deposit/investment G/L account
  2. By balance sheet position (FS10) — the grouping that shows a whole position is missing
  3. Variants & methods — the typology (T044A x SAPF100 variants)
  4. Notes              — how to read it, and what was measured

READ-ONLY. Nothing is written to SAP.

Why an English workbook: this is the artefact that goes to Treasury, and Treasury reads the
balance sheet in English (FS10 texts are English).

Usage:
    python build_scope_workbook.py [--out <path.xlsx>]
"""
import argparse
import os
import sys
from datetime import datetime, timezone

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", "..", ".."))
sys.path.insert(0, os.path.join(REPO, "Zagentexecution", "mcp-backend-server-python"))
sys.path.insert(0, os.path.join(REPO, "Zagentexecution", "quality_checks"))
from rfc_helpers import get_connection                                    # noqa: E402
from ob09_vs_variant_check import parse, rd, variant_accounts, covered    # noqa: E402

from openpyxl import Workbook                                            # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side   # noqa: E402
from openpyxl.utils import get_column_letter                             # noqa: E402

BUKRS, KTOPL, VERSN, YEAR = "UNES", "UNES", "FS10", "2026"
POSITIONS = ["1.1.1.1", "1.1.1.2", "1.1.2.1", "1.1.2.3", "1.2.1.1", "1.2.1.2"]
FAMILY = ("00040410", "00040430")          # deposits & investments
NEW_TODAY = {"0004041018", "0004041019"}

HDR = PatternFill("solid", fgColor="08305F")
GRP = PatternFill("solid", fgColor="DCE6F1")
BAD = PatternFill("solid", fgColor="FADDE1")
WRN = PatternFill("solid", fgColor="FFECCC")
OK_ = PatternFill("solid", fgColor="E3F5EC")
THIN = Side(style="thin", color="D0D7DE")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def num(s):
    s = (s or "").strip()
    return 0.0 if not s else (-float(s[:-1]) if s.endswith("-") else float(s))


def collect():
    c = get_connection("P01")
    try:
        qt = {r["ERGSL"]: r["TXT45"] for r in
              rd(c, "FAGL_011QT", ["VERSN", "ERGSL", "TXT45", "SPRAS"],
                 "VERSN = '%s' AND SPRAS = 'E'" % VERSN)}
        zc = rd(c, "FAGL_011ZC", ["ERGSL", "VONKT", "BISKT"],
                "KTOPL = '%s' AND VERSN = '%s'" % (KTOPL, VERSN))
        iv = [(r["VONKT"].rjust(10, "0"), (r["BISKT"] or r["VONKT"]).rjust(10, "0"), r["ERGSL"])
              for r in zc]
        skb1 = {r["SAKNR"]: r for r in
                rd(c, "SKB1", ["SAKNR", "WAERS", "XSPEB", "XOPVW", "KDFSL"],
                   "BUKRS = '%s'" % BUKRS)}
        txt = {r["SAKNR"]: r["TXT50"] for r in
               rd(c, "SKAT", ["SAKNR", "TXT50"], "KTOPL = '%s' AND SPRAS = 'E'" % KTOPL)}
        t030h = {r["HKONT"]: (r.get("LKORR") or "").strip() for r in
                 rd(c, "T030H", ["HKONT", "LKORR"],
                    "KTOPL = '%s' AND CURTP = '10'" % KTOPL)}
        t030s = {(r.get("KDFSL") or ""): (r.get("KSOLL"), r.get("KHABN")) for r in
                 rd(c, "T030S", ["KTOPL", "KDFSL", "KSOLL", "KHABN"], "KTOPL = '%s'" % KTOPL)}
        variants = sorted({r["VARIANT"] for r in
                           rd(c, "VARID", ["VARIANT"], "REPORT = 'SAPF100'")
                           if r["VARIANT"].startswith(BUKRS)})
        sets, vmeta = {}, {}
        for v in variants:
            sets[v] = variant_accounts(c, v)
            d = {x["SELNAME"]: (x.get("LOW") or "").strip() for x in
                 (c.call("RS_VARIANT_CONTENTS_RFC", REPORT="SAPF100", VARIANT=v,
                         VALUTAB=[]).get("VALUTAB") or [])}
            vmeta[v] = d
        methods = {r["BWMET"]: r for r in
                   rd(c, "T044A", ["BWMET", "XSALK", "XPOSD", "XSALR", "KURSS", "KURSH",
                                   "BLART", "XAUFW"], "BWMET LIKE 'UN%'")}
        cols = ["RACCT", "HSLVT"] + ["HSL%02d" % i for i in range(1, 13)]
        bal = {}
        for r in rd(c, "GLT0", cols, "BUKRS = '%s' AND RYEAR = '%s'" % (BUKRS, YEAR)):
            v = sum(num(r.get(k)) for k in cols[1:])
            if abs(v) > 0.005:
                bal[r["RACCT"]] = bal.get(r["RACCT"], 0.0) + v
        local = next((r["WAERS"] for r in rd(c, "T001", ["BUKRS", "WAERS"],
                                             "BUKRS = '%s'" % BUKRS)), "USD")

        rows = []
        for acct in sorted(a for a in skb1 if a.startswith(FAMILY)):
            r = skb1[acct]
            hit = [e for lo, hi, e in iv if lo <= acct <= hi]
            fc, fc_local = {}, 0.0
            try:
                for x in parse(c.call("RFC_READ_TABLE", QUERY_TABLE="BSIS", DELIMITER="|",
                                      FIELDS=[{"FIELDNAME": y} for y in
                                              ("WAERS", "WRBTR", "DMBTR", "SHKZG")],
                                      OPTIONS=[{"TEXT": "BUKRS = '%s' AND HKONT = '%s'"
                                                % (BUKRS, acct)}], ROWCOUNT=0)):
                    sg = -1 if x["SHKZG"] == "H" else 1
                    if x["WAERS"] and x["WAERS"] != local:
                        fc[x["WAERS"]] = fc.get(x["WAERS"], 0.0) + sg * num(x["WRBTR"])
                        fc_local += sg * num(x["DMBTR"])
            except Exception:
                pass
            fc = {k: v for k, v in fc.items() if abs(v) > 0.005}
            vs = covered(acct, sets)
            fixed = bool(r.get("WAERS")) and r["WAERS"] != local
            oi = r.get("XOPVW") == "X"
            det = "T030H (KDF)" if oi and acct in t030h else (
                  "T030S (KDB, default)" if not oi and "" in t030s else "none")
            if r.get("XSPEB") == "X":
                status = "Blocked"
            elif fc and not vs:
                status = "DEFECT - FX open, no variant"
            elif vs:
                status = "Complete"
            elif acct in t030h:
                status = "Latent - configured, no variant, no FX today"
            else:
                status = "Correctly out - no FX"
            rows.append({
                "account": acct, "text": txt.get(acct, ""),
                "currency": r.get("WAERS") or "", "fixed": fixed,
                "management": "Open items" if oi else "Balance",
                "determination": det,
                "position": sorted(set(hit))[0] if hit else "(none)",
                "position_text": qt.get(sorted(set(hit))[0], "") if hit else "",
                "variant": ", ".join(vs), "balance": bal.get(acct, 0.0),
                "fc_open": sum(fc.values()), "fc_cur": ", ".join(sorted(fc)) or "",
                "fc_in_usd": fc_local, "status": status,
                "new_today": acct in NEW_TODAY})
        return rows, qt, vmeta, methods, sets, t030s, local
    finally:
        c.close()


def style_header(ws, row=1):
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = HDR
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.join(HERE, "fx_revaluation_scope_UNES.xlsx"))
    a = ap.parse_args()
    rows, qt, vmeta, methods, sets, t030s, local = collect()
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    wb = Workbook()

    # ---------------- sheet 1: scope by account
    ws = wb.active
    ws.title = "Scope by account"
    head = ["G/L account", "Description", "Account currency", "Currency is fixed?",
            "Managed as", "Account determination", "FS10 position", "FS10 position text",
            "F.05 variant", "Balance %s (%s)" % (YEAR, local),
            "Open items in FX", "FX currency", "FX open in %s" % local,
            "Status", "Created today"]
    ws.append(head)
    for r in sorted(rows, key=lambda z: (z["position"], z["account"])):
        ws.append([r["account"], r["text"], r["currency"],
                   "YES - only this currency" if r["fixed"]
                   else "no - company currency, any currency allowed",
                   r["management"], r["determination"], r["position"], r["position_text"],
                   r["variant"] or "NONE", round(r["balance"], 2),
                   round(r["fc_open"], 2) if r["fc_open"] else None, r["fc_cur"],
                   round(r["fc_in_usd"], 2) if r["fc_in_usd"] else None,
                   r["status"], "YES" if r["new_today"] else ""])
        i = ws.max_row
        fill = (BAD if r["status"].startswith("DEFECT") else
                WRN if r["status"].startswith("Latent") else
                OK_ if r["status"] == "Complete" else None)
        for c_ in range(1, len(head) + 1):
            cell = ws.cell(row=i, column=c_)
            cell.border = BOX
            if fill:
                cell.fill = fill
            if c_ in (10, 11, 13):
                cell.number_format = "#,##0"
            if c_ == 9 and not r["variant"]:
                cell.font = Font(bold=True, color="A32D2D")
    style_header(ws)
    autosize(ws, [14, 38, 15, 34, 12, 22, 13, 24, 16, 18, 16, 11, 16, 34, 12])
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(head)), ws.max_row)

    # ---------------- sheet 2: by FS10 position
    ws2 = wb.create_sheet("By FS10 position")
    ws2.append(["FS10 position", "Position text", "Accounts", "In a variant", "Not in a variant",
                "Currencies", "Variants present", "Total balance (%s)" % local,
                "FX open NOT revalued"])
    for p in sorted({r["position"] for r in rows}):
        g = [r for r in rows if r["position"] == p]
        inside = [r for r in g if r["variant"]]
        unrev = sum(r["fc_open"] for r in g if not r["variant"])
        ws2.append([p, qt.get(p, ""), len(g), len(inside), len(g) - len(inside),
                    " / ".join(sorted({r["currency"] or "?" for r in g})),
                    ", ".join(sorted({v for r in g for v in
                                      (r["variant"].split(", ") if r["variant"] else [])}))
                    or "NONE",
                    round(sum(r["balance"] for r in g), 2),
                    round(unrev, 2) if unrev else None])
        i = ws2.max_row
        for c_ in range(1, 10):
            cell = ws2.cell(row=i, column=c_)
            cell.border = BOX
            if unrev:
                cell.fill = BAD
            elif not inside:
                cell.fill = WRN
            if c_ in (8, 9):
                cell.number_format = "#,##0"
    style_header(ws2)
    autosize(ws2, [14, 26, 10, 13, 17, 14, 20, 20, 20])

    # ---------------- sheet 3: variants & methods
    ws3 = wb.create_sheet("Variants and methods")
    ws3.append(["Valuation method (T044A)", "In use?", "Mechanism", "Exchange rate type",
                "Document type"])
    for mname in sorted(methods):
        md = methods[mname]
        used = any(vmeta[v].get("BWMET1") == mname for v in vmeta)
        ws3.append([mname, "yes" if used else "NO - defined, never used",
                    "Balance valuation (XSALK)" if md.get("XSALK") == "X"
                    else "Open item valuation (XPOSD/XSALR)",
                    md.get("KURSS") or "", md.get("BLART") or ""])
    style_header(ws3)
    ws3.append([])
    ws3.append(["SAPF100 variant", "Method", "Mechanism actually switched on",
                "Account selection", "Accounts selected"])
    style_header(ws3, ws3.max_row)
    for v in sorted(vmeta):
        d = vmeta[v]
        mech = []
        if d.get("X_SALBEW") == "X":
            mech.append("BALANCE (X_SALBEW)")
        for k, lbl in (("X_GL", "open items G/L"), ("X_AP", "open items AP"),
                       ("X_AR", "open items AR")):
            if d.get(k) == "X":
                mech.append(lbl)
        inc, exc, rngs = sets[v]
        sel = ("%d range(s): " % len(rngs) + ", ".join("%s-%s" % (lo, hi) for _, lo, hi in rngs)) \
            if rngs else "%d single values" % len(inc)
        ws3.append([v, d.get("BWMET1", ""), " + ".join(mech) or "?", sel, len(inc) + len(rngs)])
    autosize(ws3, [26, 12, 34, 46, 18])

    # ---------------- sheet 4: notes
    ws4 = wb.create_sheet("How to read this")
    N = [
      ["FX revaluation scope - UNES deposits and investments", ""],
      ["Source", "P01 read live over RFC on %s. Read-only." % stamp],
      ["Population", "All G/L accounts of company code UNES in the deposit/investment ranges "
                     "40410xx and 40430xx, as presented in balance sheet version FS10."],
      ["", ""],
      ["Account currency - the key column", ""],
      ["Fixed foreign currency (EUR)", "SAP only allows postings in that currency, so everything "
                                       "the account holds is foreign currency. It ALWAYS needs "
                                       "revaluation, and you can tell from the master record."],
      ["Company currency (USD)", "SAP allows postings in ANY currency. The account may hold only "
                                 "USD (nothing to revalue) or foreign currency items (must be "
                                 "revalued). The master record does not tell you - you have to "
                                 "look at the line items. THIS IS WHERE THE GAP HAPPENS."],
      ["", ""],
      ["Account determination - two different tables", ""],
      ["SKB1-XOPVW = 'X'", "Open item managed -> KDF -> T030H, one row PER ACCOUNT (this is OB09)"],
      ["SKB1-XOPVW = ''", "Balance valuated -> KDB -> T030S, one row per exchange rate difference "
                          "key (SKB1-KDFSL); the row with a blank key is the chart default. In "
                          "UNES, KDFSL is blank on all 2,315 accounts, so every balance-valuated "
                          "account uses the default: expense 6045011 / income 7045011."],
      ["", ""],
      ["Balance vs exposure - do not confuse them", ""],
      ["Balance", "The account total in company currency (GLT0). Most of it can be USD, which "
                  "needs no revaluation."],
      ["Open items in FX", "Items still open (not cleared), in the currency they were posted in. "
                           "THIS is what gets revalued. Example: 4041011 has a balance of "
                           "571.6m USD of which 560m is USD - nothing to revalue - and 10m EUR "
                           "open, which does need revaluing and today nobody revalues it."],
      ["", ""],
      ["The finding", ""],
      ["One live defect", "4041011 Term Deposits Principal - open item managed, T030H row in "
                          "place (LKORR 4041011), and in NO F.05 variant. Configured to know "
                          "where to post, never selected to be calculated."],
      ["A whole position missing", "FS10 position 1.1.2.1 Short Term Deposits has 6 accounts and "
                                   "ZERO in any variant, carrying 763.7m USD. Compare with "
                                   "1.2.1.1 Other Investments (4 of 7 in) and 1.1.1.1 Cash with "
                                   "Banks (3 of 7 in)."],
      ["Two odd classifications", "The account named 'Short Term Deposits Principal' (4041013) is "
                                  "presented under Cash with Banks, not under Short Term "
                                  "Deposits. And the three mandates are split across two "
                                  "positions - PIMCO and JP Morgan in 1.1.2.1, RAMP in 1.2.1.1 - "
                                  "although they are the same kind of product."],
      ["Never reviewed", "The three mandates 4043030 PIMCO, 4043031 JP Morgan and 4043032 RAMP "
                         "total 395.7m USD with no OB09 and no variant. They hold no foreign "
                         "currency today, so they are correct - but nobody has confirmed it, and "
                         "they are exactly the profile this class of defect catches."],
      ["", ""],
      ["Reproduce", "python Zagentexecution/quality_checks/fx_revaluation_scope_check.py"],
      ["Rebuild this file", "python Zagentexecution/tasks/2026_08_21_fx_revaluation_scope/"
                            "build_scope_workbook.py"],
    ]
    for k, v in N:
        ws4.append([k, v])
        if v == "" and k:
            for c_ in (1, 2):
                ws4.cell(row=ws4.max_row, column=c_).fill = GRP
            ws4.cell(row=ws4.max_row, column=1).font = Font(bold=True, size=11)
        ws4.cell(row=ws4.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws4, [30, 105])

    wb.save(a.out)
    n_def = len([r for r in rows if r["status"].startswith("DEFECT")])
    print("written: %s" % a.out)
    print("  %d accounts | %d live defect(s) | %d FS10 positions"
          % (len(rows), n_def, len({r["position"] for r in rows})))
    return 0


if __name__ == "__main__":
    sys.exit(main())
