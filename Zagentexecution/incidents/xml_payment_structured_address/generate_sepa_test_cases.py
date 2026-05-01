"""Generate SEPA_CT_UNES DMEE Test Data cases — 3 tabs, 3 tables per tab.

Each tab represents a representative scenario from the top-15 coverage:
- Case 1: SEPA-UNES-FR-EUR-S-O (Standard vendor payment, scenario #1, 118,678 payments, 18.4% volume)
- Case 2: SEPA-UNES-FR-EUR-S-P (Payroll, scenario #2, 103,126 payments, 16% volume)
- Case 3: SEPA cross-border EUR to Spanish vendor (SEPA-zone non-FR, scenario coverage)

Each tab has 3 tables matching the DMEE Test Data screen layout:
  FPAYHX (Dbtr / UNESCO HQ — header per batch)
  FPAYH (Cdtr / Payee — data per payment)
  FPAYP (line items)

Output: sepa_ct_unes_test_cases.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT = Path(__file__).parent / "sepa_ct_unes_test_cases.xlsx"

# Styling
HEADER_FILL = PatternFill(start_color="1E3A4A", end_color="1E3A4A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TABLE_TITLE_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
TABLE_TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="15803D")
DATA_FONT = Font(name="Calibri", size=10)
NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="475569")
THIN = Side(border_style="thin", color="94A3B8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CASES = [
    {
        "name": "Case_1_Standard_Vendor",
        "title": "Case 1 — Standard SEPA vendor payment (FR vendor, EUR)",
        "subtitle": "Scenario #1: SEPA-UNES-FR-EUR-S-O · 118,678 payments · 18.4% in-scope volume · Doc cat=01 (FI vendor invoice)",
        "expected_xml_notes": [
            "Document/CstmrCdtTrfInitn/GrpHdr/InitgPty/Nm should reflect UNESCO (or SOG01 per current observed binding — flag for review)",
            "PmtInf/Dbtr/Nm = UNESCO · /Dbtr/PstlAdr/Ctry = FR · /Dbtr/PstlAdr/AdrLine = PARIS (V000 Hybrid: Ctry + 1 AdrLine)",
            "PmtInf/CdtTrfTxInf/Cdtr/Nm = ACME SUPPLIES SARL",
            "PmtInf/CdtTrfTxInf/Cdtr/PstlAdr — depends on Z-field bindings (ZLISO/ZPFST/ZPLOR); empty unless those are populated via expanded layout",
            "Amount: enter 1500.00 — observed 100x output (CtrlSum=150000.00) due to SEPA tree CV_RULE; expected behavior, not a bug for our purposes",
        ],
        "fpayhx": {
            "headers": ["Row", "PCd/city", "POBox/str.", "ISO Code", "Name", "Ctr", "City", "Issuer", "IBAN at our House Bank", "ISO", "SWIFT/BIC", "Reference number"],
            "rows": [
                [1, "75007", "7 PLACE DE FONTENOY", "FR", "UNESCO", "FR", "PARIS", "SOG01", "FR7630003000110000123456789", "EUR", "SOGEFRPP", "TEST-V000-CASE1"],
            ],
        },
        "fpayh": {
            "headers": ["Row", "No. FPAYHX", "Name of the payee", "SWIFT/BIC", "IBAN of the Payee", "Acct hold.", "Doc cat.", "Ref. Doc.", "Amount", "Due date"],
            "rows": [
                [1, 1, "ACME SUPPLIES SARL", "BNPAFRPPXXX", "FR1420041010050500013M02606", "ACME SUPPLIES SARL", "01", "5100012345", "1500.00", "01.04.2025"],
            ],
        },
        "fpayp": {
            "headers": ["Row", "No. FPAYH", "Name", "Ctr", "Name (2)"],
            "rows": [
                [1, 1, "Invoice 2025-001 services", "FR", "ACME SUPPLIES SARL"],
            ],
        },
        "z_fields_hint": [
            ["ZLISO", "FR", "Cdtr country code (ISO) — for V000 Cdtr/PstlAdr/Ctry"],
            ["ZPFST", "15 RUE DE LA PAIX", "Cdtr address line 1 — for V000 Cdtr/PstlAdr/AdrLine"],
            ["ZPLOR", "75002 PARIS", "Cdtr address line 2 — for V000 Cdtr/PstlAdr/AdrLine"],
            ["ZSTRA", "15 RUE DE LA PAIX", "Cdtr street (for V001 StrtNm if testing V001 instead)"],
        ],
    },
    {
        "name": "Case_2_Payroll",
        "title": "Case 2 — SEPA Payroll payment (employee, EUR)",
        "subtitle": "Scenario #2: SEPA-UNES-FR-EUR-S-P · 103,126 payments · 16% in-scope volume · Doc cat=03 (Payroll info)",
        "expected_xml_notes": [
            "Same Dbtr block as Case 1 (UNESCO HQ as payer)",
            "Doc cat=03 — Payroll category — this is what differentiates from Case 1",
            "Cdtr is the employee, not a vendor — different name pattern (PIERRE MARTIN, fictional)",
            "Ref. Doc. follows Payroll naming pattern (PR-YYYY-MM-EMPID rather than 51000xxx invoice number)",
            "Due date is end-of-month for payroll runs",
            "Amount field semantics same as Case 1 (100x scaling at output)",
        ],
        "fpayhx": {
            "headers": ["Row", "PCd/city", "POBox/str.", "ISO Code", "Name", "Ctr", "City", "Issuer", "IBAN at our House Bank", "ISO", "SWIFT/BIC", "Reference number"],
            "rows": [
                [1, "75007", "7 PLACE DE FONTENOY", "FR", "UNESCO", "FR", "PARIS", "SOG01", "FR7630003000110000123456789", "EUR", "SOGEFRPP", "TEST-V000-PAYROLL-001"],
            ],
        },
        "fpayh": {
            "headers": ["Row", "No. FPAYHX", "Name of the payee", "SWIFT/BIC", "IBAN of the Payee", "Acct hold.", "Doc cat.", "Ref. Doc.", "Amount", "Due date"],
            "rows": [
                [1, 1, "PIERRE MARTIN", "BNPAFRPPXXX", "FR7630002005500000987654321", "PIERRE MARTIN", "03", "PR-2025-04-EMP12345", "4500.00", "30.04.2025"],
            ],
        },
        "fpayp": {
            "headers": ["Row", "No. FPAYH", "Name", "Ctr", "Name (2)"],
            "rows": [
                [1, 1, "April 2025 Salary", "FR", "PIERRE MARTIN"],
            ],
        },
        "z_fields_hint": [
            ["ZLISO", "FR", "Employee country code (FR-resident employee)"],
            ["ZPFST", "12 RUE DE RIVOLI", "Employee residence street"],
            ["ZPLOR", "75001 PARIS", "Employee postal+city"],
            ["ZSTRA", "12 RUE DE RIVOLI", "Employee street (V001 StrtNm if testing V001)"],
        ],
    },
    {
        "name": "Case_3_Cross_Border_ES",
        "title": "Case 3 — SEPA cross-border vendor payment (ES vendor, EUR)",
        "subtitle": "SEPA-zone non-FR creditor · tests how tree handles non-FR Cdtr country · Doc cat=01 (FI vendor invoice)",
        "expected_xml_notes": [
            "Dbtr block identical to Cases 1+2 (UNESCO HQ)",
            "Cdtr country is ES (Spain) — verify Cdtr/PstlAdr/Ctry comes out as ES (not FR)",
            "BIC differs (BBVAESMMXXX = BBVA Spain) — verify Cdtr/CdtrAgt/FinInstnId/BIC reflects this",
            "IBAN starts ES (Spanish IBAN format) — different from Cases 1+2 FR IBAN",
            "This case validates that the SEPA tree correctly emits non-FR creditor data without Spain-specific bindings",
            "If output Cdtr country still says FR, it means the binding pulls from FPAYHX-Ctr (Dbtr) not from FPAYH or Z-fields — important V001 design implication",
        ],
        "fpayhx": {
            "headers": ["Row", "PCd/city", "POBox/str.", "ISO Code", "Name", "Ctr", "City", "Issuer", "IBAN at our House Bank", "ISO", "SWIFT/BIC", "Reference number"],
            "rows": [
                [1, "75007", "7 PLACE DE FONTENOY", "FR", "UNESCO", "FR", "PARIS", "SOG01", "FR7630003000110000123456789", "EUR", "SOGEFRPP", "TEST-V000-CROSSBORDER-001"],
            ],
        },
        "fpayh": {
            "headers": ["Row", "No. FPAYHX", "Name of the payee", "SWIFT/BIC", "IBAN of the Payee", "Acct hold.", "Doc cat.", "Ref. Doc.", "Amount", "Due date"],
            "rows": [
                [1, 1, "BARCELONA SUPPLIES S.L.", "BBVAESMMXXX", "ES9121000418450200051332", "BARCELONA SUPPLIES S.L.", "01", "5100099999", "2750.00", "15.04.2025"],
            ],
        },
        "fpayp": {
            "headers": ["Row", "No. FPAYH", "Name", "Ctr", "Name (2)"],
            "rows": [
                [1, 1, "Invoice 2025-Q2 IT services", "ES", "BARCELONA SUPPLIES S.L."],
            ],
        },
        "z_fields_hint": [
            ["ZLISO", "ES", "Cdtr country code — Spain (verify this overrides Dbtr Ctr=FR in Cdtr block)"],
            ["ZPFST", "GRAN VIA 235", "Cdtr street (Spanish vendor address line 1)"],
            ["ZPLOR", "08008 BARCELONA", "Cdtr postal+city"],
            ["ZSTRA", "GRAN VIA 235", "Cdtr street (V001 StrtNm)"],
        ],
    },
]


def write_table(ws, start_row, table_dict, title):
    """Write a single table (FPAYHX/FPAYH/FPAYP) starting at start_row. Returns next free row."""
    # Title
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.fill = TABLE_TITLE_FILL
    cell.font = TABLE_TITLE_FONT
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(table_dict["headers"]))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    start_row += 1

    # Header row
    for col_idx, hdr in enumerate(table_dict["headers"], start=1):
        c = ws.cell(row=start_row, column=col_idx, value=hdr)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = BORDER
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    start_row += 1

    # Data rows
    for row_data in table_dict["rows"]:
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=start_row, column=col_idx, value=val)
            c.font = DATA_FONT
            c.border = BORDER
            c.alignment = Alignment(horizontal="left", vertical="center")
        start_row += 1

    # Auto-size cols based on content
    for col_idx in range(1, len(table_dict["headers"]) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(table_dict["headers"][col_idx - 1])),
            *(len(str(r[col_idx - 1])) for r in table_dict["rows"]),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

    return start_row + 1  # blank line after


def write_case_tab(wb, case):
    """Write one case as one tab."""
    ws = wb.create_sheet(title=case["name"])

    # Tab title
    c = ws.cell(row=1, column=1, value=case["title"])
    c.font = Font(name="Calibri", size=14, bold=True, color="0F172A")
    ws.merge_cells("A1:L1")

    # Subtitle / scenario context
    c = ws.cell(row=2, column=1, value=case["subtitle"])
    c.font = Font(name="Calibri", size=10, italic=True, color="475569")
    ws.merge_cells("A2:L2")

    # Run order header
    c = ws.cell(row=3, column=1, value="🔢 EXECUTION ORDER FOR THIS CASE")
    c.fill = PatternFill(start_color="1E3A4A", end_color="1E3A4A", fill_type="solid")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.merge_cells("A3:L3")

    run_order = [
        "STEP 1 (V000 baseline) — tx DMEE → Tree type=PAYM + Format tree=/SEPA_CT_UNES → click Display",
        "STEP 2 — In the popup, click 'Active version' button → tree opens in V000 view",
        "STEP 3 — Menu: Tree → Test (or Test button on toolbar) → 'DMEE Test Data' screen opens with 3 empty grids",
        "STEP 4 — Copy the FPAYHX row from this Excel into the FPAYHX grid (skip the 'Row' column — it is just an Excel counter)",
        "STEP 5 — Copy the FPAYH row into the FPAYH grid (Doc cat. value is critical — use the value shown below)",
        "STEP 6 — Copy the FPAYP row into the FPAYP grid",
        "STEP 7 — (Optional but recommended) Expand FPAYHX layout: right-click header → Settings → add columns ZLISO, ZPFST, ZPLOR, ZSTRA — populate with values from the Z-FIELDS section below",
        "STEP 8 — Click Test/Run button (typically green-checkmark icon or Execute) → SAP renders V000 XML in a new screen",
        "STEP 9 — Save the V000 XML as 'case_<N>_V000.xml' (or copy the full XML text into a text file)",
        "STEP 10 (V001 comparison) — Press Back → return to DMEE Initial Screen",
        "STEP 11 — Display again → this time click 'Maintenance version' button → tree opens in V001 view",
        "STEP 12 — Tree → Test → enter the EXACT SAME values from this Excel into the 3 grids",
        "STEP 13 — Click Test/Run → SAP renders V001 XML",
        "STEP 14 — Save the V001 XML as 'case_<N>_V001.xml'",
        "STEP 15 — Diff case_<N>_V000.xml vs case_<N>_V001.xml — V001 should show ADDITIONAL <StrtNm>, <BldgNb>, <PstCd>, <TwnNm> tags inside <Cdtr><PstlAdr> and <Dbtr><PstlAdr>; if those new tags appear with the same input → V001 design empirically validated for this case",
    ]
    next_row = 4
    for step in run_order:
        c = ws.cell(row=next_row, column=1, value=step)
        c.font = Font(name="Calibri", size=10, color="1F2937")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=12)
        ws.row_dimensions[next_row].height = 22
        next_row += 1
    next_row += 1  # blank line

    # Original instructions block
    c = ws.cell(row=next_row, column=1, value=(
        "📋 NOTES — The 'Row' column in each table below is an Excel counter, do NOT enter it in SAP. "
        "Headers match SAP screen labels exactly. "
        "The 'No. FPAYHX' / 'No. FPAYH' columns are foreign keys linking the 3 tables — keep them = 1 for a single-payment test."
    ))
    c.font = NOTE_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=12)
    ws.row_dimensions[next_row].height = 50
    next_row += 2

    # FPAYHX table
    next_row = write_table(ws, next_row, case["fpayhx"], "🟢 FPAYHX — payment medium · formatting of payment data (Dbtr / UNESCO HQ side)")

    # FPAYH table
    next_row = write_table(ws, next_row, case["fpayh"], "🟢 FPAYH — payment medium · data for payment (Cdtr / Payee side)")

    # FPAYP table
    next_row = write_table(ws, next_row, case["fpayp"], "🟢 FPAYP — payment medium · data for paid items (line items)")

    # Z-fields hint section
    if case.get("z_fields_hint"):
        c = ws.cell(row=next_row, column=1, value="🔶 Z-FIELDS (hidden by default — expand FPAYHX layout to populate)")
        c.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        c.font = Font(name="Calibri", size=11, bold=True, color="B45309")
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
        next_row += 1

        z_headers = ["Z-Field", "Value", "Why this field", ""]
        for col_idx, hdr in enumerate(z_headers, start=1):
            c = ws.cell(row=next_row, column=col_idx, value=hdr)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.border = BORDER
        next_row += 1

        for z_row in case["z_fields_hint"]:
            for col_idx, val in enumerate(z_row, start=1):
                c = ws.cell(row=next_row, column=col_idx, value=val)
                c.font = DATA_FONT
                c.border = BORDER
            next_row += 1

        next_row += 1  # blank

    # Expected XML notes section
    c = ws.cell(row=next_row, column=1, value="📋 Expected XML output observations (V000)")
    c.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    c.font = Font(name="Calibri", size=11, bold=True, color="1E40AF")
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=12)
    next_row += 1

    for note in case["expected_xml_notes"]:
        c = ws.cell(row=next_row, column=1, value=f"  • {note}")
        c.font = DATA_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=12)
        ws.row_dimensions[next_row].height = 35
        next_row += 1

    next_row += 1  # blank

    # V001 testing reminder
    c = ws.cell(row=next_row, column=1, value=(
        "🔄 V001 testing — after running V000 with the values above, switch to Maintenance version "
        "(Display → Maintenance version) and re-run with the SAME inputs. The V001 output should additionally "
        "have <StrtNm>, <BldgNb>, <PstCd>, <TwnNm> tags inside <Cdtr><PstlAdr> and <Dbtr><PstlAdr> — fed from "
        "the Z-fields (ZSTRA/ZBSTR/etc.) byte-packed into FPAYHX-REF01..REF15 by Event 05. "
        "If V001 produces those structured tags with the same input → V001 design empirically validated against the SAP engine."
    ))
    c.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    c.font = Font(name="Calibri", size=10, color="15803D")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=12)
    ws.row_dimensions[next_row].height = 70


def write_overview_tab(wb):
    """First tab — overview + how to use."""
    ws = wb.create_sheet(title="README", index=0)

    rows = [
        ("UNESCO V001 — SEPA_CT_UNES — DMEE Test Data Cases", 16, True, "0F172A"),
        ("3 representative test cases for the /SEPA_CT_UNES tree, derived from the top-15 production scenarios coverage (claim 106).", 11, False, "475569"),
        ("", 10, False, "000000"),
        ("🔢 RECOMMENDED TEST EXECUTION ORDER", 13, True, "0F172A"),
        ("Run the 3 cases in this order. Each case = ~10-15 min (V000 + V001 + diff). Total time = ~30-45 min for the full SEPA tree validation.", 10, False, "1f2937"),
        ("", 10, False, "000000"),
        ("STEP 1 — Run Case 1 (Standard Vendor) FIRST", 11, True, "15803D"),
        ("   Reason: highest-volume scenario (118,678 payments / 18.4% volume) — establishes baseline V000 known-good output and the V001 reference for diff. If something fails here, no point running Cases 2-3.", 10, False, "475569"),
        ("   1.1 — Open tab 'Case_1_Standard_Vendor' in this workbook", 10, False, "1f2937"),
        ("   1.2 — Follow the 15-step EXECUTION ORDER section at the top of that tab", 10, False, "1f2937"),
        ("   1.3 — Save outputs as case_1_V000.xml + case_1_V001.xml + diff observations", 10, False, "1f2937"),
        ("", 10, False, "000000"),
        ("STEP 2 — Run Case 2 (Payroll) SECOND", 11, True, "15803D"),
        ("   Reason: second-highest-volume scenario (103,126 payments / 16% volume) and tests Doc cat=03 (Payroll) which differs from Case 1's Doc cat=01. Only point of variation: Doc cat. Validates the SEPA tree handles both invoice and payroll categories without format break.", 10, False, "475569"),
        ("   2.1 — Open tab 'Case_2_Payroll'", 10, False, "1f2937"),
        ("   2.2 — Follow the EXECUTION ORDER (same 15 steps, different values)", 10, False, "1f2937"),
        ("   2.3 — Save outputs as case_2_V000.xml + case_2_V001.xml", 10, False, "1f2937"),
        ("", 10, False, "000000"),
        ("STEP 3 — Run Case 3 (Cross-border ES) THIRD", 11, True, "15803D"),
        ("   Reason: tests SEPA-zone non-FR creditor — surfaces if the tree falls back to Dbtr-side country when Cdtr country is non-FR. This is the highest-information case for V001 design verification because it exercises Z-fields with non-FR data.", 10, False, "475569"),
        ("   3.1 — Open tab 'Case_3_Cross_Border_ES'", 10, False, "1f2937"),
        ("   3.2 — Follow the EXECUTION ORDER", 10, False, "1f2937"),
        ("   3.3 — Save outputs as case_3_V000.xml + case_3_V001.xml", 10, False, "1f2937"),
        ("", 10, False, "000000"),
        ("STEP 4 — Consolidate findings", 11, True, "15803D"),
        ("   Compile the 6 XML outputs (3 cases × 2 versions). Note any unexpected behaviors per case. Hand back to project agent for diff analysis vs Python simulator predictions and finding capture in the brain.", 10, False, "475569"),
        ("", 10, False, "000000"),
        ("📋 General workbook navigation", 12, True, "0F172A"),
        ("Each case tab is self-contained — has the EXECUTION ORDER (15 steps) at the top + the 3 tables to copy + the Z-fields hint + expected XML observations + V001 testing reminder.", 10, False, "475569"),
        ("Skip the 'Row' column from each table in the Excel — it is just a counter for the worksheet, not a SAP field.", 10, False, "475569"),
        ("Keep 'No. FPAYHX' = 1 and 'No. FPAYH' = 1 for a single-payment test.", 10, False, "475569"),
        ("For Cdtr address to render: expand the FPAYHX grid layout to add Z-fields (ZLISO, ZPFST, ZPLOR, ZSTRA, ZBSTR…) — see each case's Z-FIELDS section.", 10, False, "475569"),
        ("", 10, False, "000000"),
        ("📊 The 3 cases — coverage rationale", 12, True, "0F172A"),
        ("Case 1 — Standard SEPA vendor payment (FR vendor, EUR)", 11, True, "1f2937"),
        ("   Scenario #1 — 118,678 payments — 18.4% volume — Doc cat=01 (FI vendor invoice)", 10, False, "475569"),
        ("   The most common scenario. Validates baseline V000 SEPA output and provides the V001 reference for diff.", 10, False, "475569"),
        ("Case 2 — SEPA Payroll (employee, EUR)", 11, True, "1f2937"),
        ("   Scenario #2 — 103,126 payments — 16% volume — Doc cat=03 (Payroll info)", 10, False, "475569"),
        ("   Same tree, different document category. Validates that Doc cat=03 routes through the same SEPA chain without breaking format.", 10, False, "475569"),
        ("Case 3 — SEPA cross-border to ES vendor (Spanish vendor, EUR)", 11, True, "1f2937"),
        ("   SEPA-zone non-FR creditor — Doc cat=01 (FI vendor invoice)", 10, False, "475569"),
        ("   Tests that non-FR Cdtr country is correctly emitted (Cdtr/PstlAdr/Ctry=ES). Surfaces if the tree falls back to Dbtr country.", 10, False, "475569"),
        ("", 10, False, "000000"),
        ("⚠ Known SEPA tree behaviors (from initial V000 test session 2026-05-01)", 12, True, "B45309"),
        ("   • Amount × 100 at output: Amount=1500.00 in FPAYH produces CtrlSum=150000.00 in XML. Expected SEPA tree CV_RULE — production-correct.", 10, False, "475569"),
        ("   • InitgPty/Nm = SOG01 (HBKID): bound to FPAYHX-Issuer field, not to UNESCO name. Flag for review with N_MENARD if business intent is 'UNESCO' rather than HBKID.", 10, False, "475569"),
        ("   • UltmtCdtr/Nm bound to FPAYP-NAME1: when FPAYP 'Name' is populated, it appears as UltmtCdtr/Nm in the output XML.", 10, False, "475569"),
        ("   • Cdtr/PstlAdr binding requires Z-fields (ZLISO/ZPFST/ZPLOR): hidden by default in FPAYHX layout. Without them populated, Cdtr address comes out empty or falls back to Dbtr-side fields.", 10, False, "475569"),
        ("", 10, False, "000000"),
        ("🧪 What this enables (Layer 3 of the 4-layer testing methodology)", 12, True, "0F172A"),
        ("   This workbook delivers Layer 3 (DMEE Test Data with F110-history-replay-equivalent inputs) for the SEPA tree without requiring F110 execution.", 10, False, "475569"),
        ("   • Layer 1 (Python simulator) = already done — 794 cases × 3 schemas = 2,382 PASS (claim 105)", 10, False, "475569"),
        ("   • Layer 2 (DMEE Test Data hand-crafted) = these 3 cases", 10, False, "475569"),
        ("   • Layer 3 (DMEE Test Data F110-history replay) = next iteration — pull actual REGUH/REGUP rows for top-15 scenarios", 10, False, "475569"),
        ("   • Layer 4 (D01 F110 integration) = Phase 3 final integration test", 10, False, "475569"),
        ("   See companion BCM_StructuredAddressChange.html · Testing Simulation tab · 'Validation matrix alignment' for the full methodology.", 10, False, "475569"),
    ]

    for idx, (text, size, bold, color) in enumerate(rows, start=1):
        c = ws.cell(row=idx, column=1, value=text)
        c.font = Font(name="Calibri", size=size, bold=bold, color=color)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=12)
        if "•" in text or "   " in text[:5]:
            ws.row_dimensions[idx].height = 28
        else:
            ws.row_dimensions[idx].height = 22

    ws.column_dimensions["A"].width = 100


def main():
    wb = Workbook()
    # Remove default Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    write_overview_tab(wb)
    for case in CASES:
        write_case_tab(wb, case)

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")
    print(f"Tabs: README + {len(CASES)} cases ({', '.join(c['name'] for c in CASES)})")


if __name__ == "__main__":
    main()
